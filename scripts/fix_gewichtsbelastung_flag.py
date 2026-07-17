"""
Einmaliges Korrekturskript: aktualisiert den Flag-Text der Referenz "gewichtsbelastung" im
Werte-Sheet (Regelkorrektur Nutzer 2026-07-17).

Vorher hiess der Wert pro Ruestungslage "BE" (Behinderung) - das war falsch benannt. Korrekt:
eine Ruestungslage verursacht "RH" (Ruestungshinderlichkeit), nicht direkt eine Behinderung.
Die RH aller Lagen und Trefferzonen wird zu RHg aufsummiert; daraus ergibt sich erst die
tatsaechliche Ruestungsbehinderung:
    RBE = (RHg - ((Kon/5 + Staerke)/2 + sf_ruestungsmanoever)) / 6
(vorher faelschlich als drei separate Terme addiert statt (Kon/5+Staerke) zusammen zu halbieren).

Aufruf:
    python scripts/fix_gewichtsbelastung_flag.py "werte 0.7-claude.xlsx"

Legt vorher automatisch eine Sicherheitskopie an (siehe backup_werte.py).
"""
import sys
from pathlib import Path

import openpyxl

from backup_werte import backup

NEW_FLAG = (
    "Nutzerangabe 2026-07-16, per Regelkorrektur 2026-07-17 aktualisiert: GBE (Gewichtsbelastung) "
    "ist die Summe aus ZWEI Quellen. (1) Ruestungs-Komponente: eine Ruestungslage verursacht keine "
    "direkte Behinderung, sondern RH (Ruestungshinderlichkeit) - Einzel-Lagen-RH = "
    "RH-Basis(Ruestungsteil) + RH-Mod(Verarbeitung) + RH-Mod(Anpassung), je MAX(Lage; ...) - siehe "
    "Ruestung-Basis-Formel. RHg = Summe aller RH-Werte ueber alle 4 Koerperzonen. Daraus: "
    "RBE = (RHg - ((Kon/5 + Staerke)/2 + sf_ruestungsmanoever)) / 6 = die tatsaechliche "
    "Ruestungsbehinderung. (2) Ausruestungs-Komponente: Gewicht der getragenen/mitgefuehrten "
    "Gegenstaende (Preisliste-Gewichte), Schwellenwerte siehe gbe_unbelastet...gbe_ausgeruestet. "
    "BEIDE Komponenten sowie ihre Summe sind laut Nutzer Kampfmodul-Scope (setzt Laufzeit-"
    "Inventar/-Ruestungswahl voraus), nicht Teil der Charaktererstellung."
)


def main(path_str: str) -> None:
    path = Path(path_str)
    backup_path = backup(str(path))
    print(f"Sicherheitskopie angelegt: {backup_path}")

    wb = openpyxl.load_workbook(path, data_only=False)
    ws = wb["Werte"]
    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v:
            headers[v.strip()] = c

    for r in range(2, ws.max_row + 1):
        ref = ws.cell(row=r, column=headers["Referenz"]).value
        if ref == "gewichtsbelastung":
            ws.cell(row=r, column=headers["Flag"]).value = NEW_FLAG
            wb.save(path)
            print(f"Zeile {r}: Flag von 'gewichtsbelastung' aktualisiert.")
            print(f"Gespeichert: {path}")
            return

    raise SystemExit("Referenz 'gewichtsbelastung' nicht gefunden - nichts geaendert.")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print('Aufruf: python scripts/fix_gewichtsbelastung_flag.py "werte 0.7-claude.xlsx"')
        sys.exit(1)
    main(sys.argv[1])
