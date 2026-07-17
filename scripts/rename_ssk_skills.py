"""
Einmaliges Umbenennungsskript (Nutzer 2026-07-17): Sprache/Schrift/Kultur bekommen eine
gemeinsame Praefix-Familie "ssk_" (Sprache-Schrift-Kultur) statt im generischen "WHK"
aufzugehen, und wandern in eine eigene Kategorie "Sprache & Kultur". Jeder Sprachdialekt
bleibt eine eigene Faehigkeit (keine Zusammenlegung) - nur der Praefix aendert sich:

    whk_sprache_<dialekt>        -> ssk_sprache_<dialekt>   (23 Zeilen, unveraendert benannt)
    whk_kultur_<volk>            -> ssk_kultur_<volk>       (11 Zeilen, unveraendert benannt)
    whk_dalkinische_schrift      -> ssk_schrift_dalkini
    whk_elfische_schrift         -> ssk_schrift_elfen
    whk_gnomische_schrift        -> ssk_schrift_gnome
    whk_zwergische_schrift       -> ssk_schrift_zwerge

Kategorie aller 38 Zeilen: "WHK" -> "Sprache & Kultur".

Aufruf:
    python scripts/rename_ssk_skills.py "werte 0.7-claude.xlsx"

Legt vorher automatisch eine Sicherheitskopie an (siehe backup_werte.py).
"""
import sys
from pathlib import Path

import openpyxl

from backup_werte import backup

SCHRIFT_RENAMES = {
    "whk_dalkinische_schrift": "ssk_schrift_dalkini",
    "whk_elfische_schrift": "ssk_schrift_elfen",
    "whk_gnomische_schrift": "ssk_schrift_gnome",
    "whk_zwergische_schrift": "ssk_schrift_zwerge",
}

NEW_KATEGORIE = "Sprache & Kultur"


def main(path_str: str) -> None:
    path = Path(path_str)
    backup_path = backup(str(path))
    print(f"Sicherheitskopie angelegt: {backup_path}")

    wb = openpyxl.load_workbook(path, data_only=False)
    ws = wb["Werte"]
    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v:
            headers[v.strip()] = c

    renamed = 0
    for r in range(2, ws.max_row + 1):
        ref_cell = ws.cell(row=r, column=headers["Referenz"])
        ref = ref_cell.value
        if not ref:
            continue
        ref = str(ref).strip()

        new_ref = None
        if ref.startswith("whk_sprache_"):
            new_ref = "ssk_sprache_" + ref[len("whk_sprache_"):]
        elif ref.startswith("whk_kultur_"):
            new_ref = "ssk_kultur_" + ref[len("whk_kultur_"):]
        elif ref in SCHRIFT_RENAMES:
            new_ref = SCHRIFT_RENAMES[ref]

        if new_ref is None:
            continue

        ref_cell.value = new_ref
        ws.cell(row=r, column=headers["Kategorie"]).value = NEW_KATEGORIE
        flag_cell = ws.cell(row=r, column=headers["Flag"])
        note = (
            f"Umbenannt (Nutzer 2026-07-17): war '{ref}' in Kategorie 'WHK', jetzt "
            f"'{new_ref}' in eigener Kategorie '{NEW_KATEGORIE}' (ssk_-Praefix-Familie "
            "Sprache-Schrift-Kultur, jeder Dialekt bleibt eigene Faehigkeit)."
        )
        flag_cell.value = f"{flag_cell.value} | {note}" if flag_cell.value else note
        renamed += 1

    wb.save(path)
    print(f"{renamed} Zeilen umbenannt und nach '{NEW_KATEGORIE}' verschoben.")
    print(f"Gespeichert: {path}")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print('Aufruf: python scripts/rename_ssk_skills.py "werte 0.7-claude.xlsx"')
        sys.exit(1)
    main(sys.argv[1])
