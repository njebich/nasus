"""
Erzeugt die generierten TypeScript-Datenmodule aus einer werte-*.xlsx.

Betrifft:
  - src/data/rules.ts    (Sheet "Werte", alle Zeilen/Kategorien - die UI-Views
                           entscheiden selbst, welche Kategorien sie anzeigen)
  - src/data/lookups.ts  (Sheets "EP-Stufe-Kreis", "WHK-Spez-Kosten",
                           "Eigenschaften-Kosten" - von SVERWEIS referenziert)
  - src/data/equipment/preisliste.ts       (Sheet "Preisliste")
  - src/data/equipment/artefakte.ts        (Sheets "Artefakt-Basis" + "Artefakt-Kosten")
  - src/data/equipment/verfuegbarkeit.ts   (Sheet "Verfuegbarkeit-Modifikatoren", 2 Bloecke)
  - src/data/equipment/fernkampf.ts        (Sheets "Boegen", "Armbrust", "Pfeile", "Bolzen",
                                             "Feuerwaffen-Munition" sowie die separate Quelle
                                             "NN_Feuerwaffen_1.1.xlsx" - normalisierte Fernkampf-
                                             waffen/-Munition-Kataloge und Feuerwaffen-Ressourcen)
  - src/data/equipment/alchemika.ts        (Sheet "Alchemika", nur A1:P119 - Gifte/Heiltraenke/
                                             Kampftraenke/Parfum/Zustandstraenke-Katalog)
  - src/data/rules-jsonl/*.jsonl            (Sheet "Werte", nach Kategorie aufgesplittet - reine
                                             Lese-Projektion von rules.json fuer git-diff-
                                             Lesbarkeit und selektives Einlesen einzelner
                                             Kategorien, ohne die ganze rules.json laden zu
                                             muessen. Kein eigener Input, xlsx bleibt die Quelle.)

Aufruf:
    python scripts/generate_data_ts.py "werte 0.7-claude.xlsx"

Aendert NICHTS an der xlsx - reines Lesen. Ueberschreibt die generierten
.ts-Dateien komplett (sie sind reine Codegen-Artefakte, nicht von Hand pflegen).
"""
import sys
import json
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parent.parent
OUT_DATA_DIR = REPO_ROOT / "src" / "data"
OUT_EQUIPMENT_DIR = OUT_DATA_DIR / "equipment"
OUT_RULES_JSONL_DIR = OUT_DATA_DIR / "rules-jsonl"

UMLAUT_MAP = str.maketrans({"ä": "ae", "ö": "oe", "ü": "ue", "ß": "ss"})

RULE_COLUMNS = [
    ("Referenz", "referenz"),
    ("Kategorie", "kategorie"),
    ("Beschreibung", "beschreibung"),
    ("Abkürzung", "abkuerzung"),
    ("Info", "info"),
    ("Parent", "parent"),
    ("Art", "art"),
    ("Formel", "formelRaw"),
    ("Pool", "poolRaw"),
    ("Flag", "flag"),
    ("Grad", "grad"),
    ("Kosten", "kostenRaw"),
    ("Verfuegbarkeit", "verfuegbarkeit"),
    ("Mindest-TaW", "mindestTaw"),
    ("Eig-Bonus", "eigBonus"),
    ("Wirkung", "wirkung"),
]
REQUIRED_FIELDS = {"referenz", "kategorie", "art"}

LOOKUP_SHEETS = [
    "EP-Stufe-Kreis", "WHK-Spez-Kosten", "Eigenschaften-Kosten",
    "Sprachstufe-Kosten", "Kulturstufe-Kosten",
]


def cell_to_str(value):
    """Stringify a cell value the way it visually reads in Excel (no '16.0' for ints)."""
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def parse_numeric_or_sentinel(value):
    """Returns (numeric_value_or_None, raw_sentinel_text_or_None). Handles cells that are
    sometimes a clean number and sometimes a sentinel string like "------"/"nicht V."/"XXX"
    meaning "not available" - callers keep the raw text for traceability instead of just 0."""
    if value is None:
        return None, None
    if isinstance(value, (int, float)):
        return float(value), None
    text = str(value).strip()
    normalized = text.replace(",", ".").rstrip("D").rstrip("d").strip()
    try:
        return float(normalized), None
    except ValueError:
        return None, text


def parse_preis_dublonen_fz(raw):
    """Wie parse_numeric_or_sentinel, aber fuer die Boegen/Armbrust/Pfeile/Bolzen-Preisspalten:
    kennt zusaetzlich die Kleinwaehrung "FZ" (1000 FZ = 1 Dublone, Nutzer 2026-07-19/bereits
    2026-07-16 im Entwickeln-Sheet festgehalten) und ein optionales fuehrendes "+" bei den
    Spitzen-Modifikator-Zeilen (Preis-DELTA auf die Basis-Munition, nicht ihr eigener Preis).
    Returns (value_in_dublonen_or_None, is_delta)."""
    if raw is None:
        return None, False
    text = str(raw).strip()
    is_delta = text.startswith("+")
    if is_delta:
        text = text[1:].strip()
    normalized = text.replace(",", ".").replace(" ", "")
    if normalized.upper().endswith("FZ"):
        try:
            return float(normalized[:-2]) / 1000.0, is_delta
        except ValueError:
            return None, is_delta
    if normalized.upper().endswith("D"):
        try:
            return float(normalized[:-1]), is_delta
        except ValueError:
            return None, is_delta
    try:
        return float(normalized), is_delta
    except ValueError:
        return None, is_delta


def read_rules(wb):
    ws = wb["Werte"]
    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v:
            headers[v.strip()] = c

    missing_headers = [h for h, _ in RULE_COLUMNS if h not in headers]
    if missing_headers:
        raise SystemExit(f"Erwartete Spalten fehlen im 'Werte'-Sheet: {missing_headers}")

    rules = []
    warnings = []
    seen_referenz = {}
    for r in range(2, ws.max_row + 1):
        raw = {}
        for header, field in RULE_COLUMNS:
            raw[field] = cell_to_str(ws.cell(row=r, column=headers[header]).value)
        if all(v is None for v in raw.values()):
            continue

        entry = {k: v for k, v in raw.items() if v is not None}
        entry["sourceRow"] = r

        missing_required = REQUIRED_FIELDS - entry.keys()
        if missing_required:
            warnings.append(f"Zeile {r}: fehlende Pflichtfelder {missing_required} - uebersprungen")
            continue

        # Konvention (Nutzer 2026-07-18): eine mit "#" praefigierte Referenz markiert eine
        # bewusst deaktivierte Zeile (z.B. Duplikat einer bereits vorhandenen Talent-Wirkung) -
        # Zeile bleibt in der xlsx fuer Audit-Zwecke sichtbar, wird aber nicht importiert.
        if entry["referenz"].startswith("#"):
            warnings.append(f"Zeile {r}: Referenz '{entry['referenz']}' mit '#' auskommentiert - uebersprungen")
            continue

        ref_lower = entry["referenz"].lower()
        if ref_lower in seen_referenz:
            warnings.append(
                f"Zeile {r}: doppelte Referenz '{entry['referenz']}' "
                f"(zuerst Zeile {seen_referenz[ref_lower]}) - uebersprungen"
            )
            continue
        seen_referenz[ref_lower] = r

        rules.append(entry)

    return rules, warnings


def read_lookup_sheet(wb, sheet_name):
    ws = wb[sheet_name]
    headers = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v:
            headers.append((c, v.strip()))

    rows = []
    for r in range(2, ws.max_row + 1):
        row = {}
        for c, header in headers:
            val = cell_to_str(ws.cell(row=r, column=c).value)
            if val is not None:
                row[header] = val
        if row:
            rows.append(row)
    return rows


def read_generic_rows(wb, sheet_name, name_header, header_overrides=None):
    """Liest ein Sheet generisch: jede Zeile wird ein Dict aus {original Spaltenname: Wert}
    (nur nicht-leere Zellen) plus sourceRow und ein normalisiertes "name"-Feld (Wert der
    per name_header benannten Spalte). Passend fuer die vielen kleinen, strukturell simplen
    Modifikator-Tabellen (NK-Material, Ruestung-Verarbeitung, etc.) - spart 10+ fast-identische
    read_*-Funktionen. header_overrides={spalten_index: name} fuer Spalten ohne eigene
    Kopfzeile in der Quelle (z.B. NK-Material Spalte 11 "Volk" - bei NK-Fertigung/-Anpassung/
    -Schaftmaterial steht dort ein echter Header, bei NK-Material fehlt er in der Quelle)."""
    ws = wb[sheet_name]
    header_overrides = header_overrides or {}
    headers = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v:
            headers.append((c, v.strip()))
        elif c in header_overrides:
            headers.append((c, header_overrides[c]))

    rows = []
    for r in range(2, ws.max_row + 1):
        row = {}
        for c, header in headers:
            val = cell_to_str(ws.cell(row=r, column=c).value)
            if val is not None:
                row[header] = val
        if not row:
            continue
        row["sourceRow"] = r
        row["name"] = row.get(name_header, "")
        rows.append(row)
    return rows


def read_preisliste(wb):
    ws = wb["Preisliste"]
    rows = []
    for r in range(2, ws.max_row + 1):
        art = cell_to_str(ws.cell(row=r, column=1).value)
        name = cell_to_str(ws.cell(row=r, column=2).value)
        if not art and not name:
            continue
        anzahl_val, _ = parse_numeric_or_sentinel(ws.cell(row=r, column=3).value)
        einheit = cell_to_str(ws.cell(row=r, column=4).value)
        gewicht_val, gewicht_roh = parse_numeric_or_sentinel(ws.cell(row=r, column=5).value)
        preis_val, preis_roh = parse_numeric_or_sentinel(ws.cell(row=r, column=6).value)
        notiz = cell_to_str(ws.cell(row=r, column=7).value)
        whk = cell_to_str(ws.cell(row=r, column=8).value)
        spezialisierung = cell_to_str(ws.cell(row=r, column=9).value)

        entry = {"sourceRow": r, "art": art, "name": name}
        if anzahl_val is not None:
            entry["anzahl"] = anzahl_val
        if einheit:
            entry["einheit"] = einheit
        if gewicht_val is not None:
            entry["gewichtKg"] = gewicht_val
        elif gewicht_roh:
            entry["gewichtRoh"] = gewicht_roh
        entry["preisAvailable"] = preis_val is not None
        if preis_val is not None:
            entry["preisDublonen"] = preis_val
        elif preis_roh:
            entry["preisRoh"] = preis_roh
        if notiz:
            entry["notiz"] = notiz
        if whk:
            entry["whkCraftSkill"] = whk
        if spezialisierung:
            entry["spezialisierung"] = spezialisierung
        rows.append(entry)
    return rows


def read_artefakt_basis(wb):
    ws = wb["Artefakt-Basis"]
    cols = [
        ("Referenz", "referenz"), ("Name", "name"), ("Beschreibung", "beschreibung"),
        ("Zaubergrad", "zaubergrad"), ("Mana-Basis", "manaBasis"),
        ("Vorbereitungszeit-sec", "vorbereitungszeitSec"), ("Effektdauer-sec", "effektdauerSec"),
        ("Wirkungsdauer-Basis", "wirkungsdauerBasis"), ("Wirkungsdauer-Einheit", "wirkungsdauerEinheit"),
        ("Wirkung-Basis", "wirkungBasis"), ("Wirkung-Einheit", "wirkungEinheit"), ("Eigenschaft", "eigenschaft"),
    ]
    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v:
            headers[v.strip()] = c

    rows = []
    for r in range(2, ws.max_row + 1):
        entry = {}
        for header, field in cols:
            if header not in headers:
                continue
            entry[field] = cell_to_str(ws.cell(row=r, column=headers[header]).value)
        entry = {k: v for k, v in entry.items() if v is not None}
        if not entry.get("referenz"):
            continue
        entry["sourceRow"] = r
        rows.append(entry)
    return rows


def read_artefakt_kosten(wb):
    ws = wb["Artefakt-Kosten"]
    rows = []
    for r in range(2, ws.max_row + 1):
        referenz = cell_to_str(ws.cell(row=r, column=1).value)
        if not referenz:
            continue
        entry = {
            "sourceRow": r,
            "referenz": referenz,
            "name": cell_to_str(ws.cell(row=r, column=2).value),
            "grad": cell_to_str(ws.cell(row=r, column=3).value),
            "kostenEinmalig": cell_to_str(ws.cell(row=r, column=4).value),
            "verfuegbarkeitEinmalig": cell_to_str(ws.cell(row=r, column=5).value),
            "kostenPermanent": cell_to_str(ws.cell(row=r, column=6).value),
            "verfuegbarkeitPermanent": cell_to_str(ws.cell(row=r, column=7).value),
        }
        # Leere Zellen komplett weglassen statt "null" zu schreiben (konsistent mit den
        # anderen read_*-Funktionen) - TS-Seite kann sonst leicht null mit undefined verwechseln.
        rows.append({k: v for k, v in entry.items() if v is not None})
    return rows


def read_verfuegbarkeit_modifikatoren(wb):
    ws = wb["Verfuegbarkeit-Modifikatoren"]
    markt = []
    legende = []
    in_legende_block = False
    for r in range(2, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        b = ws.cell(row=r, column=2).value
        c = ws.cell(row=r, column=3).value
        if a is None and b is None and c is None:
            continue
        if a == "Grund-Verfuegbarkeit (Legende)":
            in_legende_block = True
            continue
        if in_legende_block and a == "Wert":
            continue  # zweite Kopfzeile des Legende-Blocks
        if in_legende_block:
            entry = {"wert": cell_to_str(a), "bedeutung": cell_to_str(b), "wuerfelwurf": cell_to_str(c)}
            legende.append({k: v for k, v in entry.items() if v is not None})
        else:
            entry = {"kategorie": cell_to_str(a), "auspraegung": cell_to_str(b), "faktor": cell_to_str(c)}
            markt.append({k: v for k, v in entry.items() if v is not None})
    return markt, legende


def ts_literal(value):
    """Render a Python value (str/int/dict/list) as a TS literal via JSON
    (JSON syntax is valid TS/JS object-literal syntax for our purposes)."""
    return json.dumps(value, ensure_ascii=False)


def write_rules_ts(rules, warnings):
    # Data goes into a plain .json file, not inlined as a TS array literal: tsc's structural
    # checking of ~1300+ distinct object literals against an interface with many optional
    # fields hits TS2590 ("union type too complex to represent"). A JSON import + single
    # `as unknown as RuleEntry[]` cast at the boundary sidesteps that entirely.
    json_rows = []
    for entry in rules:
        ordered = {k: entry[k] for k in [
            "referenz", "kategorie", "beschreibung", "abkuerzung", "info", "parent", "art",
            "formelRaw", "poolRaw", "flag", "grad", "kostenRaw", "verfuegbarkeit",
            "mindestTaw", "eigBonus", "wirkung", "sourceRow",
        ] if k in entry}
        json_rows.append(ordered)

    json_path = OUT_DATA_DIR / "rules.json"
    json_path.write_text(json.dumps(json_rows, ensure_ascii=False, indent=None), encoding="utf-8")

    lines = []
    lines.append("// GENERIERT von scripts/generate_data_ts.py - nicht von Hand bearbeiten.")
    lines.append("// Quelle: Sheet \"Werte\" (-> rules.json). Enthaelt ALLE Kategorien; welche")
    lines.append("// UI-Views welche Kategorien anzeigen, wird in src/views/*.ts entschieden.")
    lines.append("import rulesJson from './rules.json';")
    lines.append("")
    lines.append("export type Art = 'Wert' | 'Auswahl' | 'Formel' | 'Pool' | 'Fixwert' | 'Lookup';")
    lines.append("")
    lines.append("export interface RuleEntry {")
    lines.append("  referenz: string;")
    lines.append("  kategorie: string;")
    lines.append("  beschreibung?: string;")
    lines.append("  abkuerzung?: string;")
    lines.append("  info?: string;")
    lines.append("  parent?: string;")
    lines.append("  art: Art;")
    lines.append("  formelRaw?: string;")
    lines.append("  poolRaw?: string;")
    lines.append("  flag?: string;")
    lines.append("  grad?: string;")
    lines.append("  kostenRaw?: string;")
    lines.append("  verfuegbarkeit?: string;")
    lines.append("  mindestTaw?: string;")
    lines.append("  eigBonus?: string;")
    lines.append("  wirkung?: string;")
    lines.append("  sourceRow: number;")
    lines.append("}")
    lines.append("")
    lines.append("export const RULES = rulesJson as unknown as RuleEntry[];")
    lines.append("")

    if warnings:
        lines.append("// Codegen-Warnungen (siehe Konsolen-Ausgabe beim Generieren):")
        for w in warnings:
            lines.append(f"// - {w}")
        lines.append("")

    out_path = OUT_DATA_DIR / "rules.ts"
    out_path.write_text("\n".join(lines), encoding="utf-8")
    return out_path


def kategorie_slug(kategorie):
    """Dateiname-taugliches Kuerzel einer Kategorie (fuer rules-jsonl/*.jsonl)."""
    slug = kategorie.strip().lower().translate(UMLAUT_MAP)
    slug = "".join(ch if ch.isalnum() else "-" for ch in slug)
    while "--" in slug:
        slug = slug.replace("--", "-")
    return slug.strip("-") or "sonstige"


def write_rules_jsonl(rules):
    """Sharded Read-Projektion von rules.json: eine .jsonl-Datei pro Kategorie, eine Zeile pro
    Regel. Nicht die Quelle (die xlsx bleibt es) - nur damit git diffs pro Regel lesbar werden
    und man beim Nachschlagen einer Kategorie nicht die ganze rules.json laden muss."""
    if OUT_RULES_JSONL_DIR.exists():
        for old_file in OUT_RULES_JSONL_DIR.glob("*.jsonl"):
            old_file.unlink()
    OUT_RULES_JSONL_DIR.mkdir(parents=True, exist_ok=True)

    by_kategorie = {}
    for entry in rules:
        by_kategorie.setdefault(entry["kategorie"], []).append(entry)

    index = {}
    for kategorie, entries in by_kategorie.items():
        slug = kategorie_slug(kategorie)
        file_path = OUT_RULES_JSONL_DIR / f"{slug}.jsonl"
        with file_path.open("w", encoding="utf-8") as f:
            for entry in entries:
                f.write(json.dumps(entry, ensure_ascii=False, sort_keys=True))
                f.write("\n")
        index[kategorie] = {"file": file_path.name, "count": len(entries)}

    index_path = OUT_RULES_JSONL_DIR / "_index.json"
    index_path.write_text(json.dumps(index, ensure_ascii=False, indent=2, sort_keys=True), encoding="utf-8")
    return len(by_kategorie)


def write_lookups_ts(wb):
    tables = {}
    for sheet_name in LOOKUP_SHEETS:
        if sheet_name not in wb.sheetnames:
            print(f"WARNUNG: Lookup-Sheet '{sheet_name}' nicht gefunden - uebersprungen.")
            continue
        tables[sheet_name] = read_lookup_sheet(wb, sheet_name)

    json_path = OUT_DATA_DIR / "lookups.json"
    json_path.write_text(json.dumps(tables, ensure_ascii=False, indent=None), encoding="utf-8")

    lines = []
    lines.append("// GENERIERT von scripts/generate_data_ts.py - nicht von Hand bearbeiten.")
    lines.append("// Quelle: die per SVERWEIS referenzierten Lookup-Sheets (-> lookups.json).")
    lines.append("import lookupsJson from './lookups.json';")
    lines.append("")
    lines.append("export type LookupRow = Record<string, string>;")
    lines.append("")
    lines.append("export const LOOKUP_TABLES = lookupsJson as unknown as Record<string, LookupRow[]>;")
    lines.append("")

    out_path = OUT_DATA_DIR / "lookups.ts"
    out_path.write_text("\n".join(lines), encoding="utf-8")
    return out_path


def write_json_backed_module(out_dir, module_name, export_name, type_lines, ts_type_expr, data):
    """Writes {module_name}.json + {module_name}.ts (JSON import + typed export). Used for every
    generated data module - keeps data out of TS array literals (see write_rules_ts for why)."""
    json_path = out_dir / f"{module_name}.json"
    json_path.write_text(json.dumps(data, ensure_ascii=False, indent=None), encoding="utf-8")

    lines = ["// GENERIERT von scripts/generate_data_ts.py - nicht von Hand bearbeiten.", ""]
    lines.append(f"import {module_name}Json from './{module_name}.json';")
    lines.append("")
    lines.extend(type_lines)
    lines.append("")
    lines.append(f"export const {export_name} = {module_name}Json as unknown as {ts_type_expr};")
    lines.append("")

    out_path = out_dir / f"{module_name}.ts"
    out_path.write_text("\n".join(lines), encoding="utf-8")
    return out_path


def write_preisliste_ts(wb):
    rows = read_preisliste(wb)
    type_lines = [
        "export interface PreislisteRow {",
        "  sourceRow: number;",
        "  art?: string;",
        "  name?: string;",
        "  anzahl?: number;",
        "  einheit?: string;",
        "  gewichtKg?: number;",
        "  gewichtRoh?: string;",
        "  preisAvailable: boolean;",
        "  preisDublonen?: number;",
        "  preisRoh?: string;",
        "  notiz?: string;",
        "  whkCraftSkill?: string;",
        "  spezialisierung?: string;",
        "}",
    ]
    path = write_json_backed_module(OUT_EQUIPMENT_DIR, "preisliste", "PREISLISTE", type_lines, "PreislisteRow[]", rows)
    print(f"{path}: {len(rows)} Preisliste-Eintraege geschrieben.")


def write_artefakte_ts(wb):
    basis = read_artefakt_basis(wb)
    kosten = read_artefakt_kosten(wb)
    type_lines = [
        "export interface ArtefaktBasis {",
        "  sourceRow: number;",
        "  referenz: string;",
        "  name?: string;",
        "  beschreibung?: string;",
        "  zaubergrad?: string;",
        "  manaBasis?: string;",
        "  vorbereitungszeitSec?: string;",
        "  effektdauerSec?: string;",
        "  wirkungsdauerBasis?: string;",
        "  wirkungsdauerEinheit?: string;",
        "  wirkungBasis?: string;",
        "  wirkungEinheit?: string;",
        "  eigenschaft?: string;",
        "}",
        "export interface ArtefaktKosten {",
        "  sourceRow: number;",
        "  referenz: string;",
        "  name?: string;",
        "  grad?: string;",
        "  kostenEinmalig?: string;",
        "  verfuegbarkeitEinmalig?: string;",
        "  kostenPermanent?: string;",
        "  verfuegbarkeitPermanent?: string;",
        "}",
    ]
    data = {"basis": basis, "kosten": kosten}
    path = write_json_backed_module(
        OUT_EQUIPMENT_DIR, "artefakte", "ARTEFAKTE_RAW", type_lines,
        "{ basis: ArtefaktBasis[]; kosten: ArtefaktKosten[] }", data,
    )
    # Re-export as two flat consts for convenient importing (same underlying data/module).
    with open(OUT_EQUIPMENT_DIR / "artefakte.ts", "a", encoding="utf-8") as f:
        f.write("export const ARTEFAKT_BASIS = ARTEFAKTE_RAW.basis;\n")
        f.write("export const ARTEFAKT_KOSTEN = ARTEFAKTE_RAW.kosten;\n")
    print(f"{path}: {len(basis)} Artefakt-Basis, {len(kosten)} Artefakt-Kosten-Eintraege geschrieben.")


def write_verfuegbarkeit_ts(wb):
    markt, legende = read_verfuegbarkeit_modifikatoren(wb)
    type_lines = [
        "export interface MarktModifikatorRow { kategorie?: string; auspraegung?: string; faktor?: string; }",
        "export interface VerfuegbarkeitLegendeRow { wert?: string; bedeutung?: string; wuerfelwurf?: string; }",
    ]
    data = {"markt": markt, "legende": legende}
    path = write_json_backed_module(
        OUT_EQUIPMENT_DIR, "verfuegbarkeit", "VERFUEGBARKEIT_RAW", type_lines,
        "{ markt: MarktModifikatorRow[]; legende: VerfuegbarkeitLegendeRow[] }", data,
    )
    with open(OUT_EQUIPMENT_DIR / "verfuegbarkeit.ts", "a", encoding="utf-8") as f:
        f.write("export const MARKT_MODIFIKATOR = VERFUEGBARKEIT_RAW.markt;\n")
        f.write("export const VERFUEGBARKEIT_LEGENDE = VERFUEGBARKEIT_RAW.legende;\n")
    print(f"{path}: {len(markt)} Markt-Modifikator-, {len(legende)} Verfuegbarkeit-Legende-Eintraege geschrieben.")


GENERIC_ROW_TYPE_LINES = [
    "export type GenericRow = Record<string, string> & { sourceRow: number; name: string };",
]


def write_weapons_ts(wb):
    # NK-Waffen-Basis enthaelt sowohl Waffen als auch Schilde (Spezialisierung="Schild").
    # Nutzer 2026-07-18: "fang an damit, die nk-waffen inkl. herstellungs-modifikatoren zu
    # implementieren" - Material/Fertigung/Anpassung/Schaftmaterial jetzt mit aufgenommen
    # (analog zu Ruestung-Basis/-Verarbeitung/-Anpassung bzw. Schild-Material/-Fertigung/
    # -Bespannung), damit engine/weaponComposition.ts damit komponieren kann.
    basis = read_generic_rows(wb, "NK-Waffen-Basis", "Waffe")
    # Spalte 11 (Volk-Einschraenkung) hat in der Quelle keine eigene Kopfzeile - anders als bei
    # NK-Fertigung/-Anpassung/-Schaftmaterial, die dort "Volk" explizit benennen.
    material = read_generic_rows(wb, "NK-Material", "Waffe", header_overrides={11: "Volk"})
    fertigung = read_generic_rows(wb, "NK-Fertigung", "Fertigung")
    anpassung = read_generic_rows(wb, "NK-Anpassung", "Anpassung")
    schaftmaterial = read_generic_rows(wb, "NK-Schaftmaterial", "Verstärkung des Schafts")
    data = {
        "basis": basis, "material": material, "fertigung": fertigung,
        "anpassung": anpassung, "schaftmaterial": schaftmaterial,
    }
    path = write_json_backed_module(
        OUT_EQUIPMENT_DIR, "weapons", "WEAPONS_RAW", GENERIC_ROW_TYPE_LINES,
        "{ basis: GenericRow[]; material: GenericRow[]; fertigung: GenericRow[]; "
        "anpassung: GenericRow[]; schaftmaterial: GenericRow[] }", data,
    )
    with open(OUT_EQUIPMENT_DIR / "weapons.ts", "a", encoding="utf-8") as f:
        f.write("export const NK_WAFFEN_BASIS = WEAPONS_RAW.basis;\n")
        f.write("export const NK_MATERIAL = WEAPONS_RAW.material;\n")
        f.write("export const NK_FERTIGUNG = WEAPONS_RAW.fertigung;\n")
        f.write("export const NK_ANPASSUNG = WEAPONS_RAW.anpassung;\n")
        f.write("export const NK_SCHAFTMATERIAL = WEAPONS_RAW.schaftmaterial;\n")
    print(
        f"{path}: {len(basis)} NK-Waffen-Basis (Waffen+Schilde), {len(material)} NK-Material, "
        f"{len(fertigung)} NK-Fertigung, {len(anpassung)} NK-Anpassung, "
        f"{len(schaftmaterial)} NK-Schaftmaterial geschrieben."
    )


def write_armor_ts(wb):
    basis = read_generic_rows(wb, "Ruestung-Basis", "Ruestungsteil")
    verarbeitung = read_generic_rows(wb, "Ruestung-Verarbeitung", "Verarbeitung")
    anpassung = read_generic_rows(wb, "Ruestung-Anpassung", "Anpassung")
    data = {"basis": basis, "verarbeitung": verarbeitung, "anpassung": anpassung}
    path = write_json_backed_module(
        OUT_EQUIPMENT_DIR, "armor", "ARMOR_RAW", GENERIC_ROW_TYPE_LINES,
        "{ basis: GenericRow[]; verarbeitung: GenericRow[]; anpassung: GenericRow[] }", data,
    )
    with open(OUT_EQUIPMENT_DIR / "armor.ts", "a", encoding="utf-8") as f:
        f.write("export const RUESTUNG_BASIS = ARMOR_RAW.basis;\n")
        f.write("export const RUESTUNG_VERARBEITUNG = ARMOR_RAW.verarbeitung;\n")
        f.write("export const RUESTUNG_ANPASSUNG = ARMOR_RAW.anpassung;\n")
    print(
        f"{path}: {len(basis)} Ruestung-Basis, {len(verarbeitung)} Verarbeitung, "
        f"{len(anpassung)} Anpassung geschrieben."
    )


def write_shields_ts(wb):
    # Schild-Komposition (Nutzer 2026-07-17): "die haben auch Anpassung" - Material x Fertigung
    # x Bespannung, analog zu Ruestung-Basis x -Verarbeitung x -Anpassung. Schild-Verplatung
    # (nur Holzschilde) bewusst NICHT gelesen - Stub-Sheet ohne Werte (siehe Entwickeln-Sheet),
    # auf Nutzerwunsch vorerst weggelassen.
    material = read_generic_rows(wb, "Schild-Material", "Material")
    fertigung = read_generic_rows(wb, "Schild-Fertigung", "Fertigung")
    bespannung = read_generic_rows(wb, "Schild-Bespannung", "Bespannung")
    data = {"material": material, "fertigung": fertigung, "bespannung": bespannung}
    path = write_json_backed_module(
        OUT_EQUIPMENT_DIR, "shields", "SHIELDS_RAW", GENERIC_ROW_TYPE_LINES,
        "{ material: GenericRow[]; fertigung: GenericRow[]; bespannung: GenericRow[] }", data,
    )
    with open(OUT_EQUIPMENT_DIR / "shields.ts", "a", encoding="utf-8") as f:
        f.write("export const SCHILD_MATERIAL = SHIELDS_RAW.material;\n")
        f.write("export const SCHILD_FERTIGUNG = SHIELDS_RAW.fertigung;\n")
        f.write("export const SCHILD_BESPANNUNG = SHIELDS_RAW.bespannung;\n")
    print(
        f"{path}: {len(material)} Schild-Material, {len(fertigung)} Schild-Fertigung, "
        f"{len(bespannung)} Schild-Bespannung geschrieben."
    )


def enrich_fernkampf_rows(rows):
    """Ergaenzt jede Zeile um geparste numerische Felder (preisDublonen/preisIstDelta ueber
    parse_preis_dublonen_fz, verfuegbarkeitStufe als int) - roh-Spalten bleiben zusaetzlich als
    Strings erhalten (z.B. fuer Notiz/Kategorie/Anzeige der ungeparsten "1.W"-Wuerfelnotation)."""
    for row in rows:
        preis_raw = row.get("Preis")
        if preis_raw is not None:
            value, is_delta = parse_preis_dublonen_fz(preis_raw)
            if value is not None:
                row["preisDublonen"] = value
                if is_delta:
                    row["preisIstDelta"] = True
        verf_raw = row.get("Verfügbarkeit")
        if verf_raw is not None:
            try:
                row["verfuegbarkeitStufe"] = int(float(str(verf_raw).replace(",", ".")))
            except ValueError:
                pass
    return rows


def apply_basis_prices(rows, wb, sheet_name):
    """Ueberschreibt die veralteten Preistexte der Ausgabe-Sheets mit den numerischen Preisen
    aus dem jeweiligen *-Basis-Sheet. Die Basis-Sheets sind die gepflegte Preisquelle; dadurch
    werden auch fruehere Platzhalter wie ``nicht V.``, ``250D+`` oder ``1 Mio. D`` eindeutig.
    """
    basis_rows = read_generic_rows(wb, sheet_name, "Name")
    prices_by_name = {row["Name"]: row.get("Preis") for row in basis_rows}
    if len(prices_by_name) != len(basis_rows):
        raise SystemExit(f"Doppelte Namen im Sheet '{sheet_name}' verhindern die Preiszuordnung")

    for row in rows:
        name = row["Name"]
        if name not in prices_by_name:
            raise SystemExit(f"'{name}' fehlt im Preis-Quellsheet '{sheet_name}'")
        price, raw = parse_numeric_or_sentinel(prices_by_name[name])
        row.pop("preisDublonen", None)
        row.pop("preisIstDelta", None)
        if price is not None:
            row["Preis"] = f"{cell_to_str(price)}D"
            row["preisDublonen"] = price
        elif raw is not None:
            row["Preis"] = raw
    return rows


FERNKAMPF_ROW_TYPE_LINES = [
    "export type GenericRow = Record<string, string> & { sourceRow: number; name: string };",
    "export type FernkampfRow = Record<string, string> & { sourceRow: number; name: string; "
    "preisDublonen?: number; preisIstDelta?: boolean; verfuegbarkeitStufe?: number };",
]


def read_feuerwaffen(feuerwaffen_values):
    """Liest die Feuerwaffen-Vorlagen plus ihre VLOOKUP-Ressourcentabelle. Die Quelle bleibt
    eine eigene Arbeitsmappe, weil dort die lebenden Formeln und benannten Bereiche gepflegt
    werden; in die SPA gelangen ausschliesslich fertig serialisierte Konstanten."""
    waffen = read_generic_rows(feuerwaffen_values, "Waffen", "Name")
    for row in waffen:
        preis_raw = row.get("Preis")
        if preis_raw is not None:
            preis, _ = parse_numeric_or_sentinel(preis_raw)
            if preis is not None:
                row["preisDublonen"] = preis
        verf_raw = row.get("Verfuegbarkeit (1-7)")
        if verf_raw is not None:
            try:
                row["verfuegbarkeitStufe"] = int(float(str(verf_raw).replace(",", ".")))
            except ValueError:
                pass

    ressourcen = read_generic_rows(
        feuerwaffen_values, "Ressourcen-Waffen", "Name",
        # Die Lookup-Hilfstabellen besitzen in der Quelle absichtlich keine eigenen Header.
        header_overrides={36: "Wuerfelindex", 37: "Wuerfel", 41: "VerfRawAb", 42: "VerfStufe"},
    )
    wuerfel = []
    verfuegbarkeit = []
    for row in ressourcen:
        index, _ = parse_numeric_or_sentinel(row.get("Wuerfelindex"))
        if index is not None and row.get("Wuerfel"):
            wuerfel.append({"index": index, "wuerfel": row["Wuerfel"]})
        raw_ab, _ = parse_numeric_or_sentinel(row.get("VerfRawAb"))
        stufe, _ = parse_numeric_or_sentinel(row.get("VerfStufe"))
        if raw_ab is not None and stufe is not None:
            verfuegbarkeit.append({"rawAb": raw_ab, "stufe": int(stufe)})

    return waffen, ressourcen, wuerfel, verfuegbarkeit


def write_fernkampf_ts(wb, wb_values, feuerwaffen_values):
    # Boegen/Armbrust/Pfeile/Bolzen sind normalisierte, reine Wertetabellen ohne Formeln (per
    # one-off Build-Skripten aus den jeweiligen docx-Quellen erzeugt) - data_only spielt fuer
    # sie keine Rolle. Feuerwaffen-Munition dagegen hat lebende SUMPRODUCT/COUNTIFS-Formeln in
    # den Preis-1-Schuss-Spalten (Lookup gegen Munition-Feuerwaffen) - die muessen aus einer
    # data_only=True geladenen Workbook gelesen werden, sonst landet der Formel-Quelltext statt
    # des berechneten Preises im JSON.
    # preisDublonen/verfuegbarkeitStufe (Nutzer 2026-07-19: 1D=1000FZ, Kaufsperre ab Stufe 5 wie
    # bei Ruestung - siehe VERFUEGBARKEIT_SPERRE_AB in characterMutations.ts) werden nur fuer die
    # 4 Kataloge mit rohen Preis-/Verfuegbarkeit-Textspalten geparst, nicht fuer Feuerwaffen-
    # Munition (deren Preis-Spalte bereits eine fertig berechnete Formel-Zahl ist).
    boegen = apply_basis_prices(
        enrich_fernkampf_rows(read_generic_rows(wb, "Boegen", "Name")),
        wb, "Boegen-Basis",
    )
    armbrust = apply_basis_prices(
        enrich_fernkampf_rows(read_generic_rows(wb, "Armbrust", "Name")),
        wb, "Armbrust-Basis",
    )
    pfeile = enrich_fernkampf_rows(read_generic_rows(wb, "Pfeile", "Name"))
    bolzen = enrich_fernkampf_rows(read_generic_rows(wb, "Bolzen", "Name"))
    feuerwaffen_munition = read_generic_rows(wb_values, "Feuerwaffen-Munition", "Name")
    feuerwaffen, feuerwaffen_ressourcen, feuerwaffen_wuerfel, feuerwaffen_verfuegbarkeit = read_feuerwaffen(feuerwaffen_values)
    data = {
        "boegen": boegen, "armbrust": armbrust, "pfeile": pfeile,
        "bolzen": bolzen, "feuerwaffenMunition": feuerwaffen_munition,
        "feuerwaffen": feuerwaffen, "feuerwaffenRessourcen": feuerwaffen_ressourcen,
        "feuerwaffenWuerfel": feuerwaffen_wuerfel,
        "feuerwaffenVerfuegbarkeit": feuerwaffen_verfuegbarkeit,
    }
    path = write_json_backed_module(
        OUT_EQUIPMENT_DIR, "fernkampf", "FERNKAMPF_RAW", FERNKAMPF_ROW_TYPE_LINES,
        "{ boegen: FernkampfRow[]; armbrust: FernkampfRow[]; pfeile: FernkampfRow[]; "
        "bolzen: FernkampfRow[]; feuerwaffenMunition: GenericRow[]; feuerwaffen: FernkampfRow[]; "
        "feuerwaffenRessourcen: GenericRow[]; feuerwaffenWuerfel: { index: number; wuerfel: string }[]; "
        "feuerwaffenVerfuegbarkeit: { rawAb: number; stufe: number }[] }", data,
    )
    with open(OUT_EQUIPMENT_DIR / "fernkampf.ts", "a", encoding="utf-8") as f:
        f.write("export const BOEGEN = FERNKAMPF_RAW.boegen;\n")
        f.write("export const ARMBRUST = FERNKAMPF_RAW.armbrust;\n")
        f.write("export const PFEILE = FERNKAMPF_RAW.pfeile;\n")
        f.write("export const BOLZEN = FERNKAMPF_RAW.bolzen;\n")
        f.write("export const FEUERWAFFEN_MUNITION = FERNKAMPF_RAW.feuerwaffenMunition;\n")
        f.write("export const FEUERWAFFEN = FERNKAMPF_RAW.feuerwaffen;\n")
        f.write("export const FEUERWAFFEN_RESSOURCEN = FERNKAMPF_RAW.feuerwaffenRessourcen;\n")
        f.write("export const FEUERWAFFEN_WUERFEL = FERNKAMPF_RAW.feuerwaffenWuerfel;\n")
        f.write("export const FEUERWAFFEN_VERFUEGBARKEIT = FERNKAMPF_RAW.feuerwaffenVerfuegbarkeit;\n")
    print(
        f"{path}: {len(boegen)} Boegen, {len(armbrust)} Armbrust, {len(pfeile)} Pfeile, "
        f"{len(bolzen)} Bolzen, {len(feuerwaffen_munition)} Feuerwaffen-Munition, "
        f"{len(feuerwaffen)} Feuerwaffen und {len(feuerwaffen_ressourcen)} Ressourcen geschrieben."
    )


def write_voelker_maxima_ts(wb):
    # Eigenschaften-Min/Max je Volk (Nutzer 2026-07-17, werte 0.8): Erstellungs-Min/-Max gelten
    # waehrend der Charaktererstellung, "Max (ab Kreis 3)" (einheitlich 31) danach - siehe
    # state/characterMutations.ts setValue().
    ws = wb["Voelker-Maxima"]
    rows = []
    for r in range(2, ws.max_row + 1):
        volk = ws.cell(row=r, column=1).value
        eigenschaft = ws.cell(row=r, column=2).value
        if not volk or not eigenschaft:
            continue
        rows.append({
            "volk": str(volk).strip(),
            "eigenschaft": str(eigenschaft).strip(),
            "erstellungsMin": ws.cell(row=r, column=3).value,
            "erstellungsMax": ws.cell(row=r, column=4).value,
            "maxAbKreis3": ws.cell(row=r, column=5).value,
            "sourceRow": r,
        })
    type_lines = [
        "export interface VoelkerMaximaRow {",
        "  volk: string;",
        "  eigenschaft: string;",
        "  erstellungsMin: number;",
        "  erstellungsMax: number;",
        "  maxAbKreis3: number;",
        "  sourceRow: number;",
        "}",
    ]
    path = write_json_backed_module(
        OUT_DATA_DIR, "voelkerMaxima", "VOELKER_MAXIMA", type_lines, "VoelkerMaximaRow[]", rows,
    )
    print(f"{path}: {len(rows)} Voelker-Maxima-Eintraege geschrieben.")


def read_alchemika(wb_values):
    """Liest Sheet "Alchemika", nur A1:P119 (Nutzer 2026-07-19: Sheet wurde ersetzt, exakter
    Bereich vorgegeben). Feste Spaltenindizes statt Header-Namen (Nutzer hat die Zuordnung per
    Spaltenbuchstabe vorgegeben, nicht per Spaltenname): A=Kategorie, B=magisch/profan, C=Name,
    J=Wirkung, K=Legalitaet, M=Verfuegbarkeit, O=Kaufpreis(D). Alle anderen Spalten (Potenz,
    Dosierbar, Zufuhr, ZbS/ZbW, Wirkungsdauer, Char_Unterwelt_Mod, Zutaten_Preis, Rezept-
    Erschwerung) sind absichtlich NICHT erfasst - nicht Teil der angeforderten Ausruestungs-
    tabellen-Spalten (Nutzer: "Keine Spielregeln ableiten"). Spalte O (Kaufpreis_in_D) hat eine
    lebende Formel (=N*(K+L)) - braucht data_only=True wie Feuerwaffen-Munition, sonst landet
    Formel-Quelltext statt berechnetem Preis im JSON."""
    ws = wb_values["Alchemika"]
    rows = []
    for r in range(2, 120):
        kategorie = cell_to_str(ws.cell(row=r, column=1).value)
        magisch_raw = cell_to_str(ws.cell(row=r, column=2).value)
        name = cell_to_str(ws.cell(row=r, column=3).value)
        if not kategorie and not name:
            continue
        wirkung = cell_to_str(ws.cell(row=r, column=10).value)
        legalitaet_val, _ = parse_numeric_or_sentinel(ws.cell(row=r, column=11).value)
        verfuegbarkeit_val, _ = parse_numeric_or_sentinel(ws.cell(row=r, column=13).value)
        preis_val, preis_roh = parse_numeric_or_sentinel(ws.cell(row=r, column=15).value)

        magisch = (magisch_raw or "").strip().lower() == "magisch"
        beschreibung_teile = []
        if magisch:
            beschreibung_teile.append("Magischer Trank.")
        if legalitaet_val is not None and legalitaet_val > 2:
            beschreibung_teile.append("Verbotener Gegenstand")

        entry = {
            "sourceRow": r,
            "kategorie": kategorie or "",
            "name": name or "",
            "magisch": magisch,
            "wirkung": wirkung or "",
            "beschreibung": " ".join(beschreibung_teile),
        }
        if legalitaet_val is not None:
            entry["legalitaet"] = int(legalitaet_val)
        if verfuegbarkeit_val is not None:
            entry["verfuegbarkeitStufe"] = int(verfuegbarkeit_val)
        entry["preisAvailable"] = preis_val is not None
        if preis_val is not None:
            entry["preisDublonen"] = preis_val
        elif preis_roh:
            entry["preisRoh"] = preis_roh
        rows.append(entry)
    return rows


def write_alchemika_ts(wb_values):
    rows = read_alchemika(wb_values)
    type_lines = [
        "export interface AlchemikaRow {",
        "  sourceRow: number;",
        "  kategorie: string;",
        "  name: string;",
        "  magisch: boolean;",
        "  wirkung: string;",
        "  beschreibung: string;",
        "  legalitaet?: number;",
        "  verfuegbarkeitStufe?: number;",
        "  preisAvailable: boolean;",
        "  preisDublonen?: number;",
        "  preisRoh?: string;",
        "}",
    ]
    path = write_json_backed_module(OUT_EQUIPMENT_DIR, "alchemika", "ALCHEMIKA", type_lines, "AlchemikaRow[]", rows)
    print(f"{path}: {len(rows)} Alchemika-Eintraege geschrieben.")


def main(xlsx_path, feuerwaffen_xlsx_path=None):
    path = Path(xlsx_path)
    if not path.exists():
        raise SystemExit(f"Datei nicht gefunden: {path}")

    OUT_DATA_DIR.mkdir(parents=True, exist_ok=True)
    OUT_EQUIPMENT_DIR.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(path, data_only=False)
    wb_values = openpyxl.load_workbook(path, data_only=True)
    feuerwaffen_path = Path(feuerwaffen_xlsx_path) if feuerwaffen_xlsx_path else path.with_name("NN_Feuerwaffen_1.1.xlsx")
    if not feuerwaffen_path.exists():
        raise SystemExit(f"Feuerwaffen-Quelle nicht gefunden: {feuerwaffen_path}")
    feuerwaffen_values = openpyxl.load_workbook(feuerwaffen_path, data_only=True)

    rules, warnings = read_rules(wb)
    rules_path = write_rules_ts(rules, warnings)
    print(f"{rules_path}: {len(rules)} Regeln geschrieben.")
    if warnings:
        print(f"{len(warnings)} Warnung(en):")
        for w in warnings:
            print(f"  - {w}")

    n_kategorien = write_rules_jsonl(rules)
    print(f"{OUT_RULES_JSONL_DIR}: {len(rules)} Regeln in {n_kategorien} Kategorie-Dateien geschrieben.")

    lookups_path = write_lookups_ts(wb)
    print(f"{lookups_path}: {len(LOOKUP_SHEETS)} Lookup-Tabellen geschrieben.")

    write_preisliste_ts(wb)
    write_artefakte_ts(wb)
    write_verfuegbarkeit_ts(wb)
    write_weapons_ts(wb)
    write_armor_ts(wb)
    write_shields_ts(wb)
    write_fernkampf_ts(wb, wb_values, feuerwaffen_values)
    write_alchemika_ts(wb_values)
    write_voelker_maxima_ts(wb)


if __name__ == "__main__":
    if len(sys.argv) not in (2, 3):
        print('Aufruf: python scripts/generate_data_ts.py "werte 0.8-claude.xlsx" ["NN_Feuerwaffen_1.1.xlsx"]')
        sys.exit(1)
    main(sys.argv[1], sys.argv[2] if len(sys.argv) == 3 else None)
