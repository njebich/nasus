"""One-off: add Geweihte gate talents + Geweihte-WHK rows to the Werte sheet, and a
Meistermodul backlog note to the Entwickeln sheet. See project_nasus_chargen_app memory
"Geweihte-Tab" for the full feature. Run once, then recalc via Excel COM (openpyxl clears
cached formula values on save) before re-running generate_data_ts.py.

Usage: python scripts/add_geweihte_rows.py "werte 0.8-claude.xlsx"
"""
import sys
import openpyxl

WERTE_HEADERS = [
    'Referenz', 'Kategorie', 'Beschreibung', 'Abkürzung', 'Info', 'Parent', 'Art',
    'Formel', 'Pool', 'Flag', 'Grad', 'Kosten', 'Verfuegbarkeit', None,
    'Mindest-TaW', 'Eig-Bonus', 'Wirkung',
]

RELIGIONEN = ['Lloth', 'Khartazh', 'Nomna', 'Tepod', 'Isch']

TALENTE_ROWS = [
    {
        'Referenz': f'talente_geweihter_{name.lower()}_orthodox',
        'Kategorie': 'Talente',
        'Beschreibung': f'Geweihter von {name}, Orthodox',
        'Parent': 'Geweihte',
        'Art': 'Auswahl',
        'Kosten': 1,
        'Wirkung': f'Dieser Charakter ist ein Geweihter von {name}, Orthodox. Der Charakter muss Karma auf mindestens 1 steigern.',
    }
    for name in RELIGIONEN
]

WHK_ROWS = [
    {
        'Referenz': 'whk_geweihte_stossgebet',
        'Kategorie': 'WHK',
        'Beschreibung': 'Geweihte: Stoßgebet',
        'Art': 'Wert',
        'Kosten': 'WENN(wert=0;0;10+(wert-1)*wert/2)',
        'Info': "Bestimmt die Probe fuer Wunder vom Typ 'Stoß' (kurze, sofortige Anrufungen).",
    },
    {
        'Referenz': 'whk_geweihte_wunder',
        'Kategorie': 'WHK',
        'Beschreibung': 'Geweihte: Wunder',
        'Art': 'Wert',
        'Kosten': 'WENN(wert=0;0;10+(wert-1)*wert/2)',
        'Info': "Bestimmt die Probe fuer Wunder vom Typ 'Wunder' (laenger wirkende Anrufungen).",
    },
    {
        'Referenz': 'whk_geweihte_ritual',
        'Kategorie': 'WHK',
        'Beschreibung': 'Geweihte: Ritual',
        'Art': 'Wert',
        'Kosten': 'WENN(wert=0;0;10+(wert-1)*wert/2)',
        'Info': "Bestimmt die Probe fuer Wunder vom Typ 'Ritual' (aufwendige, lange Zeremonien).",
    },
]

ENTWICKELN_ROW = (
    'Geweihte/Klerus',
    'Geweihtengrad-Steigerung (Grad 2-7)',
    'Der Geweihte-Gate-Talent (talente_geweihter_<religion>_orthodox) setzt den Charakter '
    'direkt auf Geweihtengrad 1 ("Niederer"). Grad 2-7 ("Minderer" bis "Heiliger", Titel/KPP '
    'siehe Paladin-Geweihtensystem_Wundertabellen.txt) sind laut Nutzer (2026-07-22) nur vom '
    'Meister vergebbar - es gibt aktuell keine spielerseitige Steigerungsmechanik dafuer. '
    'Wird im geplanten Meistermodul umgesetzt (siehe project_nasus_multi_module_plan).',
    'Offen',
    '2026-07-22',
)


def col_index(header_name):
    return WERTE_HEADERS.index(header_name) + 1


def append_werte_row(ws, next_row, data: dict):
    for header, value in data.items():
        ws.cell(row=next_row, column=col_index(header), value=value)


def main():
    path = sys.argv[1] if len(sys.argv) > 1 else 'werte 0.8-claude.xlsx'
    wb = openpyxl.load_workbook(path, read_only=False, data_only=False)

    ws = wb['Werte']
    last_referenz_row = 1
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        if row[0].value is not None:
            last_referenz_row = row[0].row
    next_row = last_referenz_row + 1

    added = []
    for data in TALENTE_ROWS + WHK_ROWS:
        append_werte_row(ws, next_row, data)
        added.append((next_row, data['Referenz']))
        next_row += 1

    ws_ent = wb['Entwickeln']
    last_ent_row = 1
    for row in ws_ent.iter_rows(min_row=2, max_row=ws_ent.max_row):
        if row[0].value is not None:
            last_ent_row = row[0].row
    for i, value in enumerate(ENTWICKELN_ROW, start=1):
        ws_ent.cell(row=last_ent_row + 1, column=i, value=value)

    wb.save(path)
    print(f'Added {len(added)} Werte rows:')
    for row_num, referenz in added:
        print(f'  row {row_num}: {referenz}')
    print(f'Added Entwickeln row {last_ent_row + 1}')


if __name__ == '__main__':
    main()
