import re
import sys
from pathlib import Path


ROOT = Path(r"E:\Das Western Rollenspiel\LLM")
sys.path.insert(0, str(ROOT / ".python-deps"))

from openpyxl import load_workbook


WERTE = ROOT / "werte 0.8-claude.xlsx"
STAGES = (5, 10, 15, 20, 25, 30)
MAGIC_OPENERS = {
    5: "Der Charakter empfindet Unbehagen gegenüber aller Magie",
    10: "Der Charakter leidet unter Nervosität gegenüber aller Magie",
    15: "Der Charakter empfindet Furcht vor aller Magie",
    20: "Der Charakter hat Angst vor aller Magie",
    25: "Der Charakter gerät angesichts aller Magie in Panik",
    30: "Der Charakter leidet unter einer Phobie gegenüber aller Magie",
}


def magic_effect(stage):
    return (
        f"{MAGIC_OPENERS[stage]} und ist nicht in der Lage, irgendeine Form von Magie außer KI "
        f"zu wirken. Um gegen einen Magiewirker vorzugehen, muss er jede KR eine um {stage} "
        f"erschwerte Mut-Probe bestehen. Um Verzauberungen oder magische Heilung zuzulassen oder "
        f"Artefakte zu nutzen, muss er eine um {stage * 2} erschwerte Mut-Probe bestehen, gefolgt "
        f"von einer um {stage * 2} erschwerten Willenskraft-Probe."
    )


def main():
    workbook = load_workbook(WERTE, read_only=False, data_only=False)
    worksheet = workbook["Werte"]
    headers = {cell.value: cell.column for cell in worksheet[1] if cell.value}
    groups = {}

    for row in range(2, worksheet.max_row + 1):
        reference = str(worksheet.cell(row, headers["Referenz"]).value or "")
        match = re.fullmatch(r"vn_angst_(.+)_(5|10|15|20|25|30)", reference)
        if not match:
            continue
        theme, stage_text = match.groups()
        groups.setdefault(theme, {})[int(stage_text)] = row

    if len(groups) != 18:
        raise RuntimeError(f"Erwartet: 18 Angstgruppen; gefunden: {len(groups)}")

    created = []
    for theme, rows_by_stage in groups.items():
        if 25 not in rows_by_stage:
            source_row = rows_by_stage[20]
            row = worksheet.max_row + 1
            parent = worksheet.cell(source_row, headers["Parent"]).value
            rate = worksheet.cell(source_row, headers["Kosten"]).value / 20
            values = {
                "Referenz": f"vn_angst_{theme}_25",
                "Kategorie": "Vor- und Nachteile",
                "Beschreibung": f"Panik: {str(parent).removeprefix('Angst: ')}",
                "Info": f"Auswahlgruppe angst_{theme}; maximal 1 Stufe; Angstwert 25.",
                "Parent": parent,
                "Art": "Auswahl",
                "Flag": f"EXKLUSIV_ANGST=angst_{theme}",
                "Grad": 25,
                "Kosten": int(rate * 25),
                "Verfuegbarkeit": "S",
            }
            for header, value in values.items():
                worksheet.cell(row, headers[header], value)
            rows_by_stage[25] = row
            created.append((row, theme))

        if set(rows_by_stage) != set(STAGES):
            raise RuntimeError(f"Unvollständige Angstgruppe {theme}: {sorted(rows_by_stage)}")

        for stage, row in rows_by_stage.items():
            worksheet.cell(row, headers["Grad"], stage)
            worksheet.cell(row, headers["Info"], f"Auswahlgruppe angst_{theme}; maximal 1 Stufe; Angstwert {stage}.")
            worksheet.cell(row, headers["Verfuegbarkeit"], "S")
            if theme == "magie":
                worksheet.cell(row, headers["Wirkung"], magic_effect(stage))

    workbook.save(WERTE)
    print(f"fear_groups={len(groups)}")
    print(f"stage_25_created={len(created)}")
    print("magic_effects_written=6")


if __name__ == "__main__":
    main()
