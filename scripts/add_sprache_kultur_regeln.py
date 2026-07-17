"""
Einmaliges Ergaenzungsskript (mit Nutzer 2026-07-17/geklaert nach Lesen von
"NN Sprachen 0.11.docx"):

1. Neue Referenz "rerolls" (Formel: Reroll = Kreis).
2. Zwei neue Lookup-Sheets "Sprachstufe-Kosten" / "Kulturstufe-Kosten" (die im Regeldokument
   gefundene 5-Stufen-EP-Tabelle: Keine Kenntnis/Grundkenntnis/Gute Kenntnis/
   Muttersprache-bzw-Vaterland/Akademisches Niveau), ersetzen die bisherige generische
   WHK-Punktekurve bei allen whk_sprache_*-Zeilen.
3. Elf neue whk_kultur_<volk>-Zeilen (eine pro Volk aus dem Sheet "Voelker-Maxima"), mit der
   neuen Kulturstufe-Kosten-Tabelle. Existierte vorher gar nicht im Datensatz.

Schrift (whk_*_schrift) bleibt bewusst UNVERAENDERT: das Regeldokument beschreibt dafuer eine
pauschale Lerngebuehr (50 EP) statt der 5-Stufen-Tabelle, und lässt offen, wie die 3 "steigerbaren"
Schriften darueber hinaus weiter ansteigen - das ist ein offener Punkt, keine Annahme wert.

Aufruf:
    python scripts/add_sprache_kultur_regeln.py "werte 0.7-claude.xlsx"

Legt vorher automatisch eine Sicherheitskopie an (siehe backup_werte.py).
"""
import sys
from pathlib import Path

import openpyxl

from backup_werte import backup

# Sprachstufe (Sprache): Keine Kenntnis/Grundkenntnis/Gute Kenntnis/Muttersprache/Akademisches Niveau
SPRACHSTUFE_ROWS = [
    (0, 0, 0),
    (1, 15, 15),
    (2, 30, 15),
    (3, 50, 20),
    (4, 75, 25),
]
# Kulturstufe (Kultur): Keine Kenntnis/Grundkenntnis/Gute Kenntnis/Vaterland/Akademisches Niveau
KULTURSTUFE_ROWS = [
    (0, 0, 0),
    (1, 10, 10),
    (2, 25, 15),
    (3, 40, 15),
    (4, 55, 15),
]

# Volk-Liste aus Sheet "Voelker-Maxima" (Spalte A, 11 eindeutige Werte).
VOELKER = [
    "Dalkini", "Draw", "Elfen", "Gnome", "Goblins", "Indianer",
    "Katzen", "Orks", "Trolle", "Zentauren", "Zwerge",
]


def write_stufen_sheet(wb, sheet_name, rows):
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    ws.append(["Wert", "Gesamt-EP", "EP-Kosten (Schritt)"])
    for wert, gesamt, schritt in rows:
        ws.append([wert, gesamt, schritt])


def main(path_str: str) -> None:
    path = Path(path_str)
    backup_path = backup(str(path))
    print(f"Sicherheitskopie angelegt: {backup_path}")

    wb = openpyxl.load_workbook(path, data_only=False)
    ws = wb["Werte"]
    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v:
            headers[v.strip()] = c

    existing_refs = {}
    for r in range(2, ws.max_row + 1):
        ref = ws.cell(row=r, column=headers["Referenz"]).value
        if ref:
            existing_refs[str(ref).strip().lower()] = r

    # 1. rerolls = kreis
    if "rerolls" in existing_refs:
        print(f"'rerolls' existiert bereits in Zeile {existing_refs['rerolls']} - uebersprungen.")
    else:
        r = ws.max_row + 1
        ws.cell(row=r, column=headers["Referenz"]).value = "rerolls"
        ws.cell(row=r, column=headers["Kategorie"]).value = "Charakterwerte"
        ws.cell(row=r, column=headers["Beschreibung"]).value = "Rerolls"
        ws.cell(row=r, column=headers["Art"]).value = "Formel"
        ws.cell(row=r, column=headers["Formel"]).value = "kreis"
        ws.cell(row=r, column=headers["Flag"]).value = (
            "Neu ergaenzt (2026-07-17, Nutzerregel): Rerolls = Kreis."
        )
        print(f"Zeile {r}: 'rerolls' (Formel: kreis) ergaenzt.")

    # 2. Sprachstufe-Kosten / Kulturstufe-Kosten Lookup-Sheets
    write_stufen_sheet(wb, "Sprachstufe-Kosten", SPRACHSTUFE_ROWS)
    write_stufen_sheet(wb, "Kulturstufe-Kosten", KULTURSTUFE_ROWS)
    print("Sheets 'Sprachstufe-Kosten' / 'Kulturstufe-Kosten' geschrieben (5 Stufen je Tabelle).")

    # 3. whk_sprache_* auf die neue Sprachstufe-Kosten-Tabelle umstellen
    ws = wb["Werte"]  # create_sheet() kann die Sheet-Reihenfolge/-Referenz nicht aendern, sicherheitshalber neu holen
    updated = 0
    for r in range(2, ws.max_row + 1):
        ref = ws.cell(row=r, column=headers["Referenz"]).value
        if ref and str(ref).strip().lower().startswith("whk_sprache_"):
            ws.cell(row=r, column=headers["Kosten"]).value = "SVERWEIS(wert;'Sprachstufe-Kosten';2;0)"
            flag_cell = ws.cell(row=r, column=headers["Flag"])
            note = (
                "Kosten 2026-07-17 auf die 5-Stufen-EP-Tabelle aus 'NN Sprachen 0.11.docx' "
                "umgestellt (vorher generische WHK-Punktekurve)."
            )
            flag_cell.value = f"{flag_cell.value} | {note}" if flag_cell.value else note
            updated += 1
    print(f"{updated} whk_sprache_*-Zeilen auf Sprachstufe-Kosten umgestellt.")

    # 4. Neue whk_kultur_<volk>-Zeilen
    added = 0
    for volk in VOELKER:
        ref = f"whk_kultur_{volk.lower()}"
        if ref in existing_refs:
            print(f"'{ref}' existiert bereits in Zeile {existing_refs[ref]} - uebersprungen.")
            continue
        r = ws.max_row + 1
        ws.cell(row=r, column=headers["Referenz"]).value = ref
        ws.cell(row=r, column=headers["Kategorie"]).value = "WHK"
        ws.cell(row=r, column=headers["Beschreibung"]).value = volk
        ws.cell(row=r, column=headers["Art"]).value = "Wert"
        ws.cell(row=r, column=headers["Kosten"]).value = "SVERWEIS(wert;'Kulturstufe-Kosten';2;0)"
        ws.cell(row=r, column=headers["Flag"]).value = (
            "Neu ergaenzt (2026-07-17, Nutzerregel): Kulturfertigkeit pro Volk (aus "
            "Sheet 'Voelker-Maxima'), Kosten nach der Kulturstufe-Tabelle aus "
            "'NN Sprachen 0.11.docx'. Existierte vorher nicht im Datensatz."
        )
        added += 1
    print(f"{added} neue whk_kultur_*-Zeilen ergaenzt (von {len(VOELKER)} Voelkern).")

    wb.save(path)
    print(f"Gespeichert: {path}")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print('Aufruf: python scripts/add_sprache_kultur_regeln.py "werte 0.7-claude.xlsx"')
        sys.exit(1)
    main(sys.argv[1])
