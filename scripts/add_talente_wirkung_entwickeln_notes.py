"""
Traegt drei Entwickeln-Eintraege fuer die noch NICHT portierten Talent-Gruppen aus
"Talente-Wirkung-chatgpt.xlsx" ein (siehe scripts/extract_talente_maximum.py sowie
src/data/talenteFaktor.ts/talenteModifikator.ts fuer den bereits portierten Teil, und
die Memory project_talente_wirkung_analyse.md fuer den Gesamtueberblick):

1. Kampfstilmodifikator (6 Talente, strukturiert, aber bewusst ausgeschlossen - wirkt
   auf ALLE at_X/pa_X-Formeln gleichzeitig statt auf eine Einzelreferenz).
2. Portierungsstatus "teilstrukturiert" (34 Talente, davon 1 bereits ueber
   MANUAL_MAXIMUM_OVERRIDES abgedeckt -> 33 offen): Probenregel-, Freischaltung/
   Manoever- und Fernkampfmodifikator-Talente.
3. Portierungsstatus "manuell modellieren" (29 Talente, davon 5 bereits ueber
   MANUAL_MAXIMUM_OVERRIDES abgedeckt -> 24 offen): Komplexer-Regeltext-Talente.
"""
import sys
from pathlib import Path
from datetime import date

ROOT = Path(r"E:\Das Western Rollenspiel\LLM")
sys.path.insert(0, str(ROOT / ".python-deps"))

from openpyxl import load_workbook

WERTE = ROOT / "werte 0.8-claude.xlsx"
HEUTE = "2026-07-20"

ENTRIES = [
    {
        "Bereich": "Nahkampf/Kampfstil",
        "Thema": "Kampfstilmodifikator-Talente (Offensiver Kampfstil, Verteidiger) noch nicht portiert",
        "Beschreibung": (
            "Talente-Wirkung-chatgpt.xlsx fuehrt 6 Talente mit Wirkungsklasse "
            "'Kampfstilmodifikator': Offensiver Kampfstil Stufe 1-3 (nAT+1/+2/+3, "
            "nPA-1/-2/-3) und Verteidiger Stufe 1-3 (nAT-1/-2/-3, nPA+1/+2/+3). Alle 6 "
            "sind strukturell sauber (Portierungsstatus 'strukturiert', Zielreferenz "
            "nAT/nPA vorhanden), wurden aber bewusst NICHT in talenteMaximum.ts/"
            "talenteFaktor.ts/talenteModifikator.ts aufgenommen (siehe Kommentar in "
            "talenteModifikator.ts), weil sie nicht auf eine einzelne Formel-Referenz "
            "wirken, sondern auf ALLE at_X/pa_X-Kampfwaffen-Formeln gleichzeitig (jede "
            "Waffen-Attacke/Parade im System). Braucht eine eigene Architektur-"
            "Entscheidung (z.B. ein globaler Kampfstil-Modifikator, der nach der "
            "einzelnen Waffenformel angewendet wird, statt eine eigene Referenz je "
            "Waffe zu pflegen) - vermutlich Teil des geplanten Kampfmoduls, nicht der "
            "Charaktererstellung."
        ),
        "Status": "Offen",
        "Hinzugefuegt": HEUTE,
    },
    {
        "Bereich": "Kampf/Talente",
        "Thema": "Probenregel- und Manoever-Talente (teilstrukturiert) noch nicht portiert",
        "Beschreibung": (
            "34 Talente haben Portierungsstatus 'teilstrukturiert' in Talente-Wirkung-"
            "chatgpt.xlsx (Spalte AW), davon ist 'Charismatischer Fuehrer' bereits ueber "
            "MANUAL_MAXIMUM_OVERRIDES in extract_talente_maximum.py abgedeckt (33 "
            "verbleibend, offen). Drei Wirkungsklassen: "
            "(a) Probenregel (13): KI gute Stufe 1/2, Entwaffnen, Mit Schild umwerfen, "
            "Blutmagie Stufe 1-3, Spruchgute Stufe 1/2, PSI gute Stufe 1/2, "
            "Fernkampfgeschick Stufe 1/2 - aendern, wie 'gute/geschenkte' Proben-"
            "Schwellen fuer bestimmte Aktionen berechnet werden. "
            "(b) Freischaltung/Manoever (11 offen von 12): Meuchler, Kampf mit zwei "
            "Waffen Stufe 1-4, Konter, Wuchtschlag, Gezielter Schuss, Linkshaendig "
            "Pistolenschiessen, Mehrfachschuss Stufe 1/2, Mit Zwei Pistolen schiessen - "
            "schalten einen benannten Kampf-Manoever frei, dessen Ablauf eigene Regeln "
            "braucht. "
            "(c) Fernkampfmodifikator (6): Berittenes Werfen Stufe 1/2, Fliegend "
            "Schiessen, Schnell Schiessen, Sorgfaeltiges Zielen Stufe 1/2 - Situations-"
            "abhaengige Fernkampf-Erschwernisse/-Erleichterungen. Zusaetzlich 2 "
            "'Modifikator - komplex' (Berittenes Pistolenschiessen Stufe 1/2). Keines "
            "davon ist eine einfache 'Formel +/- X'-Regel wie bei talenteMaximum/"
            "-Faktor/-Modifikator, sondern aendert Ablauf/Ergebnis einer konkreten "
            "Kampfaktion - gehoert wie die Schild-Kampfregeln (siehe Eintrag "
            "'Nahkampf/Schilde') ins geplante Kampfmodul, nicht in die "
            "Charaktererstellung. Volltext je Talent (Spalten AT Bedingung/Ausloeser, "
            "AU Stufen-/Kumulierungslogik, AV Imperative Implementierungsanweisung) "
            "steht in Talente-Wirkung-chatgpt.xlsx bereit."
        ),
        "Status": "Offen",
        "Hinzugefuegt": HEUTE,
    },
    {
        "Bereich": "Kampf/Talente",
        "Thema": "Komplexe Freitext-Kampfregeln (manuell modellieren) noch nicht portiert",
        "Beschreibung": (
            "29 Talente haben Portierungsstatus 'manuell modellieren' (Wirkungsklasse "
            "durchgehend 'Komplexer Regeltext', keine strukturierte Zielreferenz), "
            "davon sind 5 bereits ueber MANUAL_MAXIMUM_OVERRIDES in "
            "extract_talente_maximum.py abgedeckt (Vorderlader Ladeschuetze Stufe 1/2, "
            "PSI Psinetik Stufe 1-3) - 24 verbleiben offen: ME sparen, Finte, "
            "Schnellziehen Hiebwaffen/Klingenwaffen/Stangenwaffen/Stichwaffen, Schnell "
            "Zaubern Stufe 1-3, Spruchmagie Stufe 2/3 zaubern, Berittenes "
            "Armbrustschiessen Stufe 1/2, Berittenes Bogenschiessen Stufe 1/2, "
            "Berittenes Musketenschiessen Stufe 1/2, Fernkampfgeschick Stufe 3, "
            "Point-Blank Shot, Schnell Ziehen Armbrust/Bogen/Muskete/Pistole/"
            "Wurfwaffe. Fuer jedes existiert ein ausformulierter Regeltext (Spalte F "
            "Beschreibung, jetzt auch in Werte!Wirkung) sowie Spalten AT (Bedingung/"
            "Ausloeser) und AV (Imperative Implementierungsanweisung) in Talente-"
            "Wirkung-chatgpt.xlsx, aber keine Formel-Referenz - jedes braucht "
            "handgeschriebene Engine-Logik statt einer generischen Bonus-Tabelle "
            "(aehnlich wie Kampfstilmodifikator und die Probenregel/Manoever-Gruppe: "
            "Kampfmodul-Scope, nicht Charaktererstellung)."
        ),
        "Status": "Offen",
        "Hinzugefuegt": HEUTE,
    },
]


def main():
    workbook = load_workbook(WERTE, read_only=False, data_only=False)
    worksheet = workbook["Entwickeln"]
    headers = {cell.value: cell.column for cell in worksheet[1] if cell.value}

    last_row = 0
    for row in worksheet.iter_rows(min_row=2):
        if any(c.value is not None for c in row):
            last_row = row[0].row

    row = last_row + 1
    for entry in ENTRIES:
        for header, value in entry.items():
            worksheet.cell(row, headers[header], value)
        row += 1

    workbook.save(WERTE)
    print(f"{len(ENTRIES)} Entwickeln-Eintraege geschrieben (Zeilen {last_row + 1}-{row - 1}).")


if __name__ == "__main__":
    main()
