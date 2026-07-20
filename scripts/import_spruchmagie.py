"""
Importiert/korrigiert die Spruchmagie-Kategorie im "Werte"-Sheet von werte-0.8-claude.xlsx
aus der Originalquelle NN_Spruchmagie_0.57.xlsx (Sheet "Spruchmagie komplett").

Hintergrund (siehe Memory project_spruchmagie_regeln.md, Punkte 11-15):
  - Die Quelle hat eine Spalte "Gueltig" (Ja/NN/NR/Nein). Fuer 70 Zaubernamen gibt es je
    eine NN- und eine NR-Kampagnenvariante (exakt 1x NN + 1x NR pro Name, verifiziert).
    Der letzte Import hat nicht gezielt die NN-Variante behalten. Fix: nur Gueltig in
    {Ja, NN} (bzw. leer, 1 Zeile "Material Stabilisieren, Konservieren") wird importiert -
    das loest den NN/NR-Fall automatisch mit, da fuer jedes Duplikat-Paar die NR-Zeile
    durch den Filter herausfaellt.
  - Neue Spalten (Min.Int, RW, Ziel, Form, Art, Aufr., VZ, ED, WD, St.1/2/3, Mana,
    Gegenprobe) gehen NICHT in RuleEntry (feste Spalten im "Werte"-Sheet), sondern in ein
    neues Begleit-Sheet "Spruchmagie-Details", analog zu "Artefakt-Basis"/"Artefakt-Kosten".

Vorgehen im "Werte"-Sheet (KEIN Zeilen-Loeschen/-Verschieben, da die 439 bestehenden
Spruchmagie-Zeilen ueber 4 nicht-zusammenhaengende Bloecke verteilt sind und andere
Kategorien dazwischen liegen): bestehende Spruchmagie-Zeilen werden per etabliertem
"#"-Praefix auf der Referenz deaktiviert (siehe generate_data_ts.py:153-158,
read_rules() ueberspringt "#..."-Referenzen), die frischen Zeilen werden ans Ende des
Sheets angehaengt.

Aufruf:
    python scripts/import_spruchmagie.py --dry-run     (nur Report, keine Aenderung)
    python scripts/import_spruchmagie.py --write        (Backup + echte Aenderung, kein Recalc)
"""
import argparse
import json
import re
import shutil
import sys
from collections import Counter
from datetime import datetime
from pathlib import Path

import openpyxl

ROOT = Path(r"E:\Das Western Rollenspiel\LLM")
SOURCE_XLSX = Path(
    r"E:\Das Western Rollenspiel\_Aktuelle Daten\Nasus\Nasus Nasus das Western Rollenspiel"
    r"\Die Magie\Spruch-Magie\NN_Spruchmagie_0.57.xlsx"
)
WERTE_XLSX = ROOT / "werte 0.8-claude.xlsx"
BACKUP_DIR = ROOT / "backups"

UMLAUT_MAP = str.maketrans({"ä": "ae", "ö": "oe", "ü": "ue", "ß": "ss"})

KOSTEN_RAW = "WENN(wert=0;0;10+(wert-1)*grad)"

# Eig-Spaltenwert (Quelle) -> eig_bonus_*-Referenz (siehe eigenschaft.jsonl/
# eigenschaftsbonus.jsonl: 4 "g_"-Eigenschaften + 6 "k_"-Eigenschaften).
EIG_TO_BONUS_REF = {
    "intelligenz": "eig_bonus_g_intelligenz",
    "mut": "eig_bonus_g_mut",
    "sinnesschaerfe": "eig_bonus_g_sinneschaerfe",
    "willenskraft": "eig_bonus_g_willenskraft",
    "athletik": "eig_bonus_k_athletik",
    "ausstrahlung": "eig_bonus_k_ausstrahlung",
    "geschicklichkeit": "eig_bonus_k_geschicklichkeit",
    "konstitution": "eig_bonus_k_konstitution",
    "schnelligkeit": "eig_bonus_k_schnelligkeit",
    "staerke": "eig_bonus_k_staerke",
}

RULE_COLUMNS = [
    ("Referenz", "referenz"),
    ("Kategorie", "kategorie"),
    ("Beschreibung", "beschreibung"),
    ("Abkürzung", "abkuerzung"),
    ("Info", "info"),
    ("Parent", "parent"),
    ("Art", "art"),
    ("Formel", "formelRaw"),
    ("Pool", "poolRaw"),
    ("Flag", "flag"),
    ("Grad", "grad"),
    ("Kosten", "kostenRaw"),
    ("Verfuegbarkeit", "verfuegbarkeit"),
    ("Mindest-TaW", "mindestTaw"),
    ("Eig-Bonus", "eigBonus"),
    ("Wirkung", "wirkung"),
]

DETAIL_HEADERS = [
    "Referenz", "MinInt", "Gegenprobe", "RW", "Ziel", "Form", "ZauberArt",
    "Aufrechterhaltung", "Vorbereitungszeit", "Einwirkdauer", "Wirkungsdauer",
    "Stufe1", "Stufe2", "Stufe3", "Mana",
]


def slugify(text):
    text = (text or "").strip().lower().translate(UMLAUT_MAP)
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return text.strip("_")


def normalize_eig(text):
    return slugify(text).replace("_", "")


def cell_to_str(value):
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def read_source_spells():
    wb = openpyxl.load_workbook(SOURCE_XLSX, data_only=True)
    ws = wb["Spruchmagie komplett"]
    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v:
            headers[v.strip()] = c

    def col(header):
        return headers[header]

    spells = []
    warnings = []
    seen_referenz = {}
    gueltig_counts = Counter()
    for r in range(2, ws.max_row + 1):
        name = cell_to_str(ws.cell(r, col("Name")).value)
        if not name:
            continue
        schule = cell_to_str(ws.cell(r, col(" Schule")).value if " Schule" in headers else ws.cell(r, col("Schule")).value)
        grad = cell_to_str(ws.cell(r, col("Grad")).value)
        gueltig = cell_to_str(ws.cell(r, col("Gültig")).value)
        gueltig_counts[gueltig] += 1

        if gueltig not in ("Ja", "NN", None):
            continue  # Nein/NR -> "auskommentiert" (Regel 2/4/11)

        min_int = cell_to_str(ws.cell(r, col("Min.Int")).value)
        eig = cell_to_str(ws.cell(r, col("Eig")).value)
        wirkung = cell_to_str(ws.cell(r, col("Wirkung")).value)
        gegenprobe = cell_to_str(ws.cell(r, col("Gegenprobe")).value)
        rw = cell_to_str(ws.cell(r, col("RW")).value)
        ziel = cell_to_str(ws.cell(r, col("Ziel")).value)
        form = cell_to_str(ws.cell(r, col("Form")).value)
        zauber_art = cell_to_str(ws.cell(r, col("Art")).value)
        aufr = cell_to_str(ws.cell(r, col("Aufr.")).value)
        vz = cell_to_str(ws.cell(r, col("VZ in")).value)
        ed = cell_to_str(ws.cell(r, col("ED")).value)
        wd = cell_to_str(ws.cell(r, col("WD")).value)
        st1 = cell_to_str(ws.cell(r, col("St. 1")).value)
        st2 = cell_to_str(ws.cell(r, col("St. 2")).value)
        st3 = cell_to_str(ws.cell(r, col("St. 3")).value)
        mana = cell_to_str(ws.cell(r, col("Mana")).value)

        referenz = f"spruchmagie_{slugify(schule)}_{grad}_{slugify(name)}"
        if referenz in seen_referenz:
            warnings.append(
                f"Quellzeile {r}: doppelte generierte Referenz '{referenz}' "
                f"(zuerst Quellzeile {seen_referenz[referenz]}) - uebersprungen"
            )
            continue
        seen_referenz[referenz] = r

        eig_bonus_ref = EIG_TO_BONUS_REF.get(normalize_eig(eig))
        if eig is not None and eig_bonus_ref is None:
            warnings.append(f"Quellzeile {r} ('{name}'): unbekannter Eig-Wert '{eig}' - eigBonus bleibt leer")

        spells.append({
            "referenz": referenz,
            "kategorie": "Spruchmagie",
            "beschreibung": name,
            "parent": schule,
            "art": "Wert",
            "grad": grad,
            "kostenRaw": KOSTEN_RAW,
            "eigBonus": eig_bonus_ref,
            "wirkung": wirkung,
            "sourceRow": r,
            "detail": {
                "Referenz": referenz,
                "MinInt": min_int,
                "Gegenprobe": gegenprobe,
                "RW": rw,
                "Ziel": ziel,
                "Form": form,
                "ZauberArt": zauber_art,
                "Aufrechterhaltung": aufr,
                "Vorbereitungszeit": vz,
                "Einwirkdauer": ed,
                "Wirkungsdauer": wd,
                "Stufe1": st1,
                "Stufe2": st2,
                "Stufe3": st3,
                "Mana": mana,
            },
        })

    return spells, warnings, gueltig_counts


def read_existing_werte_spruchmagie(wb):
    ws = wb["Werte"]
    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v:
            headers[v.strip()] = c
    ref_col = headers["Referenz"]
    kat_col = headers["Kategorie"]
    besch_col = headers["Beschreibung"]
    grad_col = headers["Grad"]
    wirkung_col = headers["Wirkung"]

    rows = []
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, kat_col).value == "Spruchmagie":
            rows.append({
                "row": r,
                "referenz": ws.cell(r, ref_col).value,
                "beschreibung": ws.cell(r, besch_col).value,
                "grad": cell_to_str(ws.cell(r, grad_col).value),
                "wirkung": ws.cell(r, wirkung_col).value,
            })
    return rows, headers


def build_report(new_spells, warnings, gueltig_counts, existing_rows):
    lines = []
    lines.append(f"Quelle Gueltig-Verteilung: {dict(gueltig_counts)}")
    lines.append(f"Bestehende Werte-Sheet Spruchmagie-Zeilen: {len(existing_rows)}")
    lines.append(f"Neue aktive Zauber (Gueltig=Ja/NN/leer): {len(new_spells)}")
    if warnings:
        lines.append(f"{len(warnings)} Warnung(en):")
        lines.extend(f"  - {w}" for w in warnings)

    existing_by_key = {(e["beschreibung"], e["grad"]): e for e in existing_rows}
    new_by_key = {(s["beschreibung"], s["grad"]): s for s in new_spells}

    changed_wirkung = []
    for key, new_s in new_by_key.items():
        old = existing_by_key.get(key)
        if old and old["wirkung"] != new_s["wirkung"]:
            changed_wirkung.append(key[0])
    removed = sorted(set(existing_by_key) - set(new_by_key))
    added = sorted(set(new_by_key) - set(existing_by_key))

    lines.append(f"Zauber mit geaendertem Wirkungstext (NN/NR-Korrektur): {len(changed_wirkung)}")
    lines.extend(f"  - {n}" for n in sorted(changed_wirkung)[:15])
    if len(changed_wirkung) > 15:
        lines.append(f"  ... (+{len(changed_wirkung) - 15} weitere)")
    lines.append(f"Bisher vorhanden, jetzt entfernt (Gueltig=Nein/NR o.ae.): {len(removed)}")
    lines.extend(f"  - {n} (Grad {g})" for n, g in removed[:15])
    if len(removed) > 15:
        lines.append(f"  ... (+{len(removed) - 15} weitere)")
    lines.append(f"Neu hinzugekommen (vorher nicht importiert): {len(added)}")
    lines.extend(f"  - {n} (Grad {g})" for n, g in added[:15])
    if len(added) > 15:
        lines.append(f"  ... (+{len(added) - 15} weitere)")
    return "\n".join(lines)


def do_write(new_spells, existing_rows, headers):
    BACKUP_DIR.mkdir(exist_ok=True)
    stamp = datetime.now().strftime("%Y-%m-%d_%H%M")
    backup_path = BACKUP_DIR / f"werte 0.8-claude_{stamp}_pre-spruchmagie-import.xlsx"
    shutil.copy2(WERTE_XLSX, backup_path)
    print(f"Backup geschrieben: {backup_path}")

    wb = openpyxl.load_workbook(WERTE_XLSX, data_only=False)
    ws = wb["Werte"]
    ref_col = headers["Referenz"]

    for row in existing_rows:
        cell = ws.cell(row["row"], ref_col)
        if cell.value and not str(cell.value).startswith("#"):
            cell.value = f"#{cell.value}"

    next_row = ws.max_row + 1
    for spell in new_spells:
        for header, field in RULE_COLUMNS:
            col = headers[header]
            ws.cell(next_row, col, spell.get(field))
        next_row += 1

    if "Spruchmagie-Details" in wb.sheetnames:
        del wb["Spruchmagie-Details"]
    details_ws = wb.create_sheet("Spruchmagie-Details")
    for c, header in enumerate(DETAIL_HEADERS, start=1):
        details_ws.cell(1, c, header)
    for r, spell in enumerate(new_spells, start=2):
        detail = spell["detail"]
        for c, header in enumerate(DETAIL_HEADERS, start=1):
            details_ws.cell(r, c, detail.get(header))

    wb.save(WERTE_XLSX)
    print(f"{WERTE_XLSX} geschrieben: {len(existing_rows)} alte Zeilen deaktiviert (#-Praefix), "
          f"{len(new_spells)} neue Zeilen angehaengt, Sheet 'Spruchmagie-Details' geschrieben.")
    print("WICHTIG: Datei braucht noch einen Excel-COM-Recalc-Pass vor dem naechsten "
          "openpyxl(data_only=True)-Lesen (siehe Memory reference_xlsx_recalc_windows).")


def main():
    parser = argparse.ArgumentParser()
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--dry-run", action="store_true")
    group.add_argument("--write", action="store_true")
    args = parser.parse_args()

    if not SOURCE_XLSX.exists():
        raise SystemExit(f"Quelle nicht gefunden: {SOURCE_XLSX}")
    if not WERTE_XLSX.exists():
        raise SystemExit(f"Ziel nicht gefunden: {WERTE_XLSX}")

    new_spells, warnings, gueltig_counts = read_source_spells()

    wb = openpyxl.load_workbook(WERTE_XLSX, data_only=True)
    existing_rows, _ = read_existing_werte_spruchmagie(wb)

    print(build_report(new_spells, warnings, gueltig_counts, existing_rows))

    if args.write:
        wb_formulas = openpyxl.load_workbook(WERTE_XLSX, data_only=False)
        _, headers = read_existing_werte_spruchmagie(wb_formulas)
        do_write(new_spells, existing_rows, headers)


if __name__ == "__main__":
    main()
