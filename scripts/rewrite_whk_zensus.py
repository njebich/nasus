"""One-off migration: replace the entire WHK catalog in 'werte 0.8-claude.xlsx' (Werte sheet)
with the Zensus-Entwurf taxonomy from WHK-ZENSUS-ENTWURF.md (Runde 8, 2026-08-05).

Deletes every row with Kategorie == "WHK", then inserts the full target taxonomy:
- Hauptfertigkeit rows get Parent = Kategorie-Gruppe (e.g. "Bau") for the new 3rd UI level.
- Spezialisierung rows keep Parent = Hauptfertigkeit-Name (unchanged existing convention).
- Geweihte: Stossgebet/Wunder/Ritual keep their exact existing referenzen (hardcoded in
  src/views/geweihte.ts) and Info text - everything else gets fresh referenz slugs.

Usage: python rewrite_whk_zensus.py "werte 0.8-claude.xlsx"
"""
import re
import sys

import openpyxl

HF_KOSTEN = "WENN(wert=0;0;10+(wert-1)*wert/2)"
SPEZ_KOSTEN = "SVERWEIS(wert;'WHK-Spez-Kosten';3;1)"

UMLAUT_MAP = str.maketrans({
    "ä": "ae", "ö": "oe", "ü": "ue", "ß": "ss",
    "Ä": "Ae", "Ö": "Oe", "Ü": "Ue",
})


def slugify(name: str) -> str:
    s = name.translate(UMLAUT_MAP).lower()
    s = re.sub(r"[^a-z0-9]+", "_", s)
    return s.strip("_")


# (Kategorie-Gruppe, [(Hauptfertigkeit, [Spezialisierungen] | [] fuer keine/Freitext-Spez)])
TAXONOMY = [
    ("Bau", [
        ("Zimmerei", ["Dachstuhlbauer", "Fenster- und Türenbauer", "Schindelmacher", "Fachwerkbauer"]),
        ("Maurerhandwerk", ["Ziegelmaurer", "Stuckateur", "Dachdecker", "Fliesenleger"]),
        ("Architektur", ["Brückenbauingenieur", "Festungsbaumeister", "Stadtplaner", "Baumeister"]),
        ("Holzbearbeitung", ["Drechsler", "Bürstenbinder", "Fassbinder", "Korbflechter", "Köhler",
                              "Schnitzer", "Tischler", "Wagner", "Holzfäller"]),
        ("Bootsbauer", ["Großschiffbauer", "Kleinbootbauer", "Galeerenbauer"]),
        ("Glaser", ["Apparateglaser", "Glasbläser", "Kristallschleifer", "Optiker", "Fensterglaser"]),
        ("Maler", ["Fassadenmaler", "Anstreicher", "Tapezierer"]),
    ]),
    ("Metall", [
        ("Schmied", ["Feinschmied", "Grobschmied", "Hufschmied"]),
        ("Buntmetallschmied", ["Bronzeschmied", "Goldschmied", "Kupferschmied", "Silberschmied", "Zinngießer"]),
        ("Waffenschmied", ["Klingenwaffen", "Hiebwaffen", "Stangenwaffen", "Stichwaffen"]),
        ("Rüstungsschmied", ["Tierrüstungen", "Kettenrüstungen", "Plattenpanzer"]),
        ("Feuerwaffenbauer", ["Vorderlader", "Stein-/Perkussionsschloss", "Kipplaufverschluss",
                               "Trommelverschluss", "Repetierverschluss"]),
        ("Metallurgie", ["Edelmetalle", "Schwermetalle", "Sonderlegierungen", "Eisen & Stahl", "Messing"]),
        ("Gießer", ["Glockengießer", "Werkzeuggießer", "Stückgießer", "Munitionsgießer"]),
        ("Bogenbauer", ["Holzbögen", "Kompositbögen", "Umspannbögen"]),
    ]),
    ("Textil/Leder", [
        ("Schneider", ["Segelmacher", "Konfektionsschneider", "Hutmacher", "Teppichknüpfer", "Stoffrüstungen"]),
        ("Weber", ["Seidenweber", "Leinweber", "Tuchmacher", "Färber", "Stricker"]),
        ("Lederbearbeitung", ["Handschuhmacher", "Oberbekleidung", "Sattler", "Schuster", "Täschner",
                               "Gerber", "Kürschner", "Lederrüstungen"]),
        ("Seiler", ["Netzmacher", "Reepschläger"]),
    ]),
    ("Nahrung", [
        ("Bäcker", ["Brotbäcker", "Konditor"]),
        ("Koch", ["Backen", "Braten", "Räuchern", "Konservieren"]),
        ("Metzger", ["Schlachtung", "Pökeln", "Trocknen", "Wursterei", "Talgsieder"]),
        ("Brenner", ["Alchemistische Zutaten", "Geist", "Kräutergeist"]),
        ("Fischer", ["Angler", "Tranbrenner", "Netzfischer", "Schwammtaucher", "Walfänger", "Austernfischer"]),
    ]),
    ("Sonstiges Handwerk", [
        ("Töpfer", ["Emailleur", "Geschirr"]),
        ("Kunsthandwerker", ["Bildhauer", "Maler", "Tätowierer", "Elfenbeinschnitzer"]),
        ("Instrumentenbauer", ["Blasinstrumente", "Saiteninstrumente", "Schlaginstrumente", "Tasteninstrumente"]),
        ("Mechanikus", ["Feinmechaniker", "Rechenmaschinen", "Uhrmacher", "Maschinenbauer",
                         "Dampftechniker", "Pumpenbauer"]),
        ("Barbier", ["Bader", "Frisör", "Kosmetiker"]),
        ("Drucker", ["Buchbinder", "Papiermacher", "Lithograf", "Stereotypeur"]),
        ("Luftschiffbauer", ["Militärschiffbauer", "Reiseschiffbauer", "Transportschiffbauer"]),
    ]),
    ("Natur/Landwirtschaft", [
        ("Landwirtschaft", ["Gemüsebauer", "Getreidebauer", "Obstbauer", "Rübenbauer", "Tabakpflanzer",
                             "Winzer", "Baumschulgärtner", "Gewürze", "Baumwollpflanzer", "Alchemistische Zutaten"]),
        ("Viehwirtschaft", ["Geflügelzüchter", "Imker", "Milchbauer", "Weidehirte", "Viehtreiber",
                             "Großechsenzüchter", "Riesenasselzüchter", "Wargzüchter"]),
        ("Jäger", ["Fallensteller", "Großwildjäger", "Treiber", "Rotwildjäger"]),
        ("Pflanzenkunde", ["Förster", "Farbstoffe", "Kräuterkundler", "Moose", "Pilzsammler",
                            "Sträucher", "Alchemistische Zutaten"]),
        ("Bergbau", ["Edelsteinschürfer", "Steinbrecher", "Kohlenhauer", "Salzbergmann", "Ölbohrer", "Prospektor"]),
        ("Alchemie", ["Feuerwerker", "Pigmentmacher", "Leimsieder", "Parfumeur", "Pulvermacher",
                       "Trankbrauer", "Giftmischer", "Salbenmischer", "Seifensieder"]),
        ("Abrichten", ["Drache", "Laufvogel", "Reitkrebse", "Reitreptilien", "Rok", "Warg", "Asseln",
                        "Elefanten", "Flugechsen", "Fledertiere", "Greifen", "Hornträger", "Hundeartige",
                        "Katzenartige", "Kamele", "Pegasi", "Pferde", "Primaten", "Raptoren", "Rollocks",
                        "Rattenartige", "Schleichkatzen", "Spinnen", "Schweine", "Vögel"]),
    ]),
    ("Handel/Verwaltung", [
        ("Kaufmann", ["Gastwirt", "Sklavenhändler", "Artefakthändler", "Bogenhändler", "Armbrusthändler",
                       "Feuerwaffenhändler", "Krämer", "Kutschenhändler", "Schiffsmakler", "Blankwaffenhändler",
                       "Rüstungshändler", "Schmuckhändler", "Baustoffhändler", "Textilhändler",
                       "Haushaltswarenhändler", "Metallwarenhändler", "Provianthändler", "Papierwarenhändler",
                       "Immobilienmakler", "Rohstoffhändler", "Musikalienhändler"]),
        ("Bankwesen", ["Kreditwesen", "Einlagengeschäft", "Wechsler", "Hypothekenwesen"]),
        ("Versicherungswesen", ["Feuerversicherung", "Lebensversicherung", "Transportversicherung", "Schiffsversicherung"]),
        ("Buchhaltung", ["Bilanzbuchhalter", "Betriebsprüfer", "Kostenrechnung"]),
        ("Administration", ["Schreiber", "Landvermesser", "Archivar", "Verwaltungsbeamter"]),
        ("Rechtskunde", ["Vertragsrecht", "Strafrecht", "Bergrecht", "Bürgerrecht"]),
        ("Logistik", ["Bahnspediteur", "Fuhrspediteur", "Hafenspediteur", "Schiffsspediteur"]),
        ("Fuhrmann", ["Lastfuhrmann", "Postkutscher", "Reisekutscher"]),
        ("Eisenbahner", ["Lokführer", "Signaler"]),
    ]),
    ("Führung/Reise", [
        ("Führung", ["Dampferkapitän", "Karawanenführer", "Luftschiffkapitän", "Schiffskapitän", "Zöllner",
                      "Gutsverwalter", "Lotse", "Werksleiter", "Kavallerieoffizier", "Artillerieoffizier",
                      "Infanterieoffizier"]),
        ("Seefahrt", ["Küstenschifffahrt", "Hochseeschifffahrt", "Ruderschifffahrt", "Flussschifffahrt"]),
        ("Luftschifffahrt", ["Höhenfahrt", "Sturmfahrt", "Langstreckenfahrt"]),
        ("Überleben", ["Berge", "Dschungel", "Ebene", "Eis", "Stadt", "Unterirdisch", "Wald", "Wüste"]),
        ("Telegrafie", ["Telegrafist", "Leitungsbauer", "Kryptografie"]),
    ]),
    ("Wissenschaft", [
        ("Chemie", ["Anorganische Chemie", "Organische Chemie", "Thermochemie"]),
        ("Physik", ["Mechanik", "Optik", "Thermodynamik"]),
        ("Mathematik", ["Arithmetik", "Geometrie", "Statistik"]),
        ("Philosophie", ["Ethik", "Logik", "Naturphilosophie"]),
        ("Wirtschaftslehre", ["Volkswirtschaftslehre", "Betriebswirtschaftslehre", "Außenhandel"]),
        ("Soziologie", ["Vergleichende Kulturkunde", "Gesellschaftsstruktur", "Migrationsforschung"]),
        ("Astronomie", ["Navigationsastronomie", "Kalenderkunde", "Kosmologie"]),
        ("Geschichte", []),  # Freitext-Spez ueber die bestehende Custom-Spez-Funktion
        ("Sprachwissenschaft", ["Vergleichende Grammatik", "Etymologie", "Dialektologie"]),
        ("Altzwergisch", []),
        ("Altelfisch", []),
        ("Biologie", ["Zoologie", "Botanik", "Anatomie"]),
        ("Theologie", ["Religionswissenschaft", "Kirchengeschichte", "Mythologie"]),
        ("Geologie", ["Lagerstättenkunde", "Mineralogie", "Tektonik"]),
        ("Geografie", ["Kartografie", "Länderkunde", "Klimakunde"]),
        ("Militärtheorie", ["Strategie", "Taktik"]),
        ("Magietheorie", ["Antimagie", "Beherrschung", "Erdbeschwörung", "Feuerbeschwörung", "Heilung",
                           "Hellsicht", "Illusion", "Luftbeschwörung", "Magiebeschwörung", "Veränderung",
                           "Verzauberung", "Wasserbeschwörung", "Dämonologie", "Nekromantie", "PSI-Theorie"]),
        ("Medizin", ["Veterinär", "Chirurg", "Zahnarzt", "Geburtshelfer", "Erste Hilfe"]),
    ]),
    ("Magie/Klerus", [
        ("Geweihte: Stoßgebet", []),
        ("Geweihte: Wunder", []),
        ("Geweihte: Ritual", []),
    ]),
    ("Sozial/Kunst", [
        ("Ermittlung", ["Spurenlesen", "Verhör", "Observation"]),
        ("Diebeskunst", ["Taschendieb", "Schlossknacker", "Taschenspieler", "Betrüger"]),
        ("Etikette", ["Gemeinvolk", "Bürgertum", "Hochfinanz"]),
        ("Pädagoge", []),
        ("Musizieren", ["Blasinstrumente", "Saiteninstrumente", "Schlaginstrumente"]),
        ("Schriftstellerei", ["Journalismus", "Belletristik", "Fachliteratur"]),
        ("Fotografie", ["Studioaufnahmen", "Außenaufnahmen", "Bewegtbild"]),
        ("Schauspielkunst", ["Schauspieler", "Redner", "Schausteller"]),
        ("Tänzer", []),
        ("Gastgewerbe", ["Kellner", "Barkeeper", "Sommelier"]),
        ("Hauswirtschaft", ["Haushälter", "Wäscher", "Hausmeister"]),
        ("Bestattungswesen", ["Einbalsamierer", "Rekonstruktion", "Präsentation"]),
        ("Stimmen imitieren", ["Dalkini", "Drow", "Elf", "Goblin", "Indianer", "Katzenmenschen",
                                "Ork", "Troll", "Zentauren", "Zwerg", "Tiere"]),
    ]),
]

# Ausnahmen: exakte Referenz + Info-Text beibehalten (hardcodiert in src/views/geweihte.ts)
GEWEIHTE_INFO = {
    "Geweihte: Stoßgebet": (
        "whk_geweihte_stossgebet",
        "Bestimmt die Probe fuer Wunder vom Typ 'Stoß' (kurze, sofortige Anrufungen).",
    ),
    "Geweihte: Wunder": (
        "whk_geweihte_wunder",
        "Bestimmt die Probe fuer Wunder vom Typ 'Wunder' (laenger wirkende Anrufungen).",
    ),
    "Geweihte: Ritual": (
        "whk_geweihte_ritual",
        "Bestimmt die Probe fuer Wunder vom Typ 'Ritual' (aufwendige, lange Zeremonien).",
    ),
}


def main():
    if len(sys.argv) != 2:
        raise SystemExit("Usage: python rewrite_whk_zensus.py <xlsx-path>")
    path = sys.argv[1]

    wb = openpyxl.load_workbook(path, data_only=False)
    ws = wb["Werte"]

    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v:
            headers[v.strip()] = c

    col_referenz = headers["Referenz"]
    col_kategorie = headers["Kategorie"]
    col_beschreibung = headers["Beschreibung"]
    col_info = headers["Info"]
    col_parent = headers["Parent"]
    col_art = headers["Art"]
    col_kosten = headers["Kosten"]

    # 1) Alle bestehenden WHK-Zeilen einsammeln und loeschen (rueckwaerts, damit Indizes stabil bleiben).
    whk_rows = [r for r in range(2, ws.max_row + 1) if ws.cell(row=r, column=col_kategorie).value == "WHK"]
    print(f"Loesche {len(whk_rows)} bestehende WHK-Zeilen...")
    for r in reversed(whk_rows):
        ws.delete_rows(r, 1)

    # 2) Neue Taxonomie anhaengen.
    seen_referenz = set()

    def write_row(row_idx, referenz, beschreibung, parent, kosten, info=None):
        if referenz in seen_referenz:
            raise SystemExit(f"Doppelte Referenz erzeugt: {referenz}")
        seen_referenz.add(referenz)
        ws.cell(row=row_idx, column=col_referenz, value=referenz)
        ws.cell(row=row_idx, column=col_kategorie, value="WHK")
        ws.cell(row=row_idx, column=col_beschreibung, value=beschreibung)
        ws.cell(row=row_idx, column=col_parent, value=parent)
        ws.cell(row=row_idx, column=col_art, value="Wert")
        ws.cell(row=row_idx, column=col_kosten, value=kosten)
        if info:
            ws.cell(row=row_idx, column=col_info, value=info)

    row_idx = ws.max_row + 1
    hf_count = 0
    spez_count = 0
    for kategorie_gruppe, hauptfertigkeiten in TAXONOMY:
        for hf_name, spez_list in hauptfertigkeiten:
            if hf_name in GEWEIHTE_INFO:
                hf_referenz, info = GEWEIHTE_INFO[hf_name]
            else:
                hf_referenz, info = f"whk_{slugify(hf_name)}", None
            write_row(row_idx, hf_referenz, hf_name, kategorie_gruppe, HF_KOSTEN, info)
            row_idx += 1
            hf_count += 1

            for spez_name in spez_list:
                spez_referenz = f"whk_spez_{slugify(hf_name)}_{slugify(spez_name)}"
                write_row(row_idx, spez_referenz, f"-> {spez_name}", hf_name, SPEZ_KOSTEN)
                row_idx += 1
                spez_count += 1

    print(f"Neu: {hf_count} Hauptfertigkeiten, {spez_count} Spezialisierungen, {hf_count + spez_count} Zeilen total.")

    wb.save(path)
    print(f"Gespeichert: {path}")


if __name__ == "__main__":
    main()
