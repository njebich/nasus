"""Extrahiert die Talente, deren Wirkung ausschliesslich eine Kampfrunden-/Proben-Mechanik ist
(Manoever, Situationsmodifikatoren, Haltungswechsel, Proben-Sonderregeln) - also KEIN
Fertigkeitsmaximum/Charakterwertmodifikator/Faktor, der sich als statischer Charakterbogenwert
speichern liesse. Schreibt src/data/talenteKampfmodul.ts, gelesen von
engine/talenteKampfmodulInfo.ts fuer die reine Info-Anzeige im Kampf-Tab (views/kampf.ts) -
KEINE mechanische Wirkung, nur ein Hinweistext fuer ein kuenftiges Kampfmodul/den Meister.

Nutzer 2026-07-21: "wire the actual effects of all talente... open cases will need to be decided
either by group or be deferred straight to Kampfmodul (because they only add Kampf abilities/
mutations of attacks, nothing stored in character but the ability), after adding info lines shown
in the kampf tab." Gruppenentscheidung (keine Einzelrueckfrage noetig): alle 6 Wirkungsklassen
unten sind durchgaengig Kampfrunden-Mechaniken ohne editierbaren Zielwert - bestaetigt durch
Gegenpruefung mit SAVEPOINT-Talente-Wirkung-2026-07-18.md (ChatGPT-Regelklaerung, die fast jeden
dieser Namen unter Begriffen wie "KR", "Ansage", "Probe" beschreibt) UND durch einen Script-Check,
dass keine dieser talente_*-Referenzen in irgendeiner formelRaw von rules.json auftaucht (also
nirgends bereits als Formel-Baustein verwendet wird).

Ausnahme: "Fernkampfgeschick Stufe 1/2/3" sieht auf den ersten Blick wie Kampfmodul-Text aus
("Komplexer Regeltext"/"Probenregel"), ist aber TATSAECHLICH schon voll wirksam - die Basis-xlsx-
Formeln fk_gute_spez_*/fk_meisterlich_spez_* referenzieren talente_fernkampfgeschick_stufe_1/2/3
direkt (WENN(...)-Verzweigung), lange bevor diese Talente-Wirkung-Analyse existierte. Deshalb hier
explizit ausgeschlossen - kein Info-Zeile noetig, die Zahl steht schon live in der Fernkampf-
Tabelle in kampf.ts.
"""
import json
import sys
from pathlib import Path

import openpyxl

OUT_PATH = Path(__file__).parent.parent / "src" / "data" / "talenteKampfmodul.ts"
RULES_JSON_PATH = Path(__file__).parent.parent / "src" / "data" / "rules.json"
DATA_DIR = Path(__file__).parent.parent / "src" / "data"

KAMPFMODUL_WIRKUNGSKLASSEN = {
    "Probenregel",
    "Freischaltung/Manöver",
    "Kampfstilmodifikator",
    "Fernkampfmodifikator",
    "Modifikator – komplex",
    "Komplexer Regeltext",
}

RETIRED_TALENTE = {"talente_ladeschuetze_schleuder"}
BEREITS_FORMELVERDRAHTET = {
    "talente_fernkampfgeschick_stufe_1",
    "talente_fernkampfgeschick_stufe_2",
    "talente_fernkampfgeschick_stufe_3",
}


def read_rows(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Talente"]
    headers = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        headers.append(v.strip() if v else f"col{c}")
    rows = []
    for r in range(2, ws.max_row + 1):
        row = {}
        for c, h in enumerate(headers, start=1):
            v = ws.cell(row=r, column=c).value
            if v is not None:
                row[h] = v
        if row:
            rows.append(row)
    return rows


def read_built_referenzen():
    """talentReferenz-Werte, die bereits ueber talenteMaximum/-Modifikator/-Faktor.ts einen
    echten mechanischen Bonus bekommen (inkl. MANUAL_MAXIMUM_OVERRIDES, z.B. Charismatischer
    Fuehrer/PSI Psinetik/Vorderlader Ladeschuetze - liefen in der Analysedatei unter einer der
    KAMPFMODUL_WIRKUNGSKLASSEN, sind aber tatsaechlich schon als Maximum-Bonus verdrahtet, siehe
    extract_talente_maximum.py). Ohne diesen Abgleich wuerden sie hier doppelt (und falsch, als
    reine Info-Zeile) auftauchen."""
    refs = set()
    for name in ("talenteMaximum.ts", "talenteModifikator.ts", "talenteFaktor.ts"):
        text = (DATA_DIR / name).read_text(encoding="utf-8")
        refs.update(__import__("re").findall(r"talentReferenz[\"']?:\s*[\"']([a-z0-9_]+)[\"']", text))
    return refs


def extract(rows):
    bereits_gebaut = read_built_referenzen()
    refs = set()
    for r in rows:
        if r.get("Wirkungsklasse") not in KAMPFMODUL_WIRKUNGSKLASSEN:
            continue
        for ref in [x.strip() for x in (r.get("Werte-Referenz(en)") or "").split("\n") if x.strip()]:
            if ref in RETIRED_TALENTE or ref in BEREITS_FORMELVERDRAHTET or ref in bereits_gebaut:
                continue
            refs.add(ref)
    return sorted(refs)


def validate(refs, rules_json_path):
    rules = json.loads(rules_json_path.read_text(encoding="utf-8"))
    referenz_set = {r["referenz"] for r in rules}
    return [f"'{ref}' existiert nicht in rules.json" for ref in refs if ref not in referenz_set]


def write_ts(refs, out_path):
    lines = [
        "// GENERIERT von scripts/extract_talente_kampfmodul.py aus Talente-Wirkung-chatgpt.xlsx -",
        "// nicht von Hand bearbeiten, stattdessen das Skript erneut laufen lassen. Siehe Skript-",
        "// Docstring fuer die Gruppenentscheidung, warum diese Talente NUR eine Info-Zeile im",
        "// Kampf-Tab bekommen (views/kampf.ts / engine/talenteKampfmodulInfo.ts) statt einer",
        "// mechanischen Wirkung: reine Kampfrunden-/Proben-Mechaniken (Manoever,",
        "// Haltungswechsel, Situationsmodifikatoren), kein editierbarer Charakterbogenwert.",
        "export const TALENTE_KAMPFMODUL: readonly string[] = " + json.dumps(refs, ensure_ascii=False, indent=2) + ";",
        "",
    ]
    out_path.write_text("\n".join(lines), encoding="utf-8")


def main(xlsx_path):
    rows = read_rows(xlsx_path)
    refs = extract(rows)
    problems = validate(refs, RULES_JSON_PATH)
    if problems:
        raise SystemExit("Validierungsfehler:\n" + "\n".join(problems))
    write_ts(refs, OUT_PATH)
    print(f"{OUT_PATH}: {len(refs)} Kampfmodul-Info-Talente geschrieben.")


if __name__ == "__main__":
    xlsx = sys.argv[1] if len(sys.argv) > 1 else "Talente-Wirkung-chatgpt.xlsx"
    main(xlsx)
