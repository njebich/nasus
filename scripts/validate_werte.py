"""
Konsistenz-Check fuer eine werte-*.xlsx Datei (Sheet "Werte").

Aufruf:
    python scripts/validate_werte.py "werte 0.7-claude.xlsx"

Prueft (siehe Funktionen unten fuer Details):
  - doppelte Referenzen
  - Zeilen mit Daten aber ohne Referenz
  - Gross-/Kleinschreibung der Art-Spalte
  - Formel-/Kosten-Spalte: Verweise auf nicht existierende Referenzen
  - Leerzeichen am Rand von Beschreibung/Referenz
  - Format der Verfuegbarkeit-Spalte (E / E&Zahl / Zahl)

Gibt am Ende eine Zusammenfassung aus. Exit-Code 0 wenn alles sauber ist,
sonst 1 (nuetzlich falls das mal automatisiert laufen soll).

Aendert NICHTS an der Datei - reines Lesen und Berichten.
"""
import sys
import re
from collections import Counter

import openpyxl

TOKEN_RE = re.compile(r"[A-Za-zÄÖÜäöüß_][A-Za-zÄÖÜäöüß0-9_]*")
QUOTE_RE = re.compile(r"'[^']*'")
KEYWORDS = {"MIN", "MAX", "WENN", "SVERWEIS"}
SELF_REF_TOKENS = {"wert", "grad"}
UNIT_WORDS = {"m", "s", "je", "nach", "Pferd", "ca"}
KNOWN_ART_VALUES = {"Wert", "Formel", "Pool", "Auswahl", "Fixwert", "Lookup"}
VERFUEGBARKEIT_RE = re.compile(r"^(E(&\d+)?|\d+)$")


def load_rows(ws, headers):
    idx = {h: i + 1 for i, h in enumerate(headers)}
    rows = []
    for r in range(2, ws.max_row + 1):
        vals = {h: ws.cell(row=r, column=idx[h]).value for h in headers}
        if all(v in (None, "") for v in vals.values()):
            continue
        rows.append((r, vals))
    return rows, idx


def check_duplicate_refs(rows):
    refs = [v["Referenz"] for r, v in rows if v.get("Referenz")]
    dupes = [k for k, c in Counter(refs).items() if c > 1]
    return dupes


def check_missing_refs(rows):
    return [
        (r, v.get("Kategorie"), v.get("Beschreibung"))
        for r, v in rows
        if not v.get("Referenz") and (v.get("Beschreibung") or v.get("Kategorie"))
    ]


def check_art_casing(rows):
    return [
        (r, v.get("Art"))
        for r, v in rows
        if v.get("Art") and v["Art"] not in KNOWN_ART_VALUES
    ]


def check_whitespace(rows):
    issues = []
    for r, v in rows:
        for col in ("Beschreibung", "Referenz"):
            val = v.get(col)
            if isinstance(val, str) and val != val.strip():
                issues.append((r, col, val))
    return issues


def check_verfuegbarkeit(rows):
    issues = []
    for r, v in rows:
        val = v.get("Verfuegbarkeit")
        if val is None:
            continue
        if not VERFUEGBARKEIT_RE.match(str(val)):
            issues.append((r, val))
    return issues


def check_formula_tokens(rows, valid_refs, colname):
    issues = []
    for r, v in rows:
        val = v.get(colname)
        if not isinstance(val, str) or not val.strip():
            continue
        if val.strip().upper() == "FEHLT":
            continue
        stripped = QUOTE_RE.sub("", val)
        tokens = TOKEN_RE.findall(stripped)
        tokens = [
            t for t in tokens
            if t.upper() not in KEYWORDS and t not in UNIT_WORDS and t not in SELF_REF_TOKENS
        ]
        unknown = [t for t in tokens if t not in valid_refs]
        if unknown:
            issues.append((r, val, unknown))
    return issues


def main(path):
    wb = openpyxl.load_workbook(path, data_only=False)
    ws = wb["Werte"]
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    headers = [h for h in headers if h]
    rows, idx = load_rows(ws, headers)

    valid_refs = {v["Referenz"] for r, v in rows if isinstance(v.get("Referenz"), str)}

    problems = 0

    dupes = check_duplicate_refs(rows)
    if dupes:
        problems += len(dupes)
        print(f"\n[FEHLER] {len(dupes)} doppelte Referenz(en):")
        for d in dupes:
            print("  ", d)

    missing = check_missing_refs(rows)
    if missing:
        problems += len(missing)
        print(f"\n[FEHLER] {len(missing)} Zeile(n) mit Daten aber ohne Referenz:")
        for r, kat, besch in missing:
            print("  ", r, kat, besch)

    art_bad = check_art_casing(rows)
    if art_bad:
        problems += len(art_bad)
        print(f"\n[WARNUNG] {len(art_bad)} Zeile(n) mit unbekanntem/unsauberem Art-Wert:")
        for r, a in art_bad:
            print("  ", r, repr(a))

    ws_issues = check_whitespace(rows)
    if ws_issues:
        problems += len(ws_issues)
        print(f"\n[WARNUNG] {len(ws_issues)} Zelle(n) mit Leerzeichen am Rand:")
        for r, col, val in ws_issues:
            print("  ", r, col, repr(val))

    verf_issues = check_verfuegbarkeit(rows)
    if verf_issues:
        problems += len(verf_issues)
        print(f"\n[WARNUNG] {len(verf_issues)} Verfuegbarkeit-Wert(e) in unerwartetem Format:")
        for r, val in verf_issues:
            print("  ", r, repr(val))

    if "Formel" in headers:
        f_issues = check_formula_tokens(rows, valid_refs, "Formel")
        if f_issues:
            problems += len(f_issues)
            print(f"\n[FEHLER] {len(f_issues)} Formel(n) mit unbekannten Referenzen:")
            for r, val, unknown in f_issues:
                print("  ", r, val, "-> unbekannt:", unknown)

    if "Kosten" in headers:
        k_issues = check_formula_tokens(rows, valid_refs, "Kosten")
        if k_issues:
            problems += len(k_issues)
            print(f"\n[FEHLER] {len(k_issues)} Kosten-Formel(n) mit unbekannten Referenzen:")
            for r, val, unknown in k_issues:
                print("  ", r, val, "-> unbekannt:", unknown)

    print(f"\n{'='*50}")
    if problems == 0:
        print(f"OK - keine Probleme gefunden ({len(rows)} Zeilen geprueft).")
        return 0
    else:
        print(f"{problems} Problem(e) gefunden ({len(rows)} Zeilen geprueft).")
        return 1


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print('Aufruf: python scripts/validate_werte.py "werte 0.7-claude.xlsx"')
        sys.exit(1)
    sys.exit(main(sys.argv[1]))
