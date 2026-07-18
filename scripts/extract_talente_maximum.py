"""Extrahiert die "Maximum"-Talente (Fertigkeits-/Eigenschafts-/Attributs-/Kampffertigkeits-/
Zauberschulmaximum) aus einer ChatGPT-Wirkungsanalyse-Datei (Talente-Wirkung-chatgpt.xlsx) und
schreibt src/data/talenteMaximum.ts. Einmaliger Import, kein Teil der regulaeren
generate_data_ts.py-Pipeline (die Quelle ist eine externe Analysedatei, kein werte-xlsx-Sheet).

Nutzer 2026-07-18: "wir muessen die talente wirkungen in den build bringen" - von 148 Talenten
sind mittlerweile 109 ("Maximum"-Talente, hier verarbeitet) mechanisch portierbar, seit die
Basis-Maximalwerte feststehen (Grundfertigkeit/Sonderfertigkeit=12, Nahkampf/Fernkampf/WHK/
Spruchmagie=24, Attribute=7, KI/PSI=24 - vom Nutzer bestaetigt, siehe engine/fertigkeitenGrenzen.ts).
Der Rest (Modifikator-Talente, Kampfzeit-Regeln) wird separat behandelt.

MANUAL_MAXIMUM_OVERRIDES (siehe unten) deckt Talente ab, die strukturell Fertigkeitsmaximum-Boni
sind, aber in der Analysedatei unter einer anderen Wirkungsklasse liefen oder keine strukturierte
Zielreferenz hatten (u.a. KI-Meister, PSI Psinetik, Vorderlader Ladeschuetze, Charismatischer
Fuehrer) - siehe Kommentar dort fuer Details je Eintrag.
"""
import json
import sys
from pathlib import Path

import openpyxl

OUT_PATH = Path(__file__).parent.parent / "src" / "data" / "talenteMaximum.ts"
RULES_JSON_PATH = Path(__file__).parent.parent / "src" / "data" / "rules.json"

# "Wirkung 1 - Ziel"-Werte, die eine ganze Kategorie statt einer Einzelreferenz meinen
# (z.B. "Grundfertigkeiten Stufe 1-4": Grundfertigkeitenmaximum +2/+4/+6/+8 fuer ALLE Grundfertigkeiten).
KATEGORIE_GRUPPEN = {
    "Grundfertigkeiten": "Grundfertigkeit",
    "Wissens,- Handwerks,- Kulturfertigkeiten": "WHK",
}

# Talente, deren Wirkung strukturell ein Fertigkeitsmaximum-Bonus ist, aber in der Quelle unter
# einer anderen Wirkungsklasse liefen (daher vom Hauptfilter unten nicht erfasst) und/oder deren
# Zielreferenz dort nicht aufgeloest wurde. Nutzer 2026-07-18 (Re-Pruefung der Analysedatei):
# "Charismatischer Fuehrer" steht als "Modifikator - komplex" ("Erhoeht die Maxima Grundfertigkeiten
# Ueberreden & Ueberzeugen zusaetzlich um +6 Punkte"), obwohl es strukturell identisch zu den
# regulaer erfassten Fertigkeitsmaximum-Talenten ist. "Ueberreden" existiert nicht als eigene
# Referenz in rules.json (nur gr_ueberzeugen) - Nutzer entschied explizit: nur auf Ueberzeugen anwenden.
#
# "PSI Psinetik" (3 Stufen) und "Vorderlader Ladeschuetze" (2 Stufen) liefen unter "Komplexer
# Regeltext" / Portierungsstatus "manuell modellieren", sind aber ebenfalls klare Fertigkeitsmaximum-
# Boni (Freitext in der Impl.-Anweisung, keine strukturierte Zielreferenz). "KI-Meister" war bisher
# komplett uebersprungen (siehe Skip-Zweig unten) - Nutzer hat 2026-07-18 in einer zweiten Runde
# eine PSI/KI-Basiswert-Entscheidung nachgeholt (beide =24, siehe engine/fertigkeitenGrenzen.ts),
# damit sind jetzt auch KI-Meister und PSI Psinetik portierbar.
MANUAL_MAXIMUM_OVERRIDES = [
    {"talentReferenz": "talente_charismatischer_fuehrer", "zielReferenz": "gr_ueberzeugen", "bonus": 6},
    {"talentReferenz": "talente_psi_psinetik_stufe_1", "zielKategorie": "PSI", "bonus": 6},
    {"talentReferenz": "talente_psi_psinetik_stufe_2", "zielKategorie": "PSI", "bonus": 12},
    {"talentReferenz": "talente_psi_psinetik_stufe_3", "zielKategorie": "PSI", "bonus": 18},
    {"talentReferenz": "talente_vorderlader_ladeschuetze_stufe1", "zielReferenz": "sf_ladeschuetze_vorderlader", "bonus": 7},
    {"talentReferenz": "talente_vorderlader_ladeschuetze_stufe2", "zielReferenz": "sf_ladeschuetze_vorderlader", "bonus": 15},
    {"talentReferenz": "talente_ki_meister", "zielKategorie": "KI", "bonus": 18},
]


def read_rows(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Talente"]
    headers = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        headers.append(v.strip() if v else f"col{c}")
    rows = []
    for r in range(2, ws.max_row + 1):
        row = {}
        for c, h in enumerate(headers, start=1):
            v = ws.cell(row=r, column=c).value
            if v is not None:
                row[h] = v
        if row:
            rows.append(row)
    return rows


def extract(rows):
    maxklassen = {
        "Fertigkeitsmaximum", "Eigenschaftsmaximum", "Attributsmaximum",
        "Kampffertigkeitsmaximum", "Zauberschulmaximum",
    }
    out = []
    skipped = []

    for r in rows:
        klasse = r.get("Wirkungsklasse")
        status = r.get("Portierungsstatus") or ""
        if klasse not in maxklassen and klasse != "KI-Fähigkeitsmaximum":
            continue
        if not status.startswith("strukturiert"):
            continue

        refs = [x.strip() for x in (r.get("Werte-Referenz(en)") or "").split("\n") if x.strip()]
        wert = r.get("Wirkung 1 – Wert")
        zielref = r.get("Wirkung 1 – Zielreferenz")
        ziel = r.get("Wirkung 1 – Ziel")

        if klasse == "KI-Fähigkeitsmaximum":
            # Zielreferenz in der Quelle ("ki_ki_faehigkeiten") existiert nicht - das eigentliche
            # Ziel ist "ALLE ki_*-Referenzen", siehe MANUAL_MAXIMUM_OVERRIDES (zielKategorie "KI").
            skipped.append((r.get("Name"), "wird stattdessen ueber MANUAL_MAXIMUM_OVERRIDES als zielKategorie=KI abgedeckt"))
            continue

        if klasse == "Zauberschulmaximum":
            # 1:n-Zeile (eine ChatGPT-Zeile deckt alle 12 Zauberschulen ab) - Schule aus dem
            # Referenz-Suffix ableiten (z.B. talente_magus_stufe_1_feuerbeschwoerungs_magus).
            for ref in refs:
                suffix = ref.split("_stufe_")[-1]  # "1_feuerbeschwoerungs_magus"
                parts = suffix.split("_")
                schule_raw = parts[1] if len(parts) > 1 else None
                if not schule_raw:
                    skipped.append((ref, "Schule nicht aus Referenz ableitbar"))
                    continue
                schule = schule_raw[:-1] if schule_raw.endswith("s") else schule_raw
                out.append({"talentReferenz": ref, "zielPraefix": f"spruchmagie_{schule}_", "bonus": wert})
            continue

        if ziel in KATEGORIE_GRUPPEN:
            for ref in refs:
                out.append({"talentReferenz": ref, "zielKategorie": KATEGORIE_GRUPPEN[ziel], "bonus": wert})
            continue

        if not zielref:
            skipped.append((r.get("Name"), "keine Zielreferenz, kein Gruppenziel erkannt"))
            continue

        for ref in refs:
            out.append({"talentReferenz": ref, "zielReferenz": zielref, "bonus": wert})

    out.extend(MANUAL_MAXIMUM_OVERRIDES)
    return out, skipped


def validate(entries, rules_json_path):
    rules = json.loads(rules_json_path.read_text(encoding="utf-8"))
    referenz_set = {r["referenz"] for r in rules}
    problems = []
    for e in entries:
        if e["talentReferenz"] not in referenz_set:
            problems.append(f"talentReferenz '{e['talentReferenz']}' existiert nicht in rules.json")
        if e.get("zielReferenz") and e["zielReferenz"] not in referenz_set:
            problems.append(f"zielReferenz '{e['zielReferenz']}' existiert nicht in rules.json")
    return problems


def write_ts(entries, out_path):
    entries = sorted(entries, key=lambda e: e["talentReferenz"])
    lines = [
        '// GENERIERT von scripts/extract_talente_maximum.py aus Talente-Wirkung-chatgpt.xlsx',
        "// (ChatGPT-Analyse der Talente-Beschreibungen, abgeglichen mit rules.json) - nicht von",
        "// Hand bearbeiten, stattdessen das Skript erneut laufen lassen. Nutzer 2026-07-18:",
        "// Basiswerte fuer Fertigkeitsmaximum bestaetigt: Grundfertigkeit/Sonderfertigkeit=12,",
        "// Nahkampf/Fernkampf/WHK/Spruchmagie=24, Attribute=7 (siehe engine/fertigkeitenGrenzen.ts).",
        "// Jeder Eintrag ist ein TALENT, das bei Auswahl das Maximum EINER Referenz (zielReferenz),",
        '// ALLER Referenzen einer Kategorie (zielKategorie, z.B. "alle Grundfertigkeiten") oder',
        '// ALLER Referenzen mit gemeinsamem Praefix (zielPraefix, z.B. "spruchmagie_feuerbeschwoerung_"',
        '// fuer eine Zauberschule) um den angegebenen Betrag erhoeht. Einige Eintraege sind manuelle',
        '// Overrides (siehe MANUAL_MAXIMUM_OVERRIDES im Skript) - liefen in der Quelle unter einer',
        '// anderen Wirkungsklasse oder hatten keine strukturierte Zielreferenz, wirken aber strukturell',
        '// identisch zu den regulaer erfassten Fertigkeitsmaximum-Talenten: Charismatischer Fuehrer',
        '// (nur gr_ueberzeugen, "Ueberreden" existiert nicht), PSI Psinetik + KI-Meister (KI/PSI-',
        '// Basiswert=24, Nutzer 2026-07-18), Vorderlader Ladeschuetze.',
        "",
        "export interface TalentMaximumBonus {",
        "  talentReferenz: string;",
        "  zielReferenz?: string;",
        "  zielKategorie?: string;",
        "  zielPraefix?: string;",
        "  bonus: number;",
        "}",
        "",
        f"export const TALENT_MAXIMUM_BONUSES: TalentMaximumBonus[] = {json.dumps(entries, ensure_ascii=False, indent=2)};",
        "",
    ]
    out_path.write_text("\n".join(lines), encoding="utf-8")


def main(xlsx_path):
    rows = read_rows(xlsx_path)
    entries, skipped = extract(rows)
    problems = validate(entries, RULES_JSON_PATH)
    if problems:
        raise SystemExit("Validierungsfehler:\n" + "\n".join(problems))
    write_ts(entries, OUT_PATH)
    print(f"{OUT_PATH}: {len(entries)} Talent-Maximum-Boni geschrieben.")
    if skipped:
        print(f"{len(skipped)} uebersprungen:")
        for name, reason in skipped:
            print(f"  - {name}: {reason}")


if __name__ == "__main__":
    xlsx = sys.argv[1] if len(sys.argv) > 1 else "Talente-Wirkung-chatgpt.xlsx"
    main(xlsx)
