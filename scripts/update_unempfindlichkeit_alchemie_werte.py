import sys
from pathlib import Path


ROOT = Path(r"E:\Das Western Rollenspiel\LLM")
sys.path.insert(0, str(ROOT / ".python-deps"))

from openpyxl import load_workbook


WERTE = ROOT / "werte 0.8-claude.xlsx"
REFERENCE = "vn_unempfindlichkeit_gegen_alchemie"
EFFECT = "Das Maximum der Sonderfertigkeit Alchemieresistenz erhöht sich um 12."


def main():
    workbook = load_workbook(WERTE, read_only=False, data_only=False)
    worksheet = workbook["Werte"]
    headers = {cell.value: cell.column for cell in worksheet[1] if cell.value}
    row = next(
        row
        for row in range(2, worksheet.max_row + 1)
        if worksheet.cell(row, headers["Referenz"]).value == REFERENCE
    )
    worksheet.cell(row, headers["Info"], "Historische Bezeichnung 'Giftresistenz' ist veraltet; gültig ist Alchemieresistenz.")
    worksheet.cell(row, headers["Wirkung"], EFFECT)
    workbook.save(WERTE)
    print(row, REFERENCE, EFFECT)


if __name__ == "__main__":
    main()
