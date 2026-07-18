import sys
from pathlib import Path


ROOT = Path(r"E:\Das Western Rollenspiel\LLM")
sys.path.insert(0, str(ROOT / ".python-deps"))

from openpyxl import load_workbook


WERTE = ROOT / "werte 0.8-claude.xlsx"
ELEMENTS = {
    "erde": "Erde",
    "feuer": "Feuer",
    "luft": "Luft",
    "magie": "Magie",
    "wasser": "Wasser",
}


def main():
    workbook = load_workbook(WERTE, read_only=False, data_only=False)
    worksheet = workbook["Werte"]
    headers = {cell.value: cell.column for cell in worksheet[1] if cell.value}
    rows_by_reference = {
        worksheet.cell(row, headers["Referenz"]).value: row
        for row in range(2, worksheet.max_row + 1)
    }
    changed = []

    for slug, element in ELEMENTS.items():
        reference = f"vn_immunitaet_gegen_{slug}"
        row = rows_by_reference.get(reference, worksheet.max_row + 1)
        effect = (
            f"Nur Elementare: Der Charakter erleidet keinen Schaden durch das eigene Element {element}. "
            "Dieser Wert ist im Spielermodus nie wählbar. Bei der Charaktererschaffung im "
            "Meistermodus ist für Elementare die zum eigenen Element passende Immunität verpflichtend."
        )
        values = {
            "Referenz": reference,
            "Kategorie": "Vor- und Nachteile",
            "Beschreibung": f"Immunität gegen {element}",
            "Info": (
                "Nur Meistercharaktere; im Spielermodus nie wählbar; bei der Charaktererschaffung "
                "im Meistermodus für Elementare passend zum eigenen Element verpflichtend."
            ),
            "Parent": "Elementar",
            "Art": "Auswahl",
            "Flag": "MEISTER_MODUS_PFLICHT",
            "Kosten": 0,
            "Verfuegbarkeit": "M",
            "Wirkung": effect,
        }
        for header, value in values.items():
            worksheet.cell(row, headers[header], value)
        changed.append((row, reference, effect))

    workbook.save(WERTE)
    print(f"updated={len(changed)}")
    for row, reference, effect in changed:
        print(row, reference, effect)


if __name__ == "__main__":
    main()
