import sys
from pathlib import Path


ROOT = Path(r"E:\Das Western Rollenspiel\LLM")
sys.path.insert(0, str(ROOT / ".python-deps"))

from openpyxl import load_workbook


WERTE = ROOT / "werte 0.8-claude.xlsx"
SCHOOLS = {
    "antimagie": "Antimagie",
    "beherrschung": "Beherrschung",
    "erdbeschwoerung": "Erdbeschwörung",
    "feuerbeschwoerung": "Feuerbeschwörung",
    "heilung": "Heilung",
    "hellsicht": "Hellsicht",
    "illusion": "Illusion",
    "luftbeschwoerung": "Luftbeschwörung",
    "magiebeschwoerung": "Magiebeschwörung",
    "veraenderung": "Veränderung",
    "verzauberung": "Verzauberung",
    "wasserbeschwoerung": "Wasserbeschwörung",
}
STAGES = {
    1: {"roman": "I", "availability": 3, "increase": "50 %", "factor": "1,5"},
    2: {"roman": "II", "availability": 5, "increase": "100 %", "factor": "2"},
    3: {"roman": "III", "availability": 7, "increase": "150 %", "factor": "2,5"},
}


def main():
    workbook = load_workbook(WERTE, read_only=False, data_only=False)
    worksheet = workbook["Werte"]
    headers = {cell.value: cell.column for cell in worksheet[1] if cell.value}
    rows_by_reference = {
        worksheet.cell(row, headers["Referenz"]).value: row
        for row in range(2, worksheet.max_row + 1)
    }
    changed = []

    for stage, stage_data in STAGES.items():
        for slug, school in SCHOOLS.items():
            reference = f"vn_meisterlicher_{slug}_magier_{stage}"
            row = rows_by_reference.get(reference, worksheet.max_row + 1)
            previous_reference = (
                f"vn_meisterlicher_{slug}_magier_{stage - 1}" if stage > 1 else None
            )
            prerequisite = (
                f"Voraussetzung: Meisterlicher {school}-Magier {STAGES[stage - 1]['roman']}. "
                if stage > 1
                else "Nur Nasus Nasus, nur Magier: "
            )
            effect = (
                f"{prerequisite}Die Wirkung von Zaubern der Schule {school} erhöht sich um "
                f"{stage_data['increase']} (Faktor {stage_data['factor']}). Nicht kumulativ mit den "
                "anderen Stufen dieses Vorteils. Eine höhere Stufe ersetzt die vorherige Stufe."
            )
            info = (
                f"Gilt ausschließlich für die Schule {school}. Nicht kumulativ; eine höhere Stufe "
                "ersetzt die vorherige."
            )
            flag = "NICHT_KUMULATIV_MEISTERLICHER_SCHULE"
            if previous_reference:
                info += f" Voraussetzung: {previous_reference}."
                flag += f";VORAUSSETZUNG={previous_reference}"
            values = {
                "Referenz": reference,
                "Kategorie": "Vor- und Nachteile",
                "Beschreibung": f"Meisterlicher {school}-Magier {stage_data['roman']}",
                "Info": info,
                "Parent": "Magier",
                "Art": "Auswahl",
                "Flag": flag,
                "Kosten": 999,
                "Verfuegbarkeit": stage_data["availability"],
                "Wirkung": effect,
            }
            for header, value in values.items():
                worksheet.cell(row, headers[header], value)
            rows_by_reference[reference] = row
            changed.append((row, reference, school))

    workbook.save(WERTE)
    print(f"updated={len(changed)}")
    for row, reference, school in changed:
        print(row, reference, school)


if __name__ == "__main__":
    main()
