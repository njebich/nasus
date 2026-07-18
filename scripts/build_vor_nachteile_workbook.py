import re
import sys
import unicodedata
from collections import defaultdict
from copy import copy
from pathlib import Path

ROOT = Path(r"E:\Das Western Rollenspiel\LLM")
sys.path.insert(0, str(ROOT / ".python-deps"))

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


SOURCE = Path(
    r"E:\Das Western Rollenspiel\_Aktuelle Daten\Nasus\Nasus Nasus das Western Rollenspiel"
    r"\---==Veraltet==---\Vor- und- Nachteile v1.1.xlsx"
)
WERTE = ROOT / "werte 0.8-claude.xlsx"
OUTPUT_DIR = ROOT / "outputs" / "vor-und-nachteile-v1.1-chatgpt"
OUTPUT = OUTPUT_DIR / "vor-und-nachteile-v1.1-chatgpt.xlsx"


def normalize(value):
    value = str(value or "").lower().replace("ß", "ss")
    value = unicodedata.normalize("NFKD", value)
    value = "".join(char for char in value if not unicodedata.combining(char))
    value = re.sub(r"\[[^\]]+\]", " x ", value)
    value = value.replace("–", "-")
    value = re.sub(r"[^a-z0-9]+", " ", value)
    return " ".join(value.split())


def without_namespace(value):
    value = re.sub(r"^(Aussehen|Fähigkeit|Krankheit|Schlaf|SF|Sicht)\s*:\s*", "", str(value or ""), flags=re.I)
    return normalize(value)


def display_value(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def availability_status(source_value, matched_values):
    source_text = display_value(source_value)
    value_texts = sorted({display_value(value) for value in matched_values if value is not None})
    if not source_text:
        return "Quelle leer"
    if not value_texts:
        return "Werte leer"
    if value_texts == [source_text]:
        return "OK"
    if source_text in value_texts and len(value_texts) > 1:
        return "Teilweise OK"
    return "Abweichung"


source_wb = load_workbook(SOURCE, read_only=False, data_only=False)
source_ws = source_wb["Komplett"]
source_rows = []
for row_number in range(2, source_ws.max_row + 1):
    name = source_ws.cell(row_number, 2).value
    effect = source_ws.cell(row_number, 3).value
    when = source_ws.cell(row_number, 4).value
    if name == "Meisterliche [Optionale Fertigkeit]":
        name = "Fähigkeit: Meisterliche WHK"
        effect = str(effect or "").replace(
            "einer ausgewählten  spezialisierbaren Optionalen Fertigkeit",
            "einer ausgewählten spezialisierbaren WHK",
        ).replace(
            "jede Optionale Fähigkeit",
            "jede WHK",
        )
    if name:
        source_rows.append((row_number, name, effect, when))

werte_wb = load_workbook(WERTE, read_only=True, data_only=False)
werte_ws = werte_wb["Werte"]
headers = [cell.value for cell in werte_ws[1]]
header_index = {header: index for index, header in enumerate(headers) if header}
werte_rows = []
talent_rows = []
for row_number, row in enumerate(werte_ws.iter_rows(min_row=2, values_only=True), start=2):
    category = row[header_index["Kategorie"]]
    if category not in ("Vor- und Nachteile", "Talente"):
        continue
    item = {
        "row": row_number,
        "reference": row[header_index["Referenz"]],
        "name": row[header_index["Beschreibung"]],
        "parent": row[header_index["Parent"]],
        "cost": row[header_index["Kosten"]],
        "availability": row[header_index["Verfuegbarkeit"]],
    }
    if category == "Vor- und Nachteile":
        werte_rows.append(item)
    else:
        talent_rows.append(item)

by_exact = defaultdict(list)
by_without_namespace = defaultdict(list)
for item in werte_rows:
    by_exact[normalize(item["name"])].append(item)
    by_without_namespace[without_namespace(item["name"])].append(item)


def select_references(pattern):
    regex = re.compile(pattern)
    return [item for item in werte_rows if regex.search(item["reference"] or "")]


manual_groups = {
    "Abneigung gegen [Element- Kampfmagie/ Beschwörung- Schule]": select_references(
        r"^vn_abneigung_gegen_(erd|feuer|luft|magie|wasser)beschwoerung$"
    ),
    "Abneigung gegen [Heilung, Verzauberung o. Beherrschung]": select_references(
        r"^vn_abneigung_gegen_(heilung|verzauberung|beherrschung)$"
    ),
    "Anfälligkeit gegen [Element]": select_references(
        r"^vn_anfaelligkeit_gegen_(erd|feuer|luft|magie|wasser)beschwoerung_1$"
    ),
    "Anfälligkeit gegen [Schule]": select_references(
        r"^vn_anfaelligkeit_gegen_beherrschung_1$"
    ),
    "Anfälligkeit gegen Alchemie 1 - 30": select_references(
        r"^vn_(leicht|mittel|schwer|sehr_schwer|extrem)_anfaelligkeit_gegen_alchemie$"
    ),
    "Angst vor Magie 1-30": select_references(
        r"^vn_angst_magie_(5|10|15|20|25|30)$"
    ),
    "Angst vor X": select_references(
        r"^vn_angst_(verlusten_humanoide|arachnophobie|bestimmtes_domestiziertes_tier|hundeartige|hunger|laecherlichkeit|magische_gegenstaende|magische_wesen|meer|raubkatzen|schmutz|tiere|untote|voegel|wasser|wilde_tiere|xenophobie)_(5|10|15|20|25|30)$"
    ),
    "Aura verhüllen": select_references(r"^vn_aura_verhuellen_i$"),
    "Eigenschafts-Maximum senken 1-9": select_references(r"^vn_schlechte_eigenschaft_"),
    "Empfindlichkeit [Element]": select_references(
        r"^vn_anfaelligkeit_gegen_(erd|feuer|luft|magie|wasser)beschwoerung_2$"
    ),
    "Empfindlichkeit [Schule]": select_references(
        r"^vn_anfaelligkeit_gegen_beherrschung_2$"
    ),
    "Empfindlichkeit gegen profane Waffen": select_references(r"^vn_anfaelligkeit_gegen_profane_waffen$"),
    "Immunität gegen [Element]": select_references(
        r"^vn_immunitaet_gegen_(erde|feuer|luft|magie|wasser)$"
    ),
    "Fähigkeit: Meisterliche WHK": select_references(r"^vn_faehigkeit_meisterliche_WHK$"),
    "Meisterliche [Grundfertigkeit]": select_references(
        r"^vn_faehigkeit_meisterliche_grundfertigkeit_.+$"
    ),
    "Meisterlicher [Schule] – Magier I": select_references(
        r"^vn_meisterlicher_.+_magier_1$"
    ),
    "Meisterlicher [Schule] – Magier II": select_references(
        r"^vn_meisterlicher_.+_magier_2$"
    ),
    "Meisterlicher [Schule] – Magier III": select_references(
        r"^vn_meisterlicher_.+_magier_3$"
    ),
    "Migräne 1 – 30": select_references(
        r"^vn_krankheit_(leicht|mittel|schwer|sehr_schwer|extrem)_migraene$"
    ),
    "Resistenz gegen [Element]": select_references(
        r"^vn_resistenz_gegen_(erde|feuer|luft|magie|wasser)$"
    ),
    "Resistenz gegen Alchemie": select_references(
        r"^vn_unempfindlichkeit_gegen_alchemie$"
    ),
    "Schlaf": select_references(r"^vn_schlaf_normal$"),
    "Unempfindlichkeit [Waffentalent]": select_references(
        r"^vn_unempfindlichkeit_gegen_profane_waffen$"
    ),
    "Xenophobie": select_references(
        r"^vn_angst_xenophobie_(5|10|15|20|25|30)$"
    ),
}


def find_matches(name):
    key = normalize(name)
    if name in manual_groups and manual_groups[name]:
        matches = manual_groups[name]
        status = "1:n – Werte aufgeteilt" if len(matches) > 1 else "Vorschlag 1:1"
        confidence = 0.95
        note = "Historischer Sammel-/Platzhaltereintrag; aktuelle Werte sind in Einzelwerte aufgeteilt."
        if name == "Fähigkeit: Meisterliche WHK":
            note = "Bestätigte Umbenennung von 'Meisterliche [Optionale Fertigkeit]' zu 'Fähigkeit: Meisterliche WHK'."
        return matches, status, confidence, note
    exact = by_exact.get(key, [])
    if exact:
        return exact, "exakt", 1.0, "Normalisierter Name stimmt überein."
    aliases = by_without_namespace.get(key, [])
    if aliases:
        return aliases, "Alias – Kategoriepräfix", 0.98, "Werte-Name ergänzt nur ein Kategoriepräfix."
    return [], "nicht gefunden", 0.0, "Kein belastbarer Namens-/Alias-Treffer in 'Werte'."


current_list_names = [
    "Schlaf: Insomnia", "Schlaf: Leichter Schlaf", "Schlaf: Normal", "Schlaf: Tiefer Schlaf",
    "Aussehen: Herausragendes Aussehen", "Aussehen: Gutes Aussehen", "Aussehen: Normal",
    "Aussehen: Allerweltsgesicht", "Aussehen: Hässlichkeit", "Aussehen: Widerwärtiges Aussehen",
    "Sicht: Astrales Auge", "Sicht: Astrales Auge II", "Sicht: Blinder Kampf", "Sicht: Blindheit",
    "Sicht: Dämmerungssicht", "Sicht: Einäugig", "Sicht: Infrarotsicht", "Sicht: Nachtsicht",
    "Sicht: Weitsichtigkeit 1-30", "Sicht: Kurzsichtigkeit 1-30",
    "Krankheit: Fettleibig", "Krankheit: Migräne 1-30", "SF: Krankheitsresistenz",
    "SF: Unempfindlichkeit gegen Alchemie N'N",
    "Resistenz gegen Magiebeschwörung", "Resistenz gegen Luftbeschwörung",
    "Resistenz gegen Feuerbeschwörung", "Resistenz gegen Wasserbeschwörung",
    "Resistenz gegen Erdbeschwörung", "Resistenz gegen Beherrschung",
    "Resistenz gegen magische Heilung", "Resistenz gegen Verzauberung", "Sprachtalent", "Weltgewandt",
    "Schlechte Eigenschaft Ausstrahlung", "Schlechte Eigenschaft Intelligenz", "Schlechte Eigenschaft Mut",
    "Schlechte Eigenschaft Sinnesschärfe", "Schlechte Eigenschaft Willenskraft",
    "Schlechte Eigenschaft Athletik", "Schlechte Eigenschaft Geschicklichkeit",
    "Schlechte Eigenschaft Konstitution", "Schlechte Eigenschaft Schnelligkeit", "Schlechte Eigenschaft Stärke",
    "Angst: Magie 1-30", "Angst: Verluste (Humanoide) 1-30", "Angst: Arachnophobie 1-30",
    "Angst: Bestimmtes domestiziertes Tier 1-30", "Angst: Hundeartige 1-30", "Angst: Hunger 1-30",
    "Angst: Lächerlichkeit 1-30", "Angst: Magische Gegenstände 1-30", "Angst: Magische Wesen 1-30",
    "Angst: Meer 1-30", "Angst: Raubkatzen 1-30", "Angst: Schmutz 1-30", "Angst: Tiere 1-30",
    "Angst: Untote 1-30", "Angst: Vögel 1-30", "Angst: Wasser 1-30", "Angst: Wilde Tiere 1-30",
    "Angst: Xenophobie 1-30", "Fähigkeit: Astrale Abschottung", "Fähigkeit: Aura verhüllen I",
    "Fähigkeit: Aura verhüllen II", "Fähigkeit: Fliegen", "Fähigkeit: Meisterliche [Grundfertigkeit]",
    "Fähigkeit: Meisterliche WHK", "Fähigkeit: Natürliche Regeneration I",
    "Fähigkeit: Natürliche Regeneration II", "Fähigkeit: Natürliche Regeneration III",
    "Fähigkeit: Schweben", "Fähigkeit: Teleportation", "Fähigkeit: Träge Reflexe",
    "Fähigkeit: Wasserwandeln", "Fähigkeit: Unempfindlichkeit gegen profane Waffen",
    "Anfälligkeit gegen Erdbeschwörung 1", "Anfälligkeit gegen Erdbeschwörung 2",
    "Anfälligkeit gegen Feuerbeschwörung 1", "Anfälligkeit gegen Feuerbeschwörung 2",
    "Anfälligkeit gegen Luftbeschwörung 1", "Anfälligkeit gegen Luftbeschwörung 2",
    "Anfälligkeit gegen Magiebeschwörung 1", "Anfälligkeit gegen Magiebeschwörung 2",
    "Anfälligkeit gegen Wasserbeschwörung 1", "Anfälligkeit gegen Wasserbeschwörung 2",
    "Anfälligkeit gegen Beherrschung 1", "Anfälligkeit gegen Beherrschung 2",
    "Anfälligkeit gegen profane Waffen", "Anfälligkeit gegen Alchemie 1-30",
]


def list_normalize(value):
    value = re.sub(r"\bgg\.?", "gegen", str(value or ""), flags=re.I)
    value = value.replace("Magische gg.stände", "Magische Gegenstände")
    return normalize(value)


current_by_name = defaultdict(list)
current_by_without_namespace = defaultdict(list)
for current_name in current_list_names:
    current_by_name[list_normalize(current_name)].append(current_name)
    expanded = re.sub(r"\bgg\.?", "gegen", current_name, flags=re.I).replace(
        "Magische gg.stände", "Magische Gegenstände"
    )
    current_by_without_namespace[without_namespace(expanded)].append(current_name)

fear_names = [name for name in current_list_names if name.startswith("Angst:") and "Magie 1-30" not in name]
element_susceptibilities = [name for name in current_list_names if name.startswith("Anfälligkeit gegen") and "beschwörung" in name.lower()]
bad_properties = [name for name in current_list_names if name.startswith("Schlechte Eigenschaft")]
element_resistances = [name for name in current_list_names if name.startswith("Resistenz gegen") and "beschwörung" in name.lower()]

current_manual = {
    "Anfälligkeit gegen [Element]": [name for name in element_susceptibilities if name.endswith(" 1")],
    "Anfälligkeit gegen [Schule]": [name for name in current_list_names if name.startswith("Anfälligkeit gegen Beherrschung") and name.endswith(" 1")],
    "Anfälligkeit gegen Alchemie 1 - 30": ["Anfälligkeit gegen Alchemie 1-30"],
    "Angst vor X": fear_names,
    "Angst vor Magie 1-30": ["Angst: Magie 1-30"],
    "Astrale Abschottung": ["Fähigkeit: Astrale Abschottung"],
    "Aura verhüllen": ["Fähigkeit: Aura verhüllen I"],
    "Aura verhüllen II": ["Fähigkeit: Aura verhüllen II"],
    "Eigenschafts-Maximum senken 1-9": bad_properties,
    "Empfindlichkeit [Element]": [name for name in element_susceptibilities if name.endswith(" 2")],
    "Empfindlichkeit [Schule]": [name for name in current_list_names if name.startswith("Anfälligkeit gegen Beherrschung") and name.endswith(" 2")],
    "Empfindlichkeit gegen profane Waffen": ["Anfälligkeit gegen profane Waffen"],
    "Empfindlichkeit gegen Alchemie": ["Anfälligkeit gegen Alchemie 1-30"],
    "Fähigkeit: Meisterliche WHK": ["Fähigkeit: Meisterliche WHK"],
    "Meisterliche [Grundfertigkeit]": ["Fähigkeit: Meisterliche [Grundfertigkeit]"],
    "Migräne 1 – 30": ["Krankheit: Migräne 1-30"],
    "Resistenz gegen [Element]": element_resistances,
    "Schlaf": ["Schlaf: Normal"],
    "Unempfindlichkeit gegen Alchemie": ["SF: Unempfindlichkeit gegen Alchemie N'N"],
    "Unempfindlichkeit gegen profane Waffen": ["Fähigkeit: Unempfindlichkeit gegen profane Waffen"],
    "Xenophobie": ["Angst: Xenophobie 1-30"],
}


def find_current_list_matches(name):
    if name in current_manual:
        matches = current_manual[name]
        status = "1:n – aktuelle Liste" if len(matches) > 1 else "Alias – aktuelle Liste"
        return matches, status
    key = list_normalize(name)
    direct = current_by_name.get(key, [])
    if direct:
        return direct, "exakt – aktuelle Liste"
    expanded = re.sub(r"\bgg\.?", "gegen", name, flags=re.I).replace(
        "Magische gg.stände", "Magische Gegenstände"
    )
    alias = current_by_without_namespace.get(without_namespace(expanded), [])
    if alias:
        return alias, "Alias – aktuelle Liste"
    return [], "nicht in aktueller Liste"


def select_talents(pattern):
    regex = re.compile(pattern)
    return [item for item in talent_rows if regex.search(item["reference"] or "")]


talent_manual = {
    "Attributs-Maximum erhöhen +1": select_talents(r"^talente_attributs_maximum_erhoehen__"),
    "Beidhändiger Kampf": select_talents(r"^talente_kampf_mit_zwei_waffen_stufe_"),
    "Charismatischer Führer": select_talents(r"^talente_charismatischer_fuehrer$"),
    "Doppelschuss": select_talents(r"^talente_mehrfachschuss_stufe_1_doppelschuss$"),
    "Eigenschaftsmaximum erhöhen +1": select_talents(r"^talente_eigenschaften_erhoehen_"),
    "Entwaffnen": select_talents(r"^talente_entwaffnen$"),
    "Geübter Zauberer I [Schule]": select_talents(r"^talente_spruchmagie_stufe_2_zaubern$"),
    "Geübter Zauberer II [Schule]": select_talents(r"^talente_spruchmagie_stufe_3_zaubern$"),
    "Meuchler": select_talents(r"^talente_meuchler$"),
    "Schnell Zaubern I": select_talents(r"^talente_schnell_zaubern_stufe_1$"),
    "Schnell Zaubern II": select_talents(r"^talente_schnell_zaubern_stufe_2$"),
    "Schnell Zaubern III": select_talents(r"^talente_schnell_zaubern_stufe_3$"),
    "Schwerer Hieb": select_talents(r"^talente_wuchtschlag$"),
    "Tripelschuss": select_talents(r"^talente_mehrfachschuss_stufe_2_tripelschuss$"),
    "Umwerfen": select_talents(r"^talente_mit_schild_umwerfen$"),
    "Zäher Bursche I": select_talents(r"^talente_zaeher_bursche_stufe_1$"),
    "Zäher Bursche II": select_talents(r"^talente_zaeher_bursche_stufe_2$"),
    "Zäher Bursche III": select_talents(r"^talente_zaeher_bursche_stufe_3$"),
}


def classify_source(name):
    list_matches, list_status = find_current_list_matches(name)
    talent_matches = talent_manual.get(name, [])
    if list_matches:
        classification = "Aktueller Vor-/Nachteil"
        recommendation = "In der Vor-/Nachteil-Arbeitsmenge behalten."
    elif talent_matches:
        classification = "Talent in alter V/N-Liste"
        recommendation = "Aus der Vor-/Nachteil-Arbeitsmenge entfernen; im Talentmodell weiterführen."
    else:
        classification = "Nicht in aktueller Liste"
        recommendation = "Nicht automatisch übernehmen; als veraltet/entfallen prüfen."
    return list_matches, list_status, talent_matches, classification, recommendation


all_source_rows = list(source_rows)
source_rows = []
removed_rows = []
for record in all_source_rows:
    source_row, name, effect, when = record
    classification = classify_source(name)[3]
    if name in {"Affinität zu [Element]", "Affinität zu [Schule]"}:
        removed_rows.append((*record, "Über Talente abgedeckt"))
    elif name == "Empfindlichkeit [Waffentalent]":
        removed_rows.append((*record, "Nicht Teil der bestätigten Anfälligkeitsliste"))
    elif name == "Empfindlichkeit gegen Alchemie":
        removed_rows.append((*record, "Nach Prüfung entfernt"))
    elif name == "Unempfindlichkeit [Waffentalent]":
        removed_rows.append((*record, "Generischer Waffentalent-Eintrag entfernt"))
    elif name == "Unsichtbarkeit":
        removed_rows.append((*record, "Nur NR"))
    elif classification == "Talent in alter V/N-Liste":
        removed_rows.append((*record, "Talent"))
    elif re.search(r"(?i)\bcyberimplant", str(name or "")):
        removed_rows.append((*record, "Cyberimplantat/Cyberimplantiert"))
    elif re.search(r"(?i)\bnur\s+(noserun|nr)\b", f"{name or ''} {effect or ''}"):
        removed_rows.append((*record, "Nur Noserun/Nur NR"))
    else:
        source_rows.append(record)


output_wb = Workbook()
ws = output_wb.active
ws.title = "Vor- und Nachteile"
ws.sheet_view.showGridLines = False
ws.freeze_panes = "E2"

output_headers = [
    "Quelle-Zeile",
    "Name",
    "Wirkung",
    "Wann wählbar",
    "Werte-Referenz(en)",
    "Referenzanzahl",
    "Zuordnungsstatus",
    "Zuordnungssicherheit",
    "Werte-Zeile(n)",
    "Werte-Name(n)",
    "Werte-Kosten",
    "Werte-Verfügbarkeit",
    "Verfügbarkeitsabgleich",
    "Wirkungsklasse",
    "Wirkung – Zieltyp",
    "Wirkung – Ziel",
    "Wirkung – Zielreferenz",
    "Wirkung – Imperativ",
    "Wirkung – Wert/Formel",
    "Wirkung – Einheit",
    "Bedingung / Auslöser",
    "Stufen-/Kumulierungslogik",
    "Imperative Implementierungsanweisung",
    "Portierungsstatus",
    "Prüfhinweis",
    "Abgleich aktuelle Liste",
    "Aktuelle Listen-Treffer",
    "Einordnung",
    "Talent-Referenz(en)",
    "Bereinigungsempfehlung",
]
ws.append(output_headers)


def compact_join(values, limit=8):
    values = [str(value) for value in values if value not in (None, "")]
    if len(values) <= limit:
        return " | ".join(values)
    return " | ".join(values[:limit]) + f" | … (+{len(values) - limit}; siehe Werte-Zuordnungen)"


mapping_records = []
deferred_reviews = set()
effect_overrides = {
    "Resistenz gegen Alchemie": "SF Alchemieresistenz Max +12",
}

for source_row, name, effect, when in source_rows:
    effect = effect_overrides.get(name, effect)
    matches, status, confidence, note = find_matches(name)
    list_matches, list_status, talent_matches, classification, recommendation = classify_source(name)
    availability_values = [item["availability"] for item in matches]
    for item in matches:
        mapping_records.append(
            [
                source_row,
                name,
                item["row"],
                item["reference"],
                item["name"],
                item["cost"],
                item["availability"],
                status,
                confidence,
            ]
        )
    row = [
        source_row,
        name,
        effect,
        when,
        compact_join(item["reference"] for item in matches),
        len(matches),
        status,
        confidence,
        compact_join(str(item["row"]) for item in matches),
        compact_join(item["name"] for item in matches),
        compact_join(display_value(item["cost"]) for item in matches),
        compact_join(display_value(item["availability"]) for item in matches),
        availability_status(when, availability_values),
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "zurückgestellt" if name in deferred_reviews else "offen",
        "Entscheidung auf Wunsch des Nutzers zurückgestellt. " + note if name in deferred_reviews else note,
        list_status,
        compact_join(list_matches),
        classification,
        compact_join(item["reference"] for item in talent_matches),
        recommendation,
    ]
    ws.append(row)

source_wb.close()
werte_wb.close()

navy = "1F4E78"
blue = "D9EAF7"
source_fill = "E2F0D9"
audit_fill = "DDEBF7"
work_fill = "FFF2CC"
classification_fill = "EADCF8"
white = "FFFFFF"
grid = "B7C9D6"
thin = Side(style="thin", color=grid)
medium = Side(style="medium", color=navy)

for cell in ws[1]:
    cell.font = Font(name="Calibri", size=10, bold=True, color=white)
    cell.fill = PatternFill("solid", fgColor=navy)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(bottom=medium)

for col in range(1, 31):
    fill = (
        source_fill
        if 2 <= col <= 4
        else audit_fill
        if 5 <= col <= 13
        else work_fill
        if 14 <= col <= 25
        else classification_fill
    )
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row, col)
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.border = Border(bottom=thin)
        cell.alignment = Alignment(
            horizontal="center" if col in (1, 4, 6, 7, 8, 13, 24, 26, 28) else "left",
            vertical="top",
            wrap_text=True,
        )
        cell.font = Font(name="Calibri", size=10)

for row in range(2, ws.max_row + 1):
    ws.cell(row, 8).number_format = "0%"
    content_lengths = [
        len(str(ws.cell(row, 2).value or "")) / 34,
        len(str(ws.cell(row, 3).value or "")) / 92,
        len(str(ws.cell(row, 5).value or "")) / 52,
        len(str(ws.cell(row, 10).value or "")) / 52,
    ]
    estimated_lines = max(2, min(8, int(max(content_lengths)) + 1))
    ws.row_dimensions[row].height = estimated_lines * 15

widths = {
    1: 12,
    2: 34,
    3: 92,
    4: 15,
    5: 52,
    6: 12,
    7: 22,
    8: 17,
    9: 22,
    10: 52,
    11: 22,
    12: 22,
    13: 22,
    14: 24,
    15: 22,
    16: 28,
    17: 30,
    18: 26,
    19: 24,
    20: 18,
    21: 38,
    22: 34,
    23: 62,
    24: 20,
    25: 62,
    26: 26,
    27: 58,
    28: 28,
    29: 56,
    30: 70,
}
for col, width in widths.items():
    ws.column_dimensions[get_column_letter(col)].width = width
ws.row_dimensions[1].height = 42
ws.auto_filter.ref = f"A1:AD{ws.max_row}"

status_validation = DataValidation(
    type="list",
    formula1='"offen,teilstrukturiert,strukturiert,manuell modellieren,zurückgestellt"',
    allow_blank=False,
)
ws.add_data_validation(status_validation)
status_validation.add(f"X2:X{ws.max_row}")

red_fill = PatternFill("solid", fgColor="F4CCCC")
amber_fill = PatternFill("solid", fgColor="FCE5CD")
green_fill = PatternFill("solid", fgColor="D9EAD3")
ws.conditional_formatting.add(
    f"G2:G{ws.max_row}", FormulaRule(formula=['$G2="nicht gefunden"'], fill=red_fill)
)
ws.conditional_formatting.add(
    f"G2:G{ws.max_row}", FormulaRule(formula=['OR(LEFT($G2,4)="1:n ",LEFT($G2,9)="Vorschlag")'], fill=amber_fill)
)
ws.conditional_formatting.add(
    f"G2:G{ws.max_row}", FormulaRule(formula=['OR($G2="exakt",LEFT($G2,5)="Alias")'], fill=green_fill)
)
ws.conditional_formatting.add(
    f"M2:M{ws.max_row}", FormulaRule(formula=['$M2="Abweichung"'], fill=red_fill)
)
ws.conditional_formatting.add(
    f"M2:M{ws.max_row}", FormulaRule(formula=['$M2="OK"'], fill=green_fill)
)
ws.conditional_formatting.add(
    f"AB2:AB{ws.max_row}", FormulaRule(formula=['$AB2="Talent in alter V/N-Liste"'], fill=red_fill)
)
ws.conditional_formatting.add(
    f"AB2:AB{ws.max_row}", FormulaRule(formula=['$AB2="Aktueller Vor-/Nachteil"'], fill=green_fill)
)
ws.conditional_formatting.add(
    f"AB2:AB{ws.max_row}", FormulaRule(formula=['$AB2="Nicht in aktueller Liste"'], fill=amber_fill)
)

legend = output_wb.create_sheet("Legende")
legend.sheet_view.showGridLines = False
legend["A1"] = "Vor- und Nachteile v1.1 – Arbeitsübergabe"
legend["A1"].font = Font(name="Calibri", size=16, bold=True, color=white)
legend["A1"].fill = PatternFill("solid", fgColor=navy)
legend.merge_cells("A1:D1")
legend["A3"] = "Quelle"
legend["B3"] = str(SOURCE)
legend["A4"] = "Werte-Abgleich"
legend["B4"] = str(WERTE)
legend["A6"] = "Bereich"
legend["B6"] = "Bedeutung"
legend["A7"] = "B–D (grün)"
legend["B7"] = "Unverändert übernommene Quellspalten: Name, Wirkung, Wann wählbar."
legend["A8"] = "E–M (blau)"
legend["B8"] = "Automatischer, konservativer Abgleich gegen Werte; unsichere Fälle bleiben markiert."
legend["A9"] = "N–Y (gelb)"
legend["B9"] = "Arbeitsfelder für die schrittweise Wirkungsmodellierung und Claude-Übergabe."
legend["A10"] = "Werte-Zuordnungen"
legend["B10"] = "Vollständige Einzelauflösung aller 1:1- und 1:n-Treffer aus der Haupttabelle."
legend["A11"] = "Z–AD (violett)"
legend["B11"] = "Abgleich gegen die vom Nutzer gelieferte aktuelle Liste und Erkennung hineingerutschter Talente."
legend["A12"] = "Entfernte Zeilen"
legend["B12"] = "Auditprotokoll der entfernten Talente, Cyberimplantat-Zeilen und Nur-Noserun/Nur-NR-Einträge."
legend["A13"] = "Zuordnungsstatus"
legend["B13"] = "Bedeutung"
legend["A14"] = "exakt"
legend["B14"] = "Normalisierter Name stimmt direkt überein."
legend["A15"] = "Alias – Kategoriepräfix"
legend["B15"] = "Werte ergänzt nur ein Präfix wie Sicht:, Fähigkeit: oder Aussehen:."
legend["A16"] = "1:n – Werte aufgeteilt"
legend["B16"] = "Historischer Sammel-/Platzhaltereintrag wurde in Werte in mehrere Einzelwerte zerlegt."
legend["A17"] = "Vorschlag 1:1"
legend["B17"] = "Plausible historische Umbenennung; fachlich prüfen."
legend["A18"] = "nicht gefunden"
legend["B18"] = "Kein belastbarer Treffer; keine unsichere Zuordnung erzwungen."
for cell in legend[6]:
    if cell.column <= 2:
        cell.font = Font(bold=True, color=white)
        cell.fill = PatternFill("solid", fgColor=navy)
for row in range(3, 19):
    legend.cell(row, 1).alignment = Alignment(vertical="top", wrap_text=True)
    legend.cell(row, 2).alignment = Alignment(vertical="top", wrap_text=True)
legend.column_dimensions["A"].width = 28
legend.column_dimensions["B"].width = 110
legend.column_dimensions["C"].width = 3
legend.column_dimensions["D"].width = 3

mapping = output_wb.create_sheet("Werte-Zuordnungen")
mapping.sheet_view.showGridLines = False
mapping.freeze_panes = "A2"
mapping_headers = [
    "Quelle-Zeile",
    "Quellname",
    "Werte-Zeile",
    "Werte-Referenz",
    "Werte-Name",
    "Werte-Kosten",
    "Werte-Verfügbarkeit",
    "Zuordnungsstatus",
    "Zuordnungssicherheit",
]
mapping.append(mapping_headers)
for record in mapping_records:
    mapping.append(record)
for cell in mapping[1]:
    cell.font = Font(name="Calibri", size=10, bold=True, color=white)
    cell.fill = PatternFill("solid", fgColor=navy)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(bottom=medium)
for row in range(2, mapping.max_row + 1):
    for col in range(1, 10):
        cell = mapping.cell(row, col)
        cell.fill = PatternFill("solid", fgColor=audit_fill)
        cell.border = Border(bottom=thin)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    mapping.cell(row, 9).number_format = "0%"
mapping_widths = [12, 40, 12, 52, 48, 18, 22, 24, 20]
for col, width in enumerate(mapping_widths, start=1):
    mapping.column_dimensions[get_column_letter(col)].width = width
mapping.row_dimensions[1].height = 36
mapping.auto_filter.ref = f"A1:I{mapping.max_row}"

list_audit = output_wb.create_sheet("Listen-Abgleich")
list_audit.sheet_view.showGridLines = False
list_audit.freeze_panes = "A2"
list_headers = [
    "Quelle-Zeile",
    "Quellname",
    "Wirkung",
    "Abgleich aktuelle Liste",
    "Aktuelle Listen-Treffer",
    "Einordnung",
    "Talent-Referenz(en)",
    "Bereinigungsempfehlung",
]
list_audit.append(list_headers)
for source_row, name, effect, when in source_rows:
    list_matches, list_status, talent_matches, classification, recommendation = classify_source(name)
    list_audit.append(
        [
            source_row,
            name,
            effect,
            list_status,
            compact_join(list_matches),
            classification,
            compact_join(item["reference"] for item in talent_matches),
            recommendation,
        ]
    )
for cell in list_audit[1]:
    cell.font = Font(name="Calibri", size=10, bold=True, color=white)
    cell.fill = PatternFill("solid", fgColor=navy)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(bottom=medium)
for row in range(2, list_audit.max_row + 1):
    classification = list_audit.cell(row, 6).value
    row_fill = (
        green_fill
        if classification == "Aktueller Vor-/Nachteil"
        else red_fill
        if classification == "Talent in alter V/N-Liste"
        else amber_fill
    )
    for col in range(1, 9):
        cell = list_audit.cell(row, col)
        cell.fill = row_fill
        cell.border = Border(bottom=thin)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    effect_lines = min(7, max(2, int(len(str(list_audit.cell(row, 3).value or "")) / 95) + 1))
    list_audit.row_dimensions[row].height = effect_lines * 15
list_widths = [12, 38, 88, 26, 60, 28, 58, 68]
for col, width in enumerate(list_widths, start=1):
    list_audit.column_dimensions[get_column_letter(col)].width = width
list_audit.row_dimensions[1].height = 38
list_audit.auto_filter.ref = f"A1:H{list_audit.max_row}"

removed = output_wb.create_sheet("Entfernte Zeilen")
removed.sheet_view.showGridLines = False
removed.freeze_panes = "A2"
removed.append(["Quelle-Zeile", "Name", "Wirkung", "Wann wählbar", "Entfernungsgrund"])
for record in removed_rows:
    removed.append(list(record))
for cell in removed[1]:
    cell.font = Font(name="Calibri", size=10, bold=True, color=white)
    cell.fill = PatternFill("solid", fgColor=navy)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(bottom=medium)
for row in range(2, removed.max_row + 1):
    for col in range(1, 6):
        cell = removed.cell(row, col)
        cell.fill = red_fill
        cell.border = Border(bottom=thin)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    effect_lines = min(7, max(2, int(len(str(removed.cell(row, 3).value or "")) / 95) + 1))
    removed.row_dimensions[row].height = effect_lines * 15
removed_widths = [12, 40, 95, 18, 34]
for col, width in enumerate(removed_widths, start=1):
    removed.column_dimensions[get_column_letter(col)].width = width
removed.row_dimensions[1].height = 38
removed.auto_filter.ref = f"A1:E{removed.max_row}"

OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
output_wb.save(OUTPUT)
print(OUTPUT)
