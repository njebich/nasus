"""
Einmaliges Ergaenzungsskript: fuegt die Referenz "talentpunkte" (TaP) neu ins Werte-Sheet ein.

Hintergrund (mit Nutzer 2026-07-17 geklaert): TaP ist ein von Steigerungspunkten (SP)
komplett getrennter Punktepool, der ausschliesslich Talente (Kategorie "Talente") bezahlt.
Formel: TaP = 20 + Stufe*5. Existierte bisher gar nicht im Datensatz.

Aufruf:
    python scripts/add_talentpunkte_row.py "werte 0.7-claude.xlsx"

Legt vorher automatisch eine Sicherheitskopie an (siehe backup_werte.py).
"""
import sys
from pathlib import Path

import openpyxl

from backup_werte import backup


def main(path_str: str) -> None:
    path = Path(path_str)
    backup_path = backup(str(path))
    print(f"Sicherheitskopie angelegt: {backup_path}")

    wb = openpyxl.load_workbook(path, data_only=False)
    ws = wb["Werte"]
    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v:
            headers[v.strip()] = c

    for r in range(2, ws.max_row + 1):
        ref = ws.cell(row=r, column=headers["Referenz"]).value
        if ref == "talentpunkte":
            print(f"'talentpunkte' existiert bereits in Zeile {r} - nichts zu tun.")
            return

    new_row = ws.max_row + 1
    ws.cell(row=new_row, column=headers["Referenz"]).value = "talentpunkte"
    ws.cell(row=new_row, column=headers["Kategorie"]).value = "Charakterwerte"
    ws.cell(row=new_row, column=headers["Beschreibung"]).value = "Talentpunkte"
    ws.cell(row=new_row, column=headers["Art"]).value = "Formel"
    ws.cell(row=new_row, column=headers["Formel"]).value = "20+stufe*5"
    ws.cell(row=new_row, column=headers["Flag"]).value = (
        "Neu ergaenzt (2026-07-17, mit Nutzer geklaert): TaP ist ein von Steigerungspunkten "
        "getrennter Pool, der ausschliesslich die Kategorie 'Talente' bezahlt. "
        "Existierte vorher nicht im Datensatz."
    )

    wb.save(path)
    print(f"Zeile {new_row}: 'talentpunkte' (Formel: 20+stufe*5) ergaenzt.")
    print(f"Gespeichert: {path}")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print('Aufruf: python scripts/add_talentpunkte_row.py "werte 0.7-claude.xlsx"')
        sys.exit(1)
    main(sys.argv[1])
