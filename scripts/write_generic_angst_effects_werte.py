import re
import sys
from pathlib import Path


ROOT = Path(r"E:\Das Western Rollenspiel\LLM")
sys.path.insert(0, str(ROOT / ".python-deps"))

from openpyxl import load_workbook


WERTE = ROOT / "werte 0.8-claude.xlsx"
STAGE_DATA = {
    5: ("SITUATION", 0, "einmal pro Situation", "mit Unbehagen"),
    10: ("NEUE_KONFRONTATION", -1, "bei jeder erneuten direkten Konfrontation", "mit Nervosität"),
    15: ("10_KR", -2, "sofort, danach alle 10 KR", "mit Furcht"),
    20: ("5_KR", -3, "sofort, danach alle 5 KR", "mit Angst"),
    25: ("2_KR", -4, "sofort, danach alle 2 KR", "mit Panik"),
    30: ("1_KR", -5, "sofort, danach jede KR", "unter einer Phobie"),
}


def opening_text(source, stage, operator):
    if stage == 30:
        return f"Der Charakter leidet {operator} gegenüber {source}."
    return f"Der Charakter reagiert auf {source} {operator}."


def presence_text(source, be):
    return (
        f" Ist {source} zwar gegenwärtig, aber weder direkter Gegner noch Gegenstand, Wesen oder "
        f"sonstiges Ziel, mit dem sich der Charakter aktuell befasst, sind keine Angst-Proben "
        f"erforderlich; der Charakter erhält BE {be}."
    )


def generic_effect_text(source, stage, interval_text, be, operator):
    text = (
        f"{opening_text(source, stage, operator)} Bei direkter Konfrontation mit {source} muss er "
        f"{interval_text} eine um {stage} erschwerte Mut-Probe bestehen, um handlungsfähig zu "
        f"bleiben. Um sich {source} freiwillig zu nähern, die Angstquelle zu berühren oder bewusst "
        f"mit ihr zu interagieren, muss er einmalig eine um {stage} erschwerte "
        f"Willenskraft-Probe bestehen."
    )
    if stage > 5:
        text += presence_text(source, be)
    return text


def magic_effect_text(stage, interval_text, be, operator):
    if stage == 5:
        opening = "Der Charakter empfindet Unbehagen gegenüber aller Magie"
    elif stage == 30:
        opening = "Der Charakter leidet unter einer Phobie gegenüber aller Magie"
    else:
        operator_name = {
            10: "Nervosität",
            15: "Furcht",
            20: "Angst",
            25: "Panik",
        }[stage]
        opening = f"Der Charakter reagiert auf alle Magie mit {operator_name}"
    text = (
        f"{opening} und ist nicht in der Lage, irgendeine Form von Magie außer KI zu wirken. "
        f"Bei direkter Konfrontation mit Magie, insbesondere wenn er gegen einen Magiewirker "
        f"vorgeht, muss er {interval_text} eine um {stage} erschwerte Mut-Probe bestehen, um "
        f"handlungsfähig zu bleiben. Um Verzauberungen oder magische Heilung zuzulassen, ein "
        f"Artefakt zu benutzen oder bewusst mit Magie zu interagieren, muss er einmalig eine um "
        f"{stage} erschwerte Willenskraft-Probe bestehen."
    )
    if stage > 5:
        text += presence_text("Magie", be)
    return text


def main():
    workbook = load_workbook(WERTE, read_only=False, data_only=False)
    worksheet = workbook["Werte"]
    headers = {cell.value: cell.column for cell in worksheet[1] if cell.value}
    updated = []

    for row in range(2, worksheet.max_row + 1):
        reference = str(worksheet.cell(row, headers["Referenz"]).value or "")
        match = re.fullmatch(r"vn_angst_(.+)_(5|10|15|20|25|30)", reference)
        if not match:
            continue
        theme, stage_text = match.groups()
        stage = int(stage_text)
        interval_code, be, interval_text, operator = STAGE_DATA[stage]
        parent = str(worksheet.cell(row, headers["Parent"]).value or "")
        source = parent.removeprefix("Angst: ")
        old_flag = str(worksheet.cell(row, headers["Flag"]).value or "")
        base_flag = old_flag.split(" | ANGST_INTERVALL=", 1)[0]
        flag = (
            f"{base_flag} | ANGST_INTERVALL={interval_code} | ANGST_DIREKTE_KONFRONTATION=1 "
            f"| ANGST_BE={be}"
        )
        info = (
            f"Auswahlgruppe angst_{theme}; maximal 1 Stufe; Angstwert {stage}; direkte "
            f"Konfrontation: {interval_text}; bloße Gegenwart: BE {be}."
        )
        effect = (
            magic_effect_text(stage, interval_text, be, operator)
            if theme == "magie"
            else generic_effect_text(source, stage, interval_text, be, operator)
        )
        worksheet.cell(row, headers["Info"], info)
        worksheet.cell(row, headers["Flag"], flag)
        worksheet.cell(row, headers["Wirkung"], effect)
        updated.append((row, reference))

    if len(updated) != 108:
        raise RuntimeError(f"Erwartet: 108 Angstwerte; gefunden: {len(updated)}")
    workbook.save(WERTE)
    print(f"fear_effects_written={len(updated)}")


if __name__ == "__main__":
    main()
