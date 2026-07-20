"""
Importiert "ki-zaubersheet import.xlsx" (Sheet1, A1:F30 - Probe/Name/Eig.Bon/VD/WD/Wirkung)
fuer die 28 KI-Faehigkeiten in werte-0.8-claude.xlsx's "Werte"-Sheet (Zeilen 223-253):

- Spalte F (Wirkung) wird in die "Wirkung"-Spalte der passenden ki_*-Zeile geschrieben
  (Tooltip-Text, analog zu add_talente_wirkung_text.py fuer Talente).
- Spalten D/E (VD/WD) werden NICHT in die xlsx geschrieben (kein Bezug zum Werte-Sheet-
  Schema) - stattdessen als src/data/kiFaehigkeiten.ts (Referenz -> {vd, wd}, numerische
  Werte mit " KR" formatiert, Text-Werte wie "sofort"/"1h" unveraendert uebernommen).

Matching: ki-Sheet-Spalte B (Name) gegen Werte-Sheet-Spalte C (Beschreibung) der KI-
Kategorie, per Alias-Map fuer 3 Namensabweichungen (Konter/Konterattacke, Schlaf der
Heilung/Heilender Schlaf, Meister des..../Meister der Grundfertigkeiten).
"""
import sys
from pathlib import Path

ROOT = Path(r"E:\Das Western Rollenspiel\LLM")
sys.path.insert(0, str(ROOT / ".python-deps"))

from openpyxl import load_workbook

SOURCE = Path(r"E:\Das Western Rollenspiel\_Entwicklung\ki-zaubersheet import.xlsx")
WERTE = ROOT / "werte 0.8-claude.xlsx"
OUT_TS = ROOT / "src" / "data" / "kiFaehigkeiten.ts"

# Name (ki-Sheet, Spalte B) -> Beschreibung (Werte-Sheet, Spalte C), nur wo sie abweichen.
NAME_ALIASE = {
    "Konter": "Konterattacke",
    "Schlaf der Heilung": "Heilender Schlaf",
    "Meister des ....": "Meister der Grundfertigkeiten",
}


def read_source_rows():
    wb = load_workbook(SOURCE, read_only=True, data_only=True)
    ws = wb["Sheet1"]
    rows = []
    for r in range(2, 31):
        name = ws.cell(r, 2).value
        if not name:
            continue
        rows.append({
            "name": str(name).strip(),
            "eig_bon": ws.cell(r, 3).value,
            "vd": ws.cell(r, 4).value,
            "wd": ws.cell(r, 5).value,
            "wirkung": ws.cell(r, 6).value,
        })
    return rows


def format_dauer(value):
    if isinstance(value, (int, float)):
        n = int(value) if float(value).is_integer() else value
        return f"{n} KR"
    return str(value).strip()


def ts_escape(s):
    return s.replace("\\", "\\\\").replace("'", "\\'").replace("\n", "\\n")


def main():
    source_rows = read_source_rows()

    workbook = load_workbook(WERTE, read_only=False, data_only=False)
    worksheet = workbook["Werte"]
    headers = {cell.value: cell.column for cell in worksheet[1] if cell.value}
    ref_col = headers["Referenz"]
    kat_col = headers["Kategorie"]
    besch_col = headers["Beschreibung"]
    wirkung_col = headers["Wirkung"]

    beschreibung_to_row = {}
    for row in range(223, 254):
        if worksheet.cell(row, kat_col).value != "KI":
            continue
        besch = worksheet.cell(row, besch_col).value
        if besch:
            beschreibung_to_row[besch.strip()] = row

    written = 0
    already_filled = []
    unmatched = []
    vd_wd_entries = []
    for entry in source_rows:
        besch = NAME_ALIASE.get(entry["name"], entry["name"])
        row = beschreibung_to_row.get(besch)
        if row is None:
            unmatched.append(entry["name"])
            continue
        referenz = worksheet.cell(row, ref_col).value

        existing = worksheet.cell(row, wirkung_col).value
        if existing:
            already_filled.append(referenz)
        else:
            worksheet.cell(row, wirkung_col, entry["wirkung"])
            written += 1

        vd_wd_entries.append((referenz, format_dauer(entry["vd"]), format_dauer(entry["wd"])))

    if unmatched:
        raise SystemExit(f"Nicht gefundene Namen (kein Beschreibung-Match Zeilen 223-253): {unmatched}")
    if already_filled:
        raise SystemExit(f"Bereits befuellte Wirkung-Zellen (nicht ueberschrieben): {already_filled}")
    if len(vd_wd_entries) != 28:
        raise SystemExit(f"Erwartet 28 KI-Faehigkeiten, gefunden {len(vd_wd_entries)}")

    workbook.save(WERTE)
    print(f"{written} Wirkung-Zellen in '{WERTE.name}' geschrieben.")

    lines = [
        "// Auto-generiert von scripts/import_ki_zaubersheet.py aus "
        "\"ki-zaubersheet import.xlsx\" (Vorbereitungsdauer/Wirkungsdauer je KI-Faehigkeit,",
        "// numerische Werte in Kampfrunden formatiert als \"<n> KR\", Text-Werte (z.B. \"sofort\", \"1h\") unveraendert).",
        "export interface KiDauer { vd: string; wd: string; }",
        "",
        "export const KI_DAUER: Record<string, KiDauer> = {",
    ]
    for referenz, vd, wd in sorted(vd_wd_entries):
        # Referenz als String-Key quoten (nicht als bare Identifier) - z.B. "ki_gleichgewicht_&_
        # feingefuehl" enthaelt ein "&" und waere sonst ungueltige TS-Syntax.
        lines.append(f"  '{ts_escape(referenz)}': {{ vd: '{ts_escape(vd)}', wd: '{ts_escape(wd)}' }},")
    lines.append("};")
    OUT_TS.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(f"{OUT_TS}: {len(vd_wd_entries)} VD/WD-Eintraege geschrieben.")


if __name__ == "__main__":
    main()
