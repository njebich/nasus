import sys
from pathlib import Path


ROOT = Path(r"E:\Das Western Rollenspiel\LLM")
sys.path.insert(0, str(ROOT / ".python-deps"))

from openpyxl import load_workbook


WERTE = ROOT / "werte 0.8-claude.xlsx"

SCHOOLS = {
    "erd": "Erdbeschwörung",
    "feuer": "Feuerbeschwörung",
    "luft": "Luftbeschwörung",
    "magie": "Magiebeschwörung",
    "wasser": "Wasserbeschwörung",
}

EFFECT = (
    "Alle Zauber mit Schadenswirkung dieses Elements (auch Verzauberungs-, Antimagie- und "
    "Beschwörungs-Zauber) kosten 9 EP und haben die 1,5-fache Erschwerung. Kontakt mit dem "
    "Element oder einem Wirker erfordert jede KR eine Mut-Probe. PSI-Zauber, die diesem Element "
    "zugeordnet werden können, kosten 18 EP pro Punkt. Magier, die diesen Nachteil und die "
    "[Schule] bei Erschaffung wählen, erhalten 650 EP zusätzlich, wovon mindestens 180 EP auf "
    "die [Schule] verteilt werden müssen."
)

OTHER_SCHOOLS = {
    "heilung": "Heilung",
    "verzauberung": "Verzauberung",
    "beherrschung": "Beherrschung",
}

OTHER_EFFECT = (
    "Alle Zauber dieser Schule (sowie Antimagie-Zauber, die diese Schule betreffen) kosten 9 EP "
    "und haben die 1,5-fache Erschwerung. Kontakt einem Wirker erfordert jede KR eine Mut-Probe. "
    "PSI-Zauber, die dieser Schule zugeordnet werden können, kosten 18 EP pro Punkt. Magier, die "
    "diesen Nachteil und die [Schule] bei Erschaffung wählen, erhalten 650 EP zusätzlich, wovon "
    "mindestens 180 EP auf die [Schule] verteilt werden müssen."
)


def main():
    workbook = load_workbook(WERTE, read_only=False, data_only=False)
    worksheet = workbook["Werte"]
    headers = {cell.value: cell.column for cell in worksheet[1] if cell.value}
    existing = {
        worksheet.cell(row, headers["Referenz"]).value
        for row in range(2, worksheet.max_row + 1)
    }
    changed = False

    old_earth_reference = "vn_abneigung_gegen_erdebeschwoerung"
    new_earth_reference = "vn_abneigung_gegen_erdbeschwoerung"
    if old_earth_reference in existing and new_earth_reference not in existing:
        for row in range(2, worksheet.max_row + 1):
            if worksheet.cell(row, headers["Referenz"]).value == old_earth_reference:
                worksheet.cell(row, headers["Referenz"], new_earth_reference)
                existing.remove(old_earth_reference)
                existing.add(new_earth_reference)
                changed = True
                break

    added = []
    for slug, school in SCHOOLS.items():
        reference = f"vn_abneigung_gegen_{slug}beschwoerung"
        if reference in existing:
            continue
        row = worksheet.max_row + 1
        values = {
            "Referenz": reference,
            "Kategorie": "Vor- und Nachteile",
            "Beschreibung": f"Abneigung gegen {school}",
            "Art": "Auswahl",
            "Kosten": -650,
            "Verfuegbarkeit": "E&1",
            "Wirkung": EFFECT,
        }
        for header, value in values.items():
            worksheet.cell(row, headers[header], value)
        added.append((row, reference, school))
        changed = True

    for slug, school in OTHER_SCHOOLS.items():
        reference = f"vn_abneigung_gegen_{slug}"
        if reference in existing:
            continue
        row = worksheet.max_row + 1
        values = {
            "Referenz": reference,
            "Kategorie": "Vor- und Nachteile",
            "Beschreibung": f"Abneigung gegen {school}",
            "Art": "Auswahl",
            "Kosten": -650,
            "Verfuegbarkeit": "E&1",
            "Wirkung": OTHER_EFFECT,
        }
        for header, value in values.items():
            worksheet.cell(row, headers[header], value)
        added.append((row, reference, school))
        changed = True

    if changed:
        workbook.save(WERTE)
    print(f"added={len(added)}")
    for row, reference, school in added:
        print(row, reference, school)


if __name__ == "__main__":
    main()
