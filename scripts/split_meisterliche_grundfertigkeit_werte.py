import sys
from pathlib import Path


ROOT = Path(r"E:\Das Western Rollenspiel\LLM")
sys.path.insert(0, str(ROOT / ".python-deps"))

from openpyxl import load_workbook


WERTE = ROOT / "werte 0.8-claude.xlsx"
GENERIC_REFERENCE = "vn_faehigkeit_meisterliche_grundfertigkeit"
WHK_REFERENCE = "vn_faehigkeit_meisterliche_WHK"


def main():
    workbook = load_workbook(WERTE, read_only=False, data_only=False)
    worksheet = workbook["Werte"]
    headers = {cell.value: cell.column for cell in worksheet[1] if cell.value}

    basic_skills = []
    for row in range(2, worksheet.max_row + 1):
        if worksheet.cell(row, headers["Kategorie"]).value != "Grundfertigkeit":
            continue
        target_reference = worksheet.cell(row, headers["Referenz"]).value
        target_name = worksheet.cell(row, headers["Beschreibung"]).value
        if target_reference and target_name:
            basic_skills.append((target_reference, target_name))
    if len(basic_skills) != 22:
        raise RuntimeError(f"Erwartet: 22 Grundfertigkeiten; gefunden: {len(basic_skills)}")

    generic_row = next(
        (
            row
            for row in range(2, worksheet.max_row + 1)
            if worksheet.cell(row, headers["Referenz"]).value == GENERIC_REFERENCE
        ),
        None,
    )
    if generic_row is not None:
        worksheet.delete_rows(generic_row, 1)

    existing_references = {
        worksheet.cell(row, headers["Referenz"]).value
        for row in range(2, worksheet.max_row + 1)
    }
    created = []
    for target_reference, target_name in basic_skills:
        suffix = target_reference.removeprefix("gr_")
        reference = f"vn_faehigkeit_meisterliche_grundfertigkeit_{suffix}"
        if reference in existing_references:
            continue
        row = worksheet.max_row + 1
        effect = (
            f"Nicht Magier: Der Charakter kann max. (M) Mana für (M)×2 min in 1 Talentpunkt pro "
            f"4 Mana der Grundfertigkeit {target_name} umwandeln. Dieser Vorteil gilt ausschließlich "
            f"für {target_name} und muss für jede Grundfertigkeit einzeln gekauft werden."
        )
        values = {
            "Referenz": reference,
            "Kategorie": "Vor- und Nachteile",
            "Beschreibung": f"Fähigkeit: Meisterliche Grundfertigkeit: {target_name}",
            "Info": f"Festes Ziel: {target_name} ({target_reference}).",
            "Parent": "Meisterliche Grundfertigkeit",
            "Art": "Auswahl",
            "Flag": f"ZIEL_GRUNDFERTIGKEIT={target_reference}",
            "Kosten": 100,
            "Verfuegbarkeit": "E&3",
            "Wirkung": effect,
        }
        for header, value in values.items():
            worksheet.cell(row, headers[header], value)
        existing_references.add(reference)
        created.append((row, reference, target_reference))

    whk_row = next(
        row
        for row in range(2, worksheet.max_row + 1)
        if worksheet.cell(row, headers["Referenz"]).value == WHK_REFERENCE
    )
    worksheet.cell(
        whk_row,
        headers["Info"],
        "Offene WHK-Liste: In der App ist ein dynamisches Dropdown über alle vorhandenen WHK "
        "erforderlich. Jede gewählte WHK ist eine eigene, separat kaufbare Instanz; dieselbe WHK "
        "darf nicht doppelt gewählt werden. Umsetzung durch Claude.",
    )
    worksheet.cell(whk_row, headers["Parent"], "Meisterliche WHK")
    worksheet.cell(whk_row, headers["Flag"], "APP_DROPDOWN_WHK_OFFENE_LISTE")
    worksheet.cell(
        whk_row,
        headers["Wirkung"],
        "Nicht Magier: Der Charakter kann max. (M) Mana für (M) h in 1 Talentpunkt pro 2 Mana "
        "einer ausgewählten spezialisierbaren WHK umwandeln. Dieser Vorteil muss für jede WHK "
        "einzeln gekauft werden.",
    )

    workbook.save(WERTE)
    print(f"created_basic_skill_entries={len(created)}")
    print(f"generic_removed={generic_row is not None}")
    print(f"whk_dropdown_marked_row={whk_row}")


if __name__ == "__main__":
    main()
