import sys
from pathlib import Path


ROOT = Path(r"E:\Das Western Rollenspiel\LLM")
sys.path.insert(0, str(ROOT / ".python-deps"))

from openpyxl import load_workbook


WERTE = ROOT / "werte 0.8-claude.xlsx"


def kb_core(erschwerung: str, wuerfel: str) -> str:
    return (
        f"Das Opfer muss eine KB-Probe (GF Körperbeherrschung + Eigenschaft Athletik) ablegen, "
        f"{erschwerung}. Bei Misserfolg wird es um so viele Meter zurückgeschleudert, wie {wuerfel} "
        f"anzeigt, und das Probenergebnis wird wie bei einer SB-Probe in der Selbstbeherrschungstabelle "
        f"nachgeschlagen, gedeckelt auf Status: Sturz (KK -2)."
    )


# referenz -> (intro, erschwerung-clause, wuerfel-clause, schaden-tail)
LUFTBESCHWOERUNG_SPELLS = {
    "spruchmagie_luftbeschwoerung_1_sturmfaust": (
        "Die Fäuste des Magiers verursachen zusätzlichen Elementarschaden. ",
        "erschwert um TP + 3",
        "der Schadenswürfel",
        "Schaden: W8 + {Magie}.",
    ),
    "spruchmagie_luftbeschwoerung_1_sturmball": (
        "Der Magus erschafft einen Ball aus Luft. ",
        "erschwert um TP + 6",
        "der höchste Schadenswürfel",
        "Schaden: 2W4 + {Magie}.",
    ),
    "spruchmagie_luftbeschwoerung_1_sturmwand": (
        "Der Magier erschafft eine Wand aus Luft. Die Wahrnehmung durch diese Wand ist stark "
        "erschwert bis unmöglich: Alle FK-Attacken sind um {Aura}*2 erschwert. Wer versucht, die "
        "Wand zu durchdringen, erleidet Elementarschaden auf jeder TZ. Während der Effektdauer "
        "entfaltet die Wand nur die halbe Wirkung. ",
        "erschwert um TP + 42",
        "der Schadenswürfel",
        "Schaden: W4+{Magie} auf jede TZ.",
    ),
    "spruchmagie_luftbeschwoerung_2_sturmpfeil": (
        "Der Magus erschafft einen Bolzen aus Luft. Treffer auf FK-TZ. ",
        "erschwert um TP + 6",
        "der höchste Schadenswürfel",
        "Schaden: 2W6 + {Magie}, RB {Magie}*2.",
    ),
    "spruchmagie_luftbeschwoerung_2_sturmaura": (
        "Der Magus hüllt sich in eine kugelförmige  Aura aus Luft. Wer sich in die Aura begibt, "
        "erleidet auf jeder TZ innerhalb dieses Gebietes Elementarschaden. Der Magier kann aus der "
        "Aura heraus normal zielen, FK-Angriffe sind um {Aura}*2 erschwert. Feindliche Magie ist mit "
        "Ziel Aura hiervon nicht betroffen. ",
        "erschwert um TP + 21",
        "der Schadenswürfel",
        "Schaden: W4+{Magie} auf jede TZ.",
    ),
    "spruchmagie_luftbeschwoerung_2_sturmstrahl": (
        "Aus der Handfläche des Magiers schießt ein {M} cm dicker Strahl aus Luft. Treffer auf eine "
        "FK-TZ mit 1 zusätzlichen TZ-Reroll. ",
        "erschwert um TP + 3",
        "der Schadenswürfel",
        "Schaden: W10+{Magie} pro sec.",
    ),
    "spruchmagie_luftbeschwoerung_3_grosse_sturmfaust": (
        "Die Fäuste des Magiers verursachen zusätzlichen Elementarschaden. ",
        "erschwert um TP + 9",
        "der höchste Schadenswürfel",
        "Schaden: 2W12+{Magie}.",
    ),
    "spruchmagie_luftbeschwoerung_3_sturmball_salve": (
        "Der Magus erschafft {Magie} : 2 Bälle aus Luft. ",
        "erschwert um TP + 6",
        "der jeweils höchste Schadenswürfel",
        "Schaden: 2W4 + {Magie}.",
    ),
    "spruchmagie_luftbeschwoerung_3_luft_klinge": (
        "Eine Waffe erhält zusätzliche Schadenswirkung entsprechend der Magie des Magus. ",
        "erschwert um TP",
        "der Zauber-Schadenswürfel",
        "Schaden: 1=W3, SB -2; 2=W4, SB -3; 3=W6, SB -3; 4=W8, SB -3; 5=W10, SB -3; 6=W12, SB -5; "
        "7=W15, SB -8; 8=W20, SB -15; 9=W30, SB -15; 10=2W20, SB -15; 11=W50, SB -15. Bei guter oder "
        "besserer Probe können die KK-Punkte zum Erhöhen der Wirkung nur dafür eingesetzt werden, die "
        "W-Klasse zu erhöhen.",
    ),
    "spruchmagie_luftbeschwoerung_3_sturm_pfeil": (
        "Ein Pfeil erhält eine zusätzliche Schadenswirkung entsprechend der Magie des Magus. ",
        "erschwert um TP",
        "der Zauber-Schadenswürfel",
        "Schaden: 1=W3, SB -2; 2=W4, SB -3; 3=W6, SB -3; 4=W8, SB -3; 5=W10, SB -3; 6=W12, SB -5; "
        "7=W15, SB -8; 8=W20, SB -15; 9=W30, SB -15; 10=2W20, SB -15; 11=W50, SB -15. Bei guter oder "
        "besserer Probe können die KK-Punkte zum Erhöhen der Wirkung nur dafür eingesetzt werden, die "
        "W-Klasse zu erhöhen.",
    ),
    "spruchmagie_luftbeschwoerung_3_sturmblitz": (
        "Der Magus erschafft um das Opfer eine Kugel aus Luft. Treffer auf FK-TZ mit 2 TZ-Rerolls "
        "auf die gewünschte Trefferzone. ",
        "erschwert um TP + 6",
        "der höchste Schadenswürfel",
        "Schaden: 2W6 + {Magie}.",
    ),
    "spruchmagie_luftbeschwoerung_4_grosser_sturmball": (
        "Der Magus erschafft einen Ball aus Luft. ",
        "erschwert um TP + 35",
        "der Schadenswürfel",
        "Schaden: W12 + {Magie} auf jede TZ.",
    ),
    "spruchmagie_luftbeschwoerung_4_sturmpfeil_salve": (
        "Der Magus erschafft {Magie} : 2 Bolzen aus Luft. Treffer auf FK-TZ. ",
        "erschwert um TP + 6",
        "der jeweils höchste Schadenswürfel",
        "Schaden: 2W6 + {Magie}; RB {Magie}*2.",
    ),
    "spruchmagie_luftbeschwoerung_4_grosse_sturmwand": (
        "Der Magier erschafft eine Wand aus Luft. Die Wahrnehmung durch diese Wand ist stark "
        "erschwert bis unmöglich: Alle FK-Attacken sind um {Aura}*2 erschwert. Wer versucht, die Wand "
        "zu durchdringen, erleidet Elementarschaden auf jeder TZ. Länge max. {M} m, Breite {M}*5cm. "
        "Während der Effektdauer entfaltet die Wand nur die halbe Wirkung. ",
        "erschwert um TP + 42",
        "der höchste Schadenswürfel",
        "Schaden: 2W6 + {Magie}.",
    ),
    "spruchmagie_luftbeschwoerung_5_grosse_sturmaura": (
        "Der Magus hüllt sich in eine kugelförmige  Aura aus Luft. Wer sich in die Aura begibt, "
        "erleidet auf jeder TZ innerhalb dieses Gebietes Elementarschaden. Der Magier kann aus der "
        "Aura heraus normal zielen, FK-Angriffe sind um {Aura}*2 erschwert. Feindliche Magie ist mit "
        "Ziel Aura hiervon nicht betroffen. ",
        "erschwert um TP + 42",
        "der höchste Schadenswürfel",
        "Schaden: 2W6+{Magie} auf jede TZ.",
    ),
    "spruchmagie_luftbeschwoerung_5_tornado": (
        "Eine vom Magus ausgehende kugelförmige Explosion verursacht Elementarschaden. Es ist "
        "unmöglich, dem Schaden zu entgehen. Jede TZ wird getroffen. ",
        "erschwert um TP + 21",
        "der Schadenswürfel",
        "Schaden: W8 + {Magie}.",
    ),
    "spruchmagie_luftbeschwoerung_5_grosser_sturmpfeil": (
        "Der Magus erschafft einen Bolzen aus Luft. Treffer auf FK-TZ. ",
        "erschwert um TP + 6",
        "der höchste Schadenswürfel",
        "Schaden: 2W10 + {Magie}; RB {M}:2.",
    ),
    "spruchmagie_luftbeschwoerung_6_grosse_sturmball_salve": (
        "Der Magus erschafft {Magie} : 2 Bälle aus Luft. ",
        "erschwert um TP + 35",
        "der jeweilige Schadenswürfel",
        "Schaden: W12 + {Magie} auf jede Trefferzone.",
    ),
    "spruchmagie_luftbeschwoerung_6_grosser_sturmstrahl": (
        "Aus der Handfläche des Magiers schießt ein {M} cm dicker Strahl aus Luft. Treffer auf eine "
        "FK-TZ mit 1 TZ-Reroll. ",
        "erschwert um TP + 6",
        "der höchste Schadenswürfel",
        "Schaden: 2W10 + {Magie} pro sec.",
    ),
    "spruchmagie_luftbeschwoerung_6_grosser_sturmblitz": (
        "Der Magus erschafft um das Opfer eine Kugel aus Luft. Treffer auf FK-TZ mit 2 TZ-Rerolls "
        "auf die gewünschte Trefferzone. ",
        "erschwert um TP + 9",
        "der höchste Schadenswürfel",
        "Schaden: 2W12 + {Magie}.",
    ),
    "spruchmagie_luftbeschwoerung_7_grosser_tornado": (
        "Eine vom Magus ausgehende kugelförmige Explosion verursacht Elementarschaden. Es ist "
        "unmöglich, dem Schaden zu entgehen. ",
        "erschwert um TP + 105",
        "der höchste Schadenswürfel",
        "Schaden: 2W15 + {Magie} auf jede TZ.",
    ),
    "spruchmagie_luftbeschwoerung_7_grosse_sturmpfeil_salve": (
        "Der Magus erschafft {Magie} : 2 Bolzen aus Luft. Jeweils Treffer auf FK-TZ. ",
        "erschwert um TP + 6",
        "der jeweils höchste Schadenswürfel",
        "Schaden: 2W10 + {Magie}, RB {M}:2.",
    ),
}

TALENT_REFERENZ = "talente_mit_schild_umwerfen"
TALENT_WIRKUNG = (
    "Vorraussetzung Schild: Wird im Nahkampf erst angesagt, nach dem der Verteidiger seine Aktion "
    "gewählt hat. Ermöglicht es, den Gegner mit einer um X erschwerten AT umzuwerfen. Die Erschwerung "
    "muss mindestens 6 betragen. Wenn der Gegner mit gleicher oder schlechterer PA-Kl. pariert als die "
    "AT-Kl. des Angreifers (was keinen Schaden verursacht) oder getroffen wird (was dann halben "
    "Schaden des Schilds verursacht), muss er eine KB-Probe (GF Körperbeherrschung + Eigenschaft "
    "Athletik) ablegen, erschwert um X × 3, um nicht zu fallen. Gute oder meisterliche AT haben keine "
    "weitere Auswirkung. Bei Misserfolg geht der Verteidiger direkt in Status: Sturz (KK -2) über. "
    "Wenn der Verteider erfolgreich ausweicht oder mit besserer PA-Kl. pariert, so wird die "
    "ngmATPA-Tabelle konsultiert und das Umwerfen schlägt fehl."
)

ZWEI_WAFFEN_WIRKUNG = {
    "talente_kampf_mit_zwei_waffen_stufe_1": (
        "Erlaubt dem Charakter, zwei Waffen mit max. WK 3,5 (unmodifizierter Listenwert) im Kampf zu "
        "führen und unbewaffnete Extra-Aktionen mit der Nebenhand auszuführen. Waffen der "
        "Spezialisierung Schild und Stangenwaffen sind ausgeschlossen.\n"
        "Die n-Mods und Mindeststärken beider Waffen werden addiert.\n"
        "Bei Attacken gilt die 1,5-fache WK der Waffe mit der höheren WK, bei Paraden werden beide WK "
        "addiert. Die gemeinsame AT- und PA-WK kann niemals unter die höhere aktuelle Einzel-WK "
        "sinken.\n"
        "Der Charakter attackiert mit den Werten der Waffe, die jeweils zum Angriff genutzt wird, und "
        "pariert immer mit dem besseren Paradewert beider Waffen."
    ),
    "talente_kampf_mit_zwei_waffen_stufe_2": (
        "Erlaubt dem Charakter, zwei Waffen mit max. WK 4,5 (unmodifizierter Listenwert) im Kampf zu "
        "führen. Waffen der Spezialisierung Schild und Stangenwaffen sind ausgeschlossen. Die "
        "gemeinsame AT- und PA-WK kann niemals unter die höhere aktuelle Einzel-WK sinken."
    ),
    "talente_kampf_mit_zwei_waffen_stufe_3": (
        "Erlaubt dem Charakter, zwei Waffen mit max. WK 5,5 (unmodifizierter Listenwert) im Kampf zu "
        "führen. Waffen der Spezialisierung Schild und Stangenwaffen sind ausgeschlossen. Die "
        "gemeinsame AT- und PA-WK kann niemals unter die höhere aktuelle Einzel-WK sinken."
    ),
    "talente_kampf_mit_zwei_waffen_stufe_4": (
        "Erlaubt dem Charakter, zwei Waffen mit max. WK 6,5 (unmodifizierter Listenwert) im Kampf zu "
        "führen. Waffen der Spezialisierung Schild und Stangenwaffen sind ausgeschlossen. Die "
        "gemeinsame AT- und PA-WK kann niemals unter die höhere aktuelle Einzel-WK sinken."
    ),
}


def main():
    workbook = load_workbook(WERTE, read_only=False, data_only=False)
    worksheet = workbook["Werte"]
    headers = {cell.value: cell.column for cell in worksheet[1] if cell.value}
    ref_col = headers["Referenz"]
    wirkung_col = headers["Wirkung"]

    by_referenz = {}
    for row in range(2, worksheet.max_row + 1):
        ref = worksheet.cell(row, ref_col).value
        if ref:
            by_referenz[ref] = row

    updated = []
    for referenz, (intro, erschwerung, wuerfel, tail) in LUFTBESCHWOERUNG_SPELLS.items():
        row = by_referenz[referenz]
        text = intro + kb_core(erschwerung, wuerfel) + " " + tail
        worksheet.cell(row, wirkung_col, text)
        updated.append((row, referenz))

    talent_row = by_referenz[TALENT_REFERENZ]
    worksheet.cell(talent_row, wirkung_col, TALENT_WIRKUNG)
    updated.append((talent_row, TALENT_REFERENZ))

    for referenz, text in ZWEI_WAFFEN_WIRKUNG.items():
        row = by_referenz[referenz]
        worksheet.cell(row, wirkung_col, text)
        updated.append((row, referenz))

    assert len(LUFTBESCHWOERUNG_SPELLS) == 22, len(LUFTBESCHWOERUNG_SPELLS)

    workbook.save(WERTE)
    for row, referenz in updated:
        print(row, referenz)


if __name__ == "__main__":
    main()
