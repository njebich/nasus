import sys
from pathlib import Path


ROOT = Path(r"E:\Das Western Rollenspiel\LLM")
sys.path.insert(0, str(ROOT / ".python-deps"))

from openpyxl import load_workbook


WERTE = ROOT / "werte 0.8-claude.xlsx"
REFERENCE = "vn_kind_der_froehlichkeit"
EFFECT = (
    "Wirft der Charakter eine fröhliche Probe, hat er den Anspruch auf eine weitere Probe auf das "
    "selbe Talent. Bei normalem/gutem/meisterlichem/fröhlichem Erfolg in dieser Probe erhöht sich "
    "das Maximum bei Eigenschaften um 0/1/2/3, bei Attributen um 0/0/0/1, bei Grundfertigkeiten "
    "um 1/2/3/4, bei Kampftalenten um 0/1/2/3, bei Optios und Sprachen um 2/4/6/8, bei PSI- und "
    "Spruchzaubern (jeweils für den einzelnen Zauber) um 1/2/3/4; bei KI-Fähigkeiten wird das "
    "Talent selbst um 1/2/3/4 erhöht (die Erhöhung wird beim Steigern nicht einbezogen, sondern "
    "immer nach dem Steigern addiert). Jeder Spielercharakter besitzt diesen Vorteil automatisch."
)


def main():
    workbook = load_workbook(WERTE, read_only=False, data_only=False)
    worksheet = workbook["Werte"]
    headers = {cell.value: cell.column for cell in worksheet[1] if cell.value}
    row = next(
        (
            row
            for row in range(2, worksheet.max_row + 1)
            if worksheet.cell(row, headers["Referenz"]).value == REFERENCE
        ),
        worksheet.max_row + 1,
    )
    values = {
        "Referenz": REFERENCE,
        "Kategorie": "Vor- und Nachteile",
        "Beschreibung": "Kind der Fröhlichkeit",
        "Info": (
            "Jeder Spielercharakter besitzt diesen Vorteil automatisch; nicht manuell wählbar; "
            "bei der Spielercharakter-Erstellung verpflichtend zu vergeben."
        ),
        "Parent": "Spielercharakter",
        "Art": "Automatisch",
        "Flag": "SPIELERCHARAKTER_PFLICHT",
        "Kosten": 0,
        "Verfuegbarkeit": "SC",
        "Wirkung": EFFECT,
    }
    for header, value in values.items():
        worksheet.cell(row, headers[header], value)
    workbook.save(WERTE)
    print(row, REFERENCE)


if __name__ == "__main__":
    main()
