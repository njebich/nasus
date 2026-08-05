"""One-off: turn the 5 flat Geweihter gate talents into a 7-step progression (Stufenkette
naming convention talente_geweihter_<religion>_stufe_<N>_orthodox, see
engine/talenteStufenKette.ts) + add the new "Gute Wunder" Talent. See project_geweihte_tab
memory for the full feature and Nutzer-Ask 2026-08-06.

Run once, then recalc via Excel COM (openpyxl clears cached formula values on save) before
re-running generate_data_ts.py.

Usage: python scripts/add_geweihte_stufen.py "werte 0.8-claude.xlsx"
"""
import sys
import openpyxl

WERTE_HEADERS = [
    'Referenz', 'Kategorie', 'Beschreibung', 'Abkürzung', 'Info', 'Parent', 'Art',
    'Formel', 'Pool', 'Flag', 'Grad', 'Kosten', 'Verfuegbarkeit', None,
    'Mindest-TaW', 'Eig-Bonus', 'Wirkung',
]

RELIGIONEN = ['Lloth', 'Khartazh', 'Nomna', 'Tepod', 'Isch']

# Grad/Titel-Tabelle wie engine/geweihte.ts GEWEIHTEN_GRADE (grad 0 = kein Titel, nicht Teil
# der Stufenkette).
TITEL_BY_GRAD = {
    1: 'Niederer', 2: 'Minderer', 3: 'Konfirmierter', 4: 'Etablierter',
    5: 'Angesehener', 6: 'Gesalbter', 7: 'Heiliger',
}

TAP_KOSTEN = 5


def col_index(header_name):
    return WERTE_HEADERS.index(header_name) + 1


def main():
    path = sys.argv[1] if len(sys.argv) > 1 else 'werte 0.8-claude.xlsx'
    wb = openpyxl.load_workbook(path, read_only=False, data_only=False)
    ws = wb['Werte']

    # 1) Die 5 bestehenden Zeilen (Grad 1) auf die Stufenketten-Namenskonvention umbenennen +
    #    Kosten auf 5 anheben (Nutzer-Ask: "Raise TaP price of Geweihter to 5").
    renamed = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        referenz_cell = row[col_index('Referenz') - 1]
        referenz = referenz_cell.value
        if not referenz or not str(referenz).startswith('talente_geweihter_') or '_stufe_' in str(referenz):
            continue
        for name in RELIGIONEN:
            if referenz == f'talente_geweihter_{name.lower()}_orthodox':
                new_referenz = f'talente_geweihter_{name.lower()}_stufe_1_orthodox'
                referenz_cell.value = new_referenz
                row[col_index('Beschreibung') - 1].value = f'Geweihter von {name}, Orthodox – Niederer'
                row[col_index('Kosten') - 1].value = TAP_KOSTEN
                row[col_index('Wirkung') - 1].value = (
                    f'Dieser Charakter ist Geweihter von {name}, Orthodox (Geweihtengrad 1: Niederer). '
                    'Der Charakter muss Karma auf mindestens 1 steigern.'
                )
                renamed.append((referenz_cell.row, referenz, new_referenz))
                break

    if len(renamed) != len(RELIGIONEN):
        raise SystemExit(f'Erwartete {len(RELIGIONEN)} bestehende Stufe-1-Zeilen, gefunden: {len(renamed)}')

    # 2) Stufe 2-7 je Religion neu anlegen (30 Zeilen) + "Gute Wunder"-Talent (1 Zeile).
    new_rows = []
    for name in RELIGIONEN:
        for grad in range(2, 8):
            titel = TITEL_BY_GRAD[grad]
            new_rows.append({
                'Referenz': f'talente_geweihter_{name.lower()}_stufe_{grad}_orthodox',
                'Kategorie': 'Talente',
                'Beschreibung': f'Geweihter von {name}, Orthodox – {titel}',
                'Parent': 'Geweihte',
                'Art': 'Auswahl',
                'Kosten': TAP_KOSTEN,
                'Wirkung': f'Dieser Charakter ist Geweihter von {name}, Orthodox (Geweihtengrad {grad}: {titel}).',
            })
    new_rows.append({
        'Referenz': 'talente_geweihte_gute_wunder',
        'Kategorie': 'Talente',
        'Beschreibung': 'Gute Wunder',
        'Parent': 'Geweihte',
        'Art': 'Auswahl',
        'Kosten': 15,
        'Wirkung': 'Ersetzt die gute Wunder-Probe durch Karma. Die höchstmögliche Gute ist Normale : 2.',
    })

    last_referenz_row = 1
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        if row[0].value is not None:
            last_referenz_row = row[0].row
    next_row = last_referenz_row + 1

    added = []
    for data in new_rows:
        for header, value in data.items():
            ws.cell(row=next_row, column=col_index(header), value=value)
        added.append((next_row, data['Referenz']))
        next_row += 1

    # 3) Entwickeln-Backlog-Zeile "Geweihtengrad-Steigerung (Grad 2-7)" auf erledigt setzen.
    ws_ent = wb['Entwickeln']
    updated_entwickeln = None
    for row in ws_ent.iter_rows(min_row=2, max_row=ws_ent.max_row):
        if row[1].value == 'Geweihtengrad-Steigerung (Grad 2-7)':
            row[3].value = 'Erledigt'
            row[2].value = (
                row[2].value
                + ' UPDATE 2026-08-06: doch spielerseitig steigerbar gemacht (Nutzer-Ask) - '
                'Grad 2-7 sind jetzt talente_geweihter_<religion>_stufe_<2-7>_orthodox, '
                'Stufenketten-Talente wie Magus (siehe engine/talenteStufenKette.ts), je 5 TaP.'
            )
            updated_entwickeln = row[0].row
            break

    wb.save(path)
    print(f'Renamed {len(renamed)} Werte rows (Stufe 1):')
    for row_num, old, new in renamed:
        print(f'  row {row_num}: {old} -> {new}')
    print(f'Added {len(added)} Werte rows:')
    for row_num, referenz in added:
        print(f'  row {row_num}: {referenz}')
    print(f'Updated Entwickeln row: {updated_entwickeln}')


if __name__ == '__main__':
    main()
