import re
import sys
from pathlib import Path


ROOT = Path(r"E:\Das Western Rollenspiel\LLM")
sys.path.insert(0, str(ROOT / ".python-deps"))

from openpyxl import load_workbook


WERTE = ROOT / "werte 0.8-claude.xlsx"
SCHOOL_LABELS = {
    "erd": "Erdbeschwörung",
    "feuer": "Feuerbeschwörung",
    "luft": "Luftbeschwörung",
    "magie": "Magiebeschwörung",
    "wasser": "Wasserbeschwörung",
    "beherrschung": "Beherrschung",
}
PATTERN = re.compile(
    r"^vn_anfaelligkeit_gegen_(erd|feuer|luft|magie|wasser)beschwoerung_([12])$"
    r"|^vn_anfaelligkeit_gegen_(beherrschung)_([12])$"
)


def main():
    workbook = load_workbook(WERTE, read_only=False, data_only=False)
    worksheet = workbook["Werte"]
    headers = {cell.value: cell.column for cell in worksheet[1] if cell.value}
    updated = []

    for row in range(2, worksheet.max_row + 1):
        reference = worksheet.cell(row, headers["Referenz"]).value
        match = PATTERN.match(str(reference or ""))
        if match:
            school_slug = match.group(1) or match.group(3)
            stage = int(match.group(2) or match.group(4))
            school = SCHOOL_LABELS[school_slug]
            factor = "1,5" if stage == 1 else "2"
            group = f"Anfälligkeit Stufe {stage}"
            effect = (
                f"Der Charakter erleidet durch {school} SP × {factor}. "
                f"Aus der Gruppe „{group}“ kann nur ein Eintrag gewählt werden."
            )
            worksheet.cell(row, headers["Kosten"], -200 if stage == 1 else -500)
            worksheet.cell(row, headers["Verfuegbarkeit"], "E")
            worksheet.cell(row, headers["Info"], f"Auswahlgruppe: {group}; maximal 1")
            worksheet.cell(row, headers["Wirkung"], effect)
            updated.append((row, reference, effect))
            continue

        if reference == "vn_anfaelligkeit_gegen_profane_waffen":
            group = "Anfälligkeit gegen profane Waffen"
            effect = (
                "Der Charakter erleidet durch profane Waffen SP × 2. "
                f"Aus der Gruppe „{group}“ kann nur ein Eintrag gewählt werden."
            )
            worksheet.cell(row, headers["Kosten"], -500)
            worksheet.cell(row, headers["Verfuegbarkeit"], "E")
            worksheet.cell(row, headers["Info"], f"Auswahlgruppe: {group}; maximal 1")
            worksheet.cell(row, headers["Wirkung"], effect)
            updated.append((row, reference, effect))

    expected = 13
    if len(updated) != expected:
        raise RuntimeError(f"Erwartet: {expected} Anfälligkeiten; gefunden: {len(updated)}")

    workbook.save(WERTE)
    print(f"updated={len(updated)}")
    for row, reference, effect in updated:
        print(row, reference, effect)


if __name__ == "__main__":
    main()
