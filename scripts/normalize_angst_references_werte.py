import re
import sys
from pathlib import Path


ROOT = Path(r"E:\Das Western Rollenspiel\LLM")
sys.path.insert(0, str(ROOT / ".python-deps"))

from openpyxl import load_workbook


WERTE = ROOT / "werte 0.8-claude.xlsx"
LEVELS = {
    "unbehagen": 5,
    "nervositaet": 10,
    "furcht": 15,
    "angst": 20,
    "panik": 25,
    "phobie": 30,
}
THEMES = {
    "magie": "Magie",
    "verlusten_humanoide": "Verlusten (Humanoide)",
    "arachnophobie": "Arachnophobie",
    "bestimmtes_domestiziertes_tier": "Bestimmtes domestiziertes Tier",
    "hundeartige": "Hundeartige",
    "hunger": "Hunger",
    "laecherlichkeit": "Lächerlichkeit",
    "magische_gegenstaende": "Magische Gegenstände",
    "magische_wesen": "Magische Wesen",
    "meer": "Meer",
    "raubkatzen": "Raubkatzen",
    "schmutz": "Schmutz",
    "tiere": "Tiere",
    "untote": "Untote",
    "voegel": "Vögel",
    "wasser": "Wasser",
    "wilde_tiere": "Wilde Tiere",
    "xenophobie": "Xenophobie",
}


def replace_reference_tokens(value, replacements):
    if not isinstance(value, str):
        return value
    for old, new in replacements.items():
        value = re.sub(
            rf"(?<![A-Za-z0-9_]){re.escape(old)}(?![A-Za-z0-9_])",
            new,
            value,
        )
    return value


def main():
    workbook = load_workbook(WERTE, read_only=False, data_only=False)
    worksheet = workbook["Werte"]
    headers = {cell.value: cell.column for cell in worksheet[1] if cell.value}

    replacements = {}
    fear_rows = []
    for row in range(2, worksheet.max_row + 1):
        reference = str(worksheet.cell(row, headers["Referenz"]).value or "")
        match = re.fullmatch(
            r"vn_(unbehagen|nervositaet|furcht|angst|phobie)_(.+)",
            reference,
        )
        if not match:
            continue
        level_name, theme = match.groups()
        if theme not in THEMES:
            continue
        value = LEVELS[level_name]
        new_reference = f"vn_angst_{theme}_{value}"
        replacements[reference] = new_reference
        fear_rows.append((row, new_reference, theme, value))

    if len(fear_rows) != 90:
        raise RuntimeError(f"Erwartet: 90 Angststufen; gefunden: {len(fear_rows)}")
    if len(set(replacements.values())) != 90:
        raise RuntimeError("Die neuen Angst-Referenzen sind nicht eindeutig.")

    for sheet in workbook.worksheets:
        for cells in sheet.iter_rows():
            for cell in cells:
                updated = replace_reference_tokens(cell.value, replacements)
                if updated != cell.value:
                    cell.value = updated

    for row, reference, theme, value in fear_rows:
        group = f"angst_{theme}"
        group_label = f"Angst: {THEMES[theme]}"
        old_info = str(worksheet.cell(row, headers["Info"]).value or "").strip()
        old_flag = str(worksheet.cell(row, headers["Flag"]).value or "").strip()
        exclusivity = f"Auswahlgruppe {group}; maximal 1 Stufe; Angstwert {value}."
        worksheet.cell(row, headers["Info"], f"{old_info} {exclusivity}".strip())
        worksheet.cell(row, headers["Parent"], group_label)
        exclusive_flag = f"EXKLUSIV_ANGST={group}"
        worksheet.cell(
            row,
            headers["Flag"],
            f"{old_flag} | {exclusive_flag}" if old_flag else exclusive_flag,
        )
        worksheet.cell(row, headers["Grad"], value)
        worksheet.cell(row, headers["Verfuegbarkeit"], "S")

    appearance_reference = "vn_aussehen_normal"
    appearance_row = next(
        (
            row
            for row in range(2, worksheet.max_row + 1)
            if worksheet.cell(row, headers["Referenz"]).value == appearance_reference
        ),
        worksheet.max_row + 1,
    )
    appearance_values = {
        "Referenz": appearance_reference,
        "Kategorie": "Vor- und Nachteile",
        "Beschreibung": "Aussehen: Normal",
        "Info": "Neutrale Aussehensstufe.",
        "Parent": "Aussehen",
        "Art": "Auswahl",
        "Kosten": 0,
        "Verfuegbarkeit": "E",
    }
    for header, value in appearance_values.items():
        worksheet.cell(appearance_row, headers[header], value)

    workbook.save(WERTE)
    print(f"fear_references_migrated={len(fear_rows)}")
    print(f"appearance_normal_row={appearance_row}")


if __name__ == "__main__":
    main()
