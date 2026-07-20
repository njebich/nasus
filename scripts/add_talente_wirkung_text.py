"""
Importiert die Beschreibung-Spalte (F) aus "Talente-Wirkung-chatgpt.xlsx" (ChatGPT-
Wirkungsanalyse aller 148 Talente) in die "Wirkung"-Spalte der Talente-Zeilen im
werte-0.8-claude.xlsx-Sheet "Werte" - fuer das Tooltip-System (RuleEntry.wirkung,
siehe generate_data_ts.py Zeile 59), analog zur bereits vorhandenen Wirkung-Befuellung
bei Vor- und Nachteile.

Matching: Talente-Wirkung R-Spalte ("Werte-Referenz(en)") enthaelt bei gruppierten
Zeilen (Schulvarianten, S/"Referenzanzahl" > 1) mehrere newline-getrennte Referenzen,
auf die dieselbe Beschreibung angewendet wird. Referenzen im Werte-Sheet koennen mit
"#"-Praefix deaktiviert sein (z.B. "#talente_ladeschuetze_schleuder"); das Matching
vergleicht daher ohne fuehrendes "#".

Alle 177 Talente-Zeilen im Werte-Sheet werden von den 144 Analyse-Zeilen abgedeckt
(gruppierte Zeilen zaehlen mehrfach), keine Ueberschneidungen, keine vorhandene
Wirkung-Spalte wird ueberschrieben (verifiziert vor dem Schreiben).
"""
import sys
from pathlib import Path

ROOT = Path(r"E:\Das Western Rollenspiel\LLM")
sys.path.insert(0, str(ROOT / ".python-deps"))

from openpyxl import load_workbook

SOURCE = ROOT / "Talente-Wirkung-chatgpt.xlsx"
WERTE = ROOT / "werte 0.8-claude.xlsx"


def read_source_rows(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Talente"]
    rows = []
    for r in range(2, 146):
        beschreibung = ws.cell(row=r, column=6).value  # F
        refs_raw = ws.cell(row=r, column=18).value  # R
        if not refs_raw:
            continue
        refs = [x.strip() for x in str(refs_raw).split("\n") if x.strip()]
        rows.append((beschreibung, refs))
    return rows


def main():
    source_rows = read_source_rows(SOURCE)

    workbook = load_workbook(WERTE, read_only=False, data_only=False)
    worksheet = workbook["Werte"]
    headers = {cell.value: cell.column for cell in worksheet[1] if cell.value}
    ref_col = headers["Referenz"]
    kat_col = headers["Kategorie"]
    wirkung_col = headers["Wirkung"]

    bare_to_row = {}
    for row in range(2, worksheet.max_row + 1):
        if worksheet.cell(row, kat_col).value != "Talente":
            continue
        ref = worksheet.cell(row, ref_col).value
        if not ref:
            continue
        bare = ref[1:] if ref.startswith("#") else ref
        bare_to_row[bare] = row

    written = 0
    missing = []
    already_filled = []
    for beschreibung, refs in source_rows:
        for ref in refs:
            row = bare_to_row.get(ref)
            if row is None:
                missing.append(ref)
                continue
            existing = worksheet.cell(row, wirkung_col).value
            if existing:
                already_filled.append(ref)
                continue
            worksheet.cell(row, wirkung_col, beschreibung)
            written += 1

    if missing:
        raise SystemExit(f"Nicht gefundene Referenzen: {missing}")
    if already_filled:
        raise SystemExit(f"Bereits befuellte Wirkung-Zellen (nicht ueberschrieben): {already_filled}")

    workbook.save(WERTE)
    print(f"{written} Wirkung-Zellen geschrieben.")


if __name__ == "__main__":
    main()
