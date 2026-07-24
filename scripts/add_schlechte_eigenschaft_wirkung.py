"""
Befuellt die "Wirkung"-Spalte fuer die 10 "Schlechte Eigenschaft: X"-Nachteile
(Referenzen vn_schlechte_eigenschaft_<x>, Zeilen 97-106 im Sheet "Werte") - fuer das
Tooltip-System (RuleEntry.wirkung, siehe generate_data_ts.py), analog zu
add_talente_wirkung_text.py. Text-Vorlage vom Nutzer 2026-07-24 (am Beispiel
Ausstrahlung), hier pro Eigenschaft mit korrektem grammatischem Geschlecht
generalisiert ("Mut" ist maskulin, alle anderen neun Eigenschaften feminin).

Ueberschreibt KEINE bereits befuellte Wirkung-Zelle (wie add_talente_wirkung_text.py).
"""
import sys
from pathlib import Path

ROOT = Path(r"E:\Das Western Rollenspiel\LLM")
sys.path.insert(0, str(ROOT / ".python-deps"))

from openpyxl import load_workbook

WERTE = ROOT / "werte 0.8-claude.xlsx"

# Referenz-Suffix -> (Anzeigename, "eine unterdurchschnittliche"/"einen unterdurchschnittlichen")
EIGENSCHAFTEN = {
    "ausstrahlung": ("Ausstrahlung", "eine unterdurchschnittliche"),
    "intelligenz": ("Intelligenz", "eine unterdurchschnittliche"),
    "mut": ("Mut", "einen unterdurchschnittlichen"),
    "sinnesschaerfe": ("Sinnesschärfe", "eine unterdurchschnittliche"),
    "willenskraft": ("Willenskraft", "eine unterdurchschnittliche"),
    "athletik": ("Athletik", "eine unterdurchschnittliche"),
    "geschicklichkeit": ("Geschicklichkeit", "eine unterdurchschnittliche"),
    "konstitution": ("Konstitution", "eine unterdurchschnittliche"),
    "schnelligkeit": ("Schnelligkeit", "eine unterdurchschnittliche"),
    "staerke": ("Stärke", "eine unterdurchschnittliche"),
}

REST = (
    " Das absolute Maximum dieser Eigenschaft ist 9 (ab Kreis 3, davor darf die Eigenschaft 7 "
    "nicht übersteigen). Dieser Nachteil überschreibt die Spezies-Vorgaben. Die Eigenschaft ist "
    "grundsätzlich nicht übersteigerbar."
)


def main():
    workbook = load_workbook(WERTE, read_only=False, data_only=False)
    worksheet = workbook["Werte"]
    headers = {cell.value: cell.column for cell in worksheet[1] if cell.value}
    ref_col = headers["Referenz"]
    wirkung_col = headers["Wirkung"]

    written = 0
    already_filled = []
    missing = []
    remaining = dict(EIGENSCHAFTEN)

    for row in range(2, worksheet.max_row + 1):
        ref = worksheet.cell(row, ref_col).value
        if not ref or not str(ref).startswith("vn_schlechte_eigenschaft_"):
            continue
        suffix = str(ref)[len("vn_schlechte_eigenschaft_"):]
        entry = remaining.pop(suffix, None)
        if entry is None:
            continue
        name, artikel = entry
        text = f"Dieser Charakter hat {artikel} {name}.{REST}"
        existing = worksheet.cell(row, wirkung_col).value
        if existing:
            already_filled.append(ref)
            continue
        worksheet.cell(row, wirkung_col, text)
        written += 1

    if remaining:
        missing.extend(remaining.keys())
    if missing:
        raise SystemExit(f"Nicht gefundene Referenzen: {missing}")
    if already_filled:
        raise SystemExit(f"Bereits befuellte Wirkung-Zellen (nicht ueberschrieben): {already_filled}")

    workbook.save(WERTE)
    print(f"{written} Wirkung-Zellen geschrieben.")


if __name__ == "__main__":
    main()
