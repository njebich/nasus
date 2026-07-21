"""
Traegt einen Entwickeln-Eintrag fuer das noch fehlende Waffen-Loadout-System ein (Nutzer
2026-07-21, im Rahmen von "talente-add-implementation-charaktererstellung.txt" / Kampf mit
zwei Waffen Stufe 1-4): der Kampf-Tab zeigt seit dieser Session pro 1H-Waffe nur noch ein
Eignungs-Flag ("2-Waffen"-Spalte, WK <= Kappungswert der hoechsten besessenen Stufe, siehe
views/kampf.ts's getZweiWaffenCap), aber KEINE kombinierte AT/PA-Mechanik fuer ein konkretes
Waffenpaar - das haengt von einer situativen Zwei-Waffen-Wahl ab, die es im Chargen-Tool noch
nicht gibt. Nutzer-Vorgabe: "create loadout system for NK, NK/Schild, NK/NK, NK/FK, FK, FK/FK".
"""
import sys
from pathlib import Path
from datetime import date

ROOT = Path(r"E:\Das Western Rollenspiel\LLM")
sys.path.insert(0, str(ROOT / ".python-deps"))

from openpyxl import load_workbook

WERTE = ROOT / "werte 0.8-claude.xlsx"
HEUTE = "2026-07-21"

ENTRIES = [
    {
        "Bereich": "Kampf/Loadout",
        "Thema": "Waffen-Loadout-System (NK, NK/Schild, NK/NK, NK/FK, FK, FK/FK) fehlt noch",
        "Beschreibung": (
            "Der Kampf-Tab behandelt jede besessene Waffe/jeden Schild als eigene, unabhaengige "
            "Zeile mit eigenem AT/PA-Pool (siehe views/kampf.ts's buildNahkampfRows) - es gibt "
            "keinen Begriff eines aktiven 'Loadouts' (welche Waffe(n) der Charakter GERADE fuehrt). "
            "Das wird fuer mehrere Talente/Situationen gebraucht, die auf einer KOMBINATION aus "
            "zwei gleichzeitig gefuehrten Waffen basieren, nicht auf einer einzelnen: "
            "(a) Kampf mit zwei Waffen Stufe 1-4 (NK/NK) - kombinierter WK-Cap, addierte n-Mods/"
            "Mindeststaerke, 1,5x-WK-Attacke mit der staerkeren Waffe, addierte WK-Parade. Diese "
            "Session hat als Zwischenloesung nur ein Eignungs-Flag pro 1H-Waffe ergaenzt (Spalte "
            "'2-Waffen' in der Nahkampf-Tabelle, WK <= Kappungswert der hoechsten besessenen Stufe, "
            "siehe getZweiWaffenCap) - die eigentliche Paar-Mechanik fehlt bewusst noch. "
            "(b) Linkshaendig Pistolenschiessen / Mit Zwei Pistolen schiessen (FK/FK) - Probenwert-"
            "Modifikator haengt davon ab, ob eine oder zwei Pistolen gleichzeitig geschossen werden. "
            "(c) NK/Schild ist die bereits bestehende Schild-Kampfregeln-Luecke (siehe frueherer "
            "Entwickeln-Eintrag 'Nahkampf/Schilde'). "
            "(d) NK/FK ist der Fall 'eine Waffe in jeder Hand, eine Nah- eine Fernkampfwaffe' "
            "(z.B. Messer + Pistole), bisher nirgends modelliert. "
            "Nutzer-Vorgabe 2026-07-21: ein generisches Loadout-System bauen, das alle sechs "
            "Kombinationen abdeckt (NK, NK/Schild, NK/NK, NK/FK, FK, FK/FK) statt fuer jede "
            "einzeln eine Sonderloesung - noch nicht spezifiziert (UI: Waffenpaar-Auswahl je "
            "Kampfsituation? Persistiert im CharacterState oder nur Kampfrunden-lokal?), daher "
            "als offener Architektur-Punkt geloggt statt spekulativ gebaut."
        ),
        "Status": "Offen",
        "Hinzugefuegt": HEUTE,
    },
]


def main():
    workbook = load_workbook(WERTE, read_only=False, data_only=False)
    worksheet = workbook["Entwickeln"]
    headers = {cell.value: cell.column for cell in worksheet[1] if cell.value}

    last_row = 0
    for row in worksheet.iter_rows(min_row=2):
        if any(c.value is not None for c in row):
            last_row = row[0].row

    row = last_row + 1
    for entry in ENTRIES:
        for header, value in entry.items():
            worksheet.cell(row, headers[header], value)
        row += 1

    workbook.save(WERTE)
    print(f"{len(ENTRIES)} Entwickeln-Eintraege geschrieben (Zeilen {last_row + 1}-{row - 1}).")


if __name__ == "__main__":
    main()
