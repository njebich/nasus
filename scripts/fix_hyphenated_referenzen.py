"""
Einmaliges Bereinigungsskript: ersetzt Bindestriche in Referenz-Namen durch Unterstriche.

Hintergrund: Referenz-Namen sollen reine snake_case-Bezeichner sein (nur [A-Za-z0-9_]),
damit der Formel-Parser Bindestrich eindeutig als Subtraktions-Operator lesen kann
(z.B. "ep_gesamt-ep_start"). Referenzen wie "nk_spez_hiebwaffen_beidhand-hiebwaffen"
widersprechen dem und zwingen den Parser sonst zu einer fehleranfaelligen
Kontext-Heuristik ("verlaengere nur wenn bekannte Referenz").

Aufruf:
    python scripts/fix_hyphenated_referenzen.py "werte 0.7-claude.xlsx"

Legt vorher automatisch eine Sicherheitskopie an (siehe backup_werte.py).
"""
import sys
import re
from pathlib import Path

import openpyxl

from backup_werte import backup


def main(path_str: str) -> None:
    path = Path(path_str)
    backup_path = backup(str(path))
    print(f"Sicherheitskopie angelegt: {backup_path}")

    wb = openpyxl.load_workbook(path, data_only=False)
    ws = wb["Werte"]
    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v:
            headers[v.strip()] = c

    ref_col = headers["Referenz"]
    flag_col = headers["Flag"]
    formula_cols = [headers[c] for c in ("Formel", "Pool", "Kosten") if c in headers]

    # 1) Alle betroffenen Referenzen sammeln + umbenennen in Spalte A.
    renamed: dict[str, str] = {}
    for r in range(2, ws.max_row + 1):
        cell = ws.cell(row=r, column=ref_col)
        old = cell.value
        if not old or "-" not in old:
            continue
        new = old.replace("-", "_")
        renamed[old] = new
        cell.value = new

        flag_cell = ws.cell(row=r, column=flag_col)
        note = "Referenz enthielt Bindestrich (mehrdeutig mit Subtraktion) - zu Unterstrich normalisiert."
        flag_cell.value = f"{flag_cell.value} | {note}" if flag_cell.value else note

    print(f"{len(renamed)} Referenz(en) umbenannt (Bindestrich -> Unterstrich).")

    # 2) Alle Formel/Pool/Kosten-Zellen im GESAMTEN Sheet nach Verwendung der ALTEN
    #    Namen durchsuchen (exakte Token-Grenzen, kein Teilstring-Treffer) und ersetzen.
    updated_formula_cells = 0
    for r in range(2, ws.max_row + 1):
        for col in formula_cols:
            cell = ws.cell(row=r, column=col)
            val = cell.value
            if not isinstance(val, str):
                continue
            new_val = val
            for old, new in renamed.items():
                pattern = r"(?<![A-Za-z0-9_])" + re.escape(old) + r"(?![A-Za-z0-9_])"
                new_val = re.sub(pattern, new, new_val)
            if new_val != val:
                cell.value = new_val
                updated_formula_cells += 1
                print(f"  Zeile {r}, Spalte {col}: \"{val}\" -> \"{new_val}\"")

    print(f"{updated_formula_cells} Formel/Pool/Kosten-Zelle(n) aktualisiert.")

    wb.save(path)
    print(f"Gespeichert: {path}")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print('Aufruf: python scripts/fix_hyphenated_referenzen.py "werte 0.7-claude.xlsx"')
        sys.exit(1)
    main(sys.argv[1])
