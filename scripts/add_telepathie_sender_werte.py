import sys
from pathlib import Path


ROOT = Path(r"E:\Das Western Rollenspiel\LLM")
sys.path.insert(0, str(ROOT / ".python-deps"))

from openpyxl import load_workbook


WERTE = ROOT / "werte 0.8-claude.xlsx"
EFFECT = "Nur für Charaktere der Spezies Elementare."
ENTRIES = {
    "vn_telepathie_sender": ("Telepathie: Sender", 3),
    "vn_telepathie_empfaenger": ("Telepathie: Empfänger", 1),
}


def main():
    workbook = load_workbook(WERTE, read_only=False, data_only=False)
    worksheet = workbook["Werte"]
    headers = {cell.value: cell.column for cell in worksheet[1] if cell.value}
    rows_by_reference = {
        worksheet.cell(row, headers["Referenz"]).value: row
        for row in range(2, worksheet.max_row + 1)
    }
    changed = []
    for reference, (description, availability) in ENTRIES.items():
        row = rows_by_reference.get(reference, worksheet.max_row + 1)
        values = {
            "Referenz": reference,
            "Kategorie": "Vor- und Nachteile",
            "Beschreibung": description,
            "Info": "Ausschließlich für die Spezies Elementare wählbar.",
            "Parent": "Elementar",
            "Art": "Auswahl",
            "Flag": "SPEZIES_ELEMENTARE",
            "Kosten": 1,
            "Verfuegbarkeit": availability,
            "Wirkung": EFFECT,
        }
        for header, value in values.items():
            worksheet.cell(row, headers[header], value)
        rows_by_reference[reference] = row
        changed.append((row, reference))
    workbook.save(WERTE)
    for row, reference in changed:
        print(row, reference)


if __name__ == "__main__":
    main()
