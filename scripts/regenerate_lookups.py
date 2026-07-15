"""
Erzeugt die generierten Nachschlage-Tabellen in einer werte-*.xlsx neu,
direkt aus "NN NPC-Rechner 0.76.xlsx" (der einzigen Referenzdatei ab jetzt).

Betrifft zwei Sheets in der werte-Datei:
  - "EP-Stufe-Kreis"   (aus Ressourcen!H17:J79)
  - "WHK-Spez-Kosten"  (aus Ressourcen!A72:C110)

Diese zwei Sheets NIEMALS von Hand bearbeiten - sie werden hier immer
komplett neu geschrieben. Alles andere in der werte-Datei bleibt unangetastet.

Aufruf (Quelle liegt im selben Ordner):
    python scripts/regenerate_lookups.py "werte 0.7-claude.xlsx"

Mit eigener Quelldatei:
    python scripts/regenerate_lookups.py "werte 0.7-claude.xlsx" --quelle "NN NPC-Rechner 0.76.xlsx"

Legt vorher automatisch eine Sicherheitskopie an (siehe backup_werte.py).
"""
import sys
import argparse
from pathlib import Path

import openpyxl

from backup_werte import backup

DEFAULT_QUELLE = "NN NPC-Rechner 0.76.xlsx"


def extract_ep_stufe_kreis(quelle_wb):
    ws = quelle_wb["Ressourcen"]
    rows = []
    for r in range(17, 80):
        ep = ws.cell(row=r, column=8).value
        stufe = ws.cell(row=r, column=9).value
        kreis = ws.cell(row=r, column=10).value
        if ep is None:
            continue
        if not float(kreis).is_integer():
            # bekannter Fehler in der Quelle bei Stufe 53-55 (7.285.../7.392.../7.499...
            # statt 7, siehe Nachbarstufen 49-52 und 56-63) - korrigieren
            kreis = 7
        rows.append((int(ep), int(stufe), int(kreis)))
    return rows


def extract_whk_spez_kosten(quelle_wb):
    ws = quelle_wb["Ressourcen"]
    rows = []
    for r in range(72, 111):
        wert = ws.cell(row=r, column=1).value
        inc = ws.cell(row=r, column=2).value
        cum = ws.cell(row=r, column=3).value
        if wert is None:
            continue
        rows.append((int(wert), int(inc), int(cum)))
    return rows


def write_ep_stufe_kreis(werte_wb, rows):
    if "EP-Stufe-Kreis" in werte_wb.sheetnames:
        del werte_wb["EP-Stufe-Kreis"]
    ws = werte_wb.create_sheet("EP-Stufe-Kreis")
    ws.append(["EP ab", "Stufe", "Kreis"])
    ws.append([0, 0, 0])
    for ep, stufe, kreis in rows:
        ws.append([ep, stufe, kreis])


def write_whk_spez_kosten(werte_wb, rows):
    if "WHK-Spez-Kosten" in werte_wb.sheetnames:
        del werte_wb["WHK-Spez-Kosten"]
    ws = werte_wb.create_sheet("WHK-Spez-Kosten")
    ws.append(["Wert", "EP-Kosten (Schritt)", "Gesamt-EP"])
    ws.append([0, 0, 0])
    for wert, inc, cum in rows:
        ws.append([wert, inc, cum])


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("werte_datei")
    parser.add_argument("--quelle", default=None, help="Pfad zur NPC-Rechner-Quelldatei")
    args = parser.parse_args()

    werte_path = Path(args.werte_datei)
    quelle_path = Path(args.quelle) if args.quelle else werte_path.parent / DEFAULT_QUELLE

    if not quelle_path.exists():
        raise SystemExit(f"Quelldatei nicht gefunden: {quelle_path}")

    backup_path = backup(str(werte_path))
    print(f"Sicherheitskopie angelegt: {backup_path}")

    quelle_wb = openpyxl.load_workbook(quelle_path, data_only=True)
    werte_wb = openpyxl.load_workbook(werte_path, data_only=False)

    ep_rows = extract_ep_stufe_kreis(quelle_wb)
    write_ep_stufe_kreis(werte_wb, ep_rows)
    print(f"EP-Stufe-Kreis neu geschrieben: {len(ep_rows)} Zeilen")

    whk_rows = extract_whk_spez_kosten(quelle_wb)
    write_whk_spez_kosten(werte_wb, whk_rows)
    print(f"WHK-Spez-Kosten neu geschrieben: {len(whk_rows)} Zeilen")

    werte_wb.save(werte_path)
    print(f"Gespeichert: {werte_path}")
    print(f"Quelle: {quelle_path}")


if __name__ == "__main__":
    main()
