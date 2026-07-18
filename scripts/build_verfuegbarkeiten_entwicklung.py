import json
import sys
from pathlib import Path


ROOT = Path(r"E:\Das Western Rollenspiel\LLM")
sys.path.insert(0, str(ROOT / ".python-deps"))

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


OUTPUT = ROOT / "Verfuegbarkeiten-Herkunftsorte-Entwicklung.xlsx"
DATA = ROOT / "src" / "data" / "equipment"

NAVY = "1F4E78"
BLUE = "D9EAF7"
LIGHT_BLUE = "EAF3F8"
GREEN = "E2F0D9"
YELLOW = "FFF2CC"
RED = "F4CCCC"
PURPLE = "E4DFEC"
GRAY = "E7E6E6"
WHITE = "FFFFFF"
DARK = "1F1F1F"
THIN_GRAY = Side(style="thin", color="D9E1F2")

STATUS_VALUES = '"BESTÄTIGT,ABGELEITET,OFFEN,SONDERFALL,NICHT KAUFBAR"'
PEOPLE_MODE_VALUES = '"ALLE,AUSWAHL"'
MERCHANT_SPECIALIZATION_EXCLUDED = {"Miete", "Post", "Reisekosten", "Tavernen-Preise", "Zoll"}
CHAIN_ARMOR_MANUFACTURERS = "Dalkini | Draw | Elfen | Goblins | Orks | Zwerge"
PLATE_ARMOR_MANUFACTURERS = "Dalkini | Draw | Elfen | Goblins | Orks | Trolle | Zwerge"
ARTIFACT_MANUFACTURERS = "Draw | Elfen | Gnome | Goblins | Orks | Zwerge"
METAL_SHIELD_MANUFACTURERS = "Dalkini | Draw | Elfen | Goblins | Orks | Trolle | Zwerge"
CHAIN_ARMOR_NAMES = {
    "Eisen-Kettenpanzer", "Stahl-Kettenpanzer", "Schwerer Eisen-Kettenpanzer",
    "Faltstahl-Kettenpanzer", "Schwerer Stahl-Kettenpanzer", "Schwerer Faltstahl-Kettenpanzer",
}
PLATE_ARMOR_NAMES = {
    "Eisenpanzer", "Stahlpanzer", "Schwerer Eisenpanzer", "Schwerer Stahlpanzer",
    "Faltstahlpanzer", "Schwerer Faltstahlpanzer",
}
METAL_SHIELD_MATERIALS = {
    "Bronze", "Eisen", "Feineisen", "Stahl", "Qualitaetsstahl", "Faltstahl",
    "Mithril", "Adamandit", "Alchemistensilb.", "Nasium",
}
METAL_WEAPON_MATERIALS = {
    "Bronze", "Eisen", "Stahl", "Qualitätsstahl", "Faltstahl", "Mithril",
    "Adamandit", "Alchemistensilb.", "Nasium",
}
SHIELD_CRAFT_AVAILABILITY = {
    "Ausschuss": (1, 1),
    "Goblin Massenfab.": (1, 1),
    "Massenfabrikation": (1, 1),
    "Gesellenarbeit": (1, 1),
    "Meisterarbeit": (2, 3),
    "Großmeisterarbeit": (3, 5),
    "Einzelstück": (7, 7),
}
GLOBAL_MATERIAL_AVAILABILITY = {
    "Holz": (1, 1),
    "Hartholz": (2, 2),
    "Leder": (1, 1),
    "Spinnenwebe": (4, 4),
    "Chitin": (5, 2),
    "Drachenschuppe": ("M", "M"),
    "Stein": (1, 1),
    "Schwarzfels": (5, 1),
    "Diamantspat": (7, 7),
    "Bronze": (1, 1),
    "Eisen": (1, 1),
    "Feineisen": (1, 1),
    "Stahl": (1, 2),
    "Qualitätsstahl": (2, 4),
    "Faltstahl": (3, 5),
    "Mithril": (7, 7),
    "Adamandit": ("M", "M"),
    "Alchemistensilber": (5, 6),
    "Vulkanglas": (7, 3),
    "Nasium": (7, 7),
    "Knochen": (1, 1),
    "Kolharz": (5, 7),
}
GLOBAL_MATERIAL_ORDER = tuple(GLOBAL_MATERIAL_AVAILABILITY)
SOURCE_MATERIAL_CANONICAL = {
    "Hart Holz": "Hartholz",
    "Drachensch.": "Drachenschuppe",
    "Schwarzfells": "Schwarzfels",
    "Diamantspart": "Diamantspat",
    "Qualitaetsstahl": "Qualitätsstahl",
    "Alchemistensilb.": "Alchemistensilber",
    "Kolhartz": "Kolharz",
    "Adamantit": "Adamandit",
}
SHIELD_COVER_AVAILABILITY = {
    "Stoff": (1, 1),
    "Spinnenwebe": (4, 4),
    "Leder": (1, 1),
    "Drachenschuppe": ("M", "M"),
    "Kohlharz": (4, 4),
}
WEAPON_ADJUSTMENT_AVAILABILITY = {
    "Von der Stange": (1, 1),
    "angepasst": (2, 3),
    "Perfekt angepasst": (3, 5),
}
WEAPON_SHAFT_BASE_MATERIAL = {
    "Eisen Verstärkt": "Eisen",
    "Eisen Voll": "Eisen",
    "Stahl Verstärkt": "Stahl",
    "Stahl Voll": "Stahl",
    "Qualitätsstahl": "Qualitätsstahl",
    "Qualitätsstahl Voll": "Qualitätsstahl",
    "Faltstahl Verstärkt": "Faltstahl",
    "Faltstahl Voll": "Faltstahl",
    "Mithril Verstärkt": "Mithril",
    "Mithril Voll": "Mithril",
    "Adamandit Verst.": "Adamandit",
    "Adamantit Voll": "Adamandit",
    "Bronze Verstärkt": "Bronze",
    "Bronze Voll": "Bronze",
}
NON_PURCHASABLE_WEAPON_PEOPLE = {
    "andere Voelker": "Dalkini | Draw | Elfen | Goblins | Indianer | Zwerge",
    "Gnom": "Gnome",
    "Ork": "Orks",
    "Troll": "Trolle",
    "Zentaur": "Zentauren",
    "Katzenmensch": "Katzen",
    "Dalkini, Zwerg, Ork": "Dalkini | Orks | Zwerge",
    "Elf": "Elfen",
    "Goblin": "Goblins",
    "alle Voelker": "ALLE",
    "Drow": "Draw",
}
CULTURAL_WEAPON_PEOPLE = {
    "Armklingen elfisch": "Elfen",
    "Streitaxt elfisch": "Elfen",
    "Bat'leth  orkisch": "Orks",
    "Dolch ork, Blutrinne": "Orks",
    "Dolch orkisch, breit": "Orks",
    "Dolch orkisch, mit 3 eingedrehten Klingen": "Orks",
    "Kriegsbeil orkisch": "Orks",
    "Ork`sche": "Orks",
    "Ork-Flegel": "Orks",
    "Orkischer Kampfhandschuh": "Orks",
    "Streitaxt orkisch": "Orks",
    "Doppelaxt zwergisch": "Zwerge",
    "Doppelbeil Zw. kurz": "Zwerge",
    "Streitaxt goblinisch": "Goblins",
    "Dornenkeule  troll,": "Trolle",
    "Keule  troll,, eisenbeschlagen": "Trolle",
    "Trollaxt": "Trolle",
    "Trollschlegel": "Trolle",
}

GLOBAL_METAL_MATERIALS = {
    "Bronze", "Eisen", "Feineisen", "Stahl", "Qualitätsstahl", "Faltstahl",
    "Mithril", "Adamandit", "Alchemistensilber", "Nasium",
}


def canonical_material_name(name):
    return SOURCE_MATERIAL_CANONICAL.get(name, name)


def global_material_people(name):
    if name in {"Mithril", "Nasium"}:
        return "Elfen | Zwerge"
    if name == "Spinnenwebe":
        return "Draw"
    if name == "Kolharz":
        return "Zentauren"
    if name == "Alchemistensilber":
        return "Dalkini | Elfen | Goblins | Orks | Zwerge"
    if name == "Diamantspat":
        return "Draw"
    if name in {"Schwarzfels", "Vulkanglas"}:
        return "Indianer | Katzen"
    if name in GLOBAL_METAL_MATERIALS:
        return METAL_SHIELD_MANUFACTURERS
    return "ALLE"


def read_json(name):
    return json.loads((DATA / name).read_text(encoding="utf-8"))


def is_availability(value):
    return str(value).strip() in {"1", "2", "3", "4", "5", "6", "7"}


def is_master_availability(value):
    return str(value or "").strip().casefold() in {"m", "meister"}


def normalize_people(raw):
    value = str(raw or "").strip()
    if not value:
        return "", "", "OFFEN"
    if value.casefold() in {"alle", "allgemein", "jedes volk"}:
        return "ALLE", "", "ABGELEITET"
    return "AUSWAHL", value, "ABGELEITET"


def add_title(ws, title, subtitle, end_col):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
    cell = ws.cell(1, 1, title)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.font = Font(name="Aptos Display", size=18, bold=True, color=WHITE)
    cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 30
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_col)
    cell = ws.cell(2, 1, subtitle)
    cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    cell.font = Font(name="Aptos", size=10, italic=True, color=DARK)
    cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[2].height = 34
    ws.sheet_view.showGridLines = False


def style_header(ws, row, start_col, end_col):
    for cell in ws.iter_cols(min_col=start_col, max_col=end_col, min_row=row, max_row=row):
        c = cell[0]
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.font = Font(name="Aptos", bold=True, color=WHITE)
        c.alignment = Alignment(wrap_text=True, vertical="center")
        c.border = Border(bottom=Side(style="medium", color=NAVY))
    ws.row_dimensions[row].height = 28


def add_table(ws, start_row, end_row, end_col, name):
    if end_row < start_row + 1:
        return
    from openpyxl.utils import get_column_letter

    ref = f"A{start_row}:{get_column_letter(end_col)}{end_row}"
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def set_widths(ws, widths):
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def add_status_formatting(ws, cell_range, first_cell):
    ws.conditional_formatting.add(
        cell_range,
        FormulaRule(formula=[f'{first_cell}="OFFEN"'], fill=PatternFill("solid", fgColor=YELLOW)),
    )
    ws.conditional_formatting.add(
        cell_range,
        FormulaRule(formula=[f'{first_cell}="SONDERFALL"'], fill=PatternFill("solid", fgColor=RED)),
    )
    ws.conditional_formatting.add(
        cell_range,
        FormulaRule(formula=[f'{first_cell}="BESTÄTIGT"'], fill=PatternFill("solid", fgColor=GREEN)),
    )


def add_list_validation(ws, formula, ranges):
    validation = DataValidation(type="list", formula1=formula, allow_blank=True)
    ws.add_data_validation(validation)
    for cell_range in ranges:
        validation.add(cell_range)


def add_availability_validation(ws, ranges):
    validation = DataValidation(type="list", formula1='"1,2,3,4,5,6,7,M"', allow_blank=True)
    validation.error = "Nur leere Zellen oder Verfügbarkeitswerte 1 bis 7 beziehungsweise M sind zulässig."
    validation.errorTitle = "Ungültige Verfügbarkeit"
    validation.showErrorMessage = True
    ws.add_data_validation(validation)
    for cell_range in ranges:
        validation.add(cell_range)


def create_overview(wb):
    ws = wb.create_sheet("Übersicht")
    add_title(
        ws,
        "Verfügbarkeiten und Herkunftsorte",
        "Entwicklungs- und Prüfarbeitsmappe. Keine dieser Änderungen wurde in werte 0.8-claude.xlsx geschrieben.",
        6,
    )
    rows = [
        ("Bestätigt", "Ausschließlich additive Verfügbarkeitsskala 1–7; niedriger ist besser."),
        ("Bestätigt", "Ruhm und Völkerbeziehungen bleiben außen vor und werden später reine Preismodifikatoren."),
        ("Bestätigt", "Heimat wird durch Herkunft ersetzt; Druckausgabe nur Ort, Region, AW/NW."),
        ("Bestätigt", "Orte und Charaktere werden lokal permanent gespeichert; Server erst nach App 1.0."),
        ("Bestätigt", "Artefakte: vorhandener Wert = AW/Grundwert; NW Grad 1–2 +0, Grad 3–5 +1, Grad 6–7 +2, maximal 7."),
        ("Bestätigt", "Skala 1–7 und M; M ist ausschließlich im Meister-Modul verfügbar. Vorhandenes 'Meister' wird als M normalisiert."),
        ("Bestätigt", "Siedlungsgröße getrennt für Rüstungen/Waffen und Artefakte; Artefaktwerte gegenüber Entwurf v0.3 jeweils +1."),
        ("Bestätigt", "Handelsstufe getrennt für Rüstungen/Waffen und Artefakte; Artefakte jeweils 1 Stufe schlechter."),
        ("Bestätigt", "Herstellungsort getrennt für Rüstungen/Waffen und Artefakte; Artefakte jeweils 1 Stufe schlechter."),
        ("Bestätigt", "Händlermodifikatoren bestätigt; spezialisierte Händler benötigen genau eine Warengruppe aus einem Dropdown."),
        ("Bestätigt", "Lokale Produktion ersetzt bei passender Warengruppe/optionalem Volk den allgemeinen Herstellungsort durch 'Herstellung direkt vor Ort'; kein Zusatzbonus."),
        ("Bestätigt", "Ortsbevölkerung: ALLE und Hauptspezies 0; etablierte Minderheit +1; keine übereinstimmende Spezies +3; beste Übereinstimmung zählt."),
        ("Bestätigt", "Zentrale Kaufsperre: Spieler dürfen effektive Verfügbarkeit 1–4 kaufen; 5–7, M und OFFEN sind gesperrt. M nur im Meister-Modul."),
        ("Bestätigt", "Meisterfreigabe: 5–7 oder M nur für konkrete Vergabe; Gegenstand/Ort unverändert; Charakter, Gegenstand, Ort, Wert und Zeitpunkt protokollieren; OFFEN nie übersteuern."),
        ("Bestätigt", "Komposition: jede gewählte Pflichtkomponente separat vollständig berechnen; numerisch Maximum, M vor numerisch, OFFEN vor allem; nicht addieren oder mitteln."),
        ("Bestätigt", "43 Warengruppen unverändert; 38 im durchsuchbaren Händler-Dropdown. Miete, Post, Reisekosten, Tavernen-Preise und Zoll sind nicht händlerspezialisierbar."),
        ("Bestätigt", "Persistenter Ortsgrundzustand plus optionaler Abenteuer-Ortszustand; verwerfen, weiterführen oder bewusst übernehmen; frühere Käufe unverändert."),
        ("Bestätigt", "Fehlende Verfügbarkeit bleibt leer und erhält Status OFFEN; niemals Platzhalter 0."),
        ("Bestätigt", "Völkerzuweisung: Modus ALLE oder AUSWAHL; eine Liste deckt Einfach- und Mehrfachauswahl ab."),
        ("Bestätigt", "Mehrere Händler je Ort; spezialisierte Händler je genau eine Warengruppe; pro Kauf niedrigster anwendbarer Modifikator, sonst 'Kein Laden'. Straitmor: 6 spezialisierte Händler."),
        ("Bestätigt", "Zwogón: für jede der 38 händlerspezialisierbaren Warengruppen ein eigener großer spezialisierter Händler."),
        ("Bestätigt", "Rüstungen: 6 Kettenbasen zusätzlich ohne Trolle; 6 Metallplattenbasen mit Trollen; beide ohne Katzen/Indianer/Zentauren/Gnome; übrige 16 ALLE."),
        ("Bestätigt", "Artefakte: 57 Grunddatensätze AUSWAHL Draw/Elfen/Gnome/Goblins/Orks/Zwerge; Vererbung an 749 Varianten; alle BESTÄTIGT."),
        ("Bestätigt", "Schilde: Kolhartz→Kolharz, Diamantspart→Diamantspat, Schwarzfells→Schwarzfels. Metallschilde ohne Katzen/Indianer/Zentauren/Gnome; Diamantspat Draw; Schwarzfels/Vulkanglas Indianer und Katzen; Kolharz/Kohlharz Zentauren; Goblin Massenfab. Goblins; übrige Komponenten ALLE."),
        ("Bestätigt", "Schildfertigung AW/NW: Ausschuss, Goblin Massenfab., Massenfabrikation und Gesellenarbeit 1/1; Meisterarbeit 2/3; Großmeisterarbeit 3/5; Einzelstück 7/7."),
        ("Bestätigt", "Globale Materialreferenz für Waffen, Schilde und abgeleitete Schaftmaterialien. Korrigiert: Schwarzfels 5/1, Chitin 5/2, Qualitätsstahl 2/4, Vulkanglas 7/3, Kolharz 5/7, Alchemistensilber 5/6; alle übrigen Werte unverändert."),
        ("Bestätigt", "Schildbespannung AW/NW: Stoff 1/1 ALLE; Leder 1/1 ALLE; Spinnenwebe 4/4 Draw; Kohlharz 4/4 Zentauren; Drachenschuppe M/M ALLE."),
        ("Bestätigt", "Waffenmaterialien: Schild-Herstellerregeln auf alle 19 gleichartigen Materialien übertragen; Kolhartz→Kolharz und Schwarzfells→Schwarzfels."),
        ("Bestätigt", "Waffenfertigung entspricht Schildfertigung: Ausschuss/Goblin Massenfab./Massenfabrikation/Gesellenarbeit 1/1; Meisterarbeit 2/3; Großmeisterarbeit 3/5; Einzelstück 7/7. Nur Goblin Massenfab. ist auf Goblins beschränkt."),
        ("Bestätigt", "Waffenanpassung: Von der Stange 1/1, angepasst 2/3, Perfekt angepasst 3/5; alle ALLE."),
        ("Bestätigt", "Waffenschaftmaterial: 15 Varianten übernehmen Verfügbarkeit/Hersteller des Grundmaterials; Verstärkt/Voll ohne Zusatzmodifikator; Adamandit-Schreibweise normalisiert."),
        ("Bestätigt", "17 Waffenbasis-Zeilen sind natürliche Angriffe oder Kampfstile, keine Ausrüstung: NICHT KAUFBAR; vorhandene Anwendungs-Spezies kanonisch normalisiert."),
        ("Bestätigt", "226 kaufbare Waffenbasis-Zeilen: 18 ausdrücklich kulturell benannte Modelle belegen 20 Zeilen und sind ihrer Spezies zugewiesen; übrige 206 Zeilen ALLE; Trolltöter Widerhaken bleibt ALLE."),
        ("Bestätigt", "Alle 226 kaufbaren Waffenbasis-Zeilen erhalten AW/NW 1/1; Seltenheit wird durch Komponenten, Spezies und Ort abgebildet. 17 Nicht-Gegenstände bleiben NICHT KAUFBAR."),
    ]
    ws.append([])
    ws.append(["Status", "Entscheidung / Arbeitsstand"])
    for row in rows:
        ws.append(row)
    style_header(ws, 4, 1, 2)
    for row in ws.iter_rows(min_row=5, max_row=4 + len(rows), min_col=1, max_col=2):
        row[0].font = Font(bold=True)
        row[1].alignment = Alignment(wrap_text=True, vertical="top")
    ws["D4"] = "Audit-Kennzahlen"
    ws["D4"].fill = PatternFill("solid", fgColor=NAVY)
    ws["D4"].font = Font(bold=True, color=WHITE)
    metrics = [
        ("Auditzeilen", "=COUNTA('Audit Ausrüstung'!D:D)-1"),
        ("Verfügbarkeit OFFEN", '=COUNTIF(\'Audit Ausrüstung\'!I:I,"OFFEN")'),
        ("Verfügbarkeit SONDERFALL", '=COUNTIF(\'Audit Ausrüstung\'!I:I,"SONDERFALL")'),
        ("Verfügbarkeit M", '=COUNTIF(\'Audit Ausrüstung\'!G:G,"M")'),
        ("Völkerzuweisung OFFEN", '=COUNTIF(\'Audit Ausrüstung\'!L:L,"OFFEN")'),
        ("Rüstungsdatensätze", "=COUNTA('Rüstungen Bestand'!B:B)-1"),
        ("Kaufbare Artefaktvarianten", "=COUNTA('Artefakte Bestand'!D:D)-1"),
    ]
    for index, (label, formula) in enumerate(metrics, 5):
        ws.cell(index, 4, label)
        ws.cell(index, 5, formula)
        ws.cell(index, 4).fill = PatternFill("solid", fgColor=BLUE)
        ws.cell(index, 4).font = Font(bold=True)
        ws.cell(index, 5).number_format = "0"
    ws["D13"] = "Quellen"
    ws["D13"].fill = PatternFill("solid", fgColor=NAVY)
    ws["D13"].font = Font(bold=True, color=WHITE)
    sources = [
        r"E:\Das Western Rollenspiel\_Aktuelle Daten\Nasus\Nasus Nasus das Western Rollenspiel\Grundregeln\NN_Verfügbarkeiten_1.0.docx",
        r"E:\Das Western Rollenspiel\_Aktuelle Daten\Nasus\Nasus Nasus das Western Rollenspiel\Grundregeln\Nasus Nasus Verfügbarkeiten Entwurf v0.3.docx",
        r"E:\Das Western Rollenspiel\LLM\werte 0.8-claude.xlsx (nur lesend; keine Änderung durch diese Arbeitsmappe)",
    ]
    for index, source in enumerate(sources, 14):
        ws.cell(index, 4, source)
        ws.merge_cells(start_row=index, start_column=4, end_row=index, end_column=6)
        ws.cell(index, 4).alignment = Alignment(wrap_text=True)
    set_widths(ws, {"A": 16, "B": 74, "C": 3, "D": 34, "E": 16, "F": 22})
    ws.freeze_panes = "A4"
    return ws


def create_location_model(wb):
    ws = wb.create_sheet("Ortsmodell")
    add_title(ws, "Ortsmodell", "Felder für Herkunfts- und Einkaufsorte sowie Header-Ausgabe.", 6)
    headers = ["Feld", "Typ", "Pflicht", "Mehrfach", "Regel", "Headerdruck"]
    ws.append([])
    ws.append(headers)
    rows = [
        ("id", "stabile ID", "ja", "nein", "Technischer Schlüssel", "nein"),
        ("name", "Text", "ja", "nein", "Benannter Ort; keine automatische Zusammensetzung", "ja: Ort"),
        ("welt", "AW | NW", "offen", "nein", "Kontrollierte Auswahl", "ja: AW/NW"),
        ("region", "Text", "offen", "nein", "Freitext bis eine Regionsliste existiert", "ja: Region"),
        ("siedlungsgroesse", "Auswahl", "offen", "nein", "Kontrollierte Ortsausprägung", "nein"),
        ("hauptspezies", "Völker-ID", "offen", "nein", "Erste Nennung", "nein"),
        ("etablierteMinderheiten", "Völker-ID[]", "nein", "ja", "Geordnete Mehrfachauswahl", "nein"),
        ("handelsstufe", "Auswahl", "offen", "nein", "Kontrollierte Ortsausprägung", "nein"),
        ("herstellungsort", "Auswahl", "offen", "nein", "Kontrollierte Ortsausprägung", "nein"),
        ("haendler", "Liste {typ, warengruppe?}", "nein", "ja", "Spezialisierte Händler: genau eine Pflicht-Warengruppe; je Kauf bester anwendbarer Händler", "nein"),
        ("lokaleProduktion", "Liste {warengruppe, volk?}", "nein", "ja", "Passender Eintrag ersetzt Herstellungsort durch 'direkt vor Ort'; kein Zusatzbonus", "nein"),
        ("erstelltAm", "Zeitstempel", "ja", "nein", "Lokale Persistenz", "nein"),
        ("aktualisiertAm", "Zeitstempel", "ja", "nein", "Lokale Persistenz", "nein"),
    ]
    for row in rows:
        ws.append(row)
    style_header(ws, 4, 1, len(headers))
    add_table(ws, 4, 4 + len(rows), len(headers), "OrtFelderTable")
    set_widths(ws, {"A": 28, "B": 20, "C": 12, "D": 12, "E": 62, "F": 18})
    ws.freeze_panes = "A5"
    return ws


def create_location_values(wb):
    ws = wb.create_sheet("Ortsausprägungen")
    add_title(
        ws,
        "Kontrollierte Ortsausprägungen",
        "Additive Modifikatoren bleiben absichtlich leer, bis sie fachlich bestätigt sind.",
        6,
    )
    headers = ["Kategorie", "Ausprägung", "Mod Rüstungen/Waffen", "Mod Artefakte", "Status", "Notiz"]
    ws.append([])
    ws.append(headers)
    values = {
        "Welt": ["AW", "NW"],
        "Siedlungsgröße": ["Wildnis", "Ansiedlung", "Dorf", "Großes Dorf", "Kleinstadt", "Stadt", "Großstadt", "Metropole"],
        "Handelsstufe": ["Völlig abgelegen von jeglichem Handel", "Abgelegen von jeglichem Handel", "Handelsroute / Kleiner Handels-Hafen", "Handelsstadt / Großer Handels-Hafen", "Handelszentrum"],
        "Herstellungsort": ["Import, wird nicht hergestellt", "Teilweiser Import, Herstellung im Reich", "Herstellung im Reich", "Herstellung in der Region", "Herstellung direkt vor Ort"],
        "Händler": ["Kein Laden / kein Händler", "Fahrender Trödelhändler", "Fahrender spezialisierter Händler", "Kleiner General Store", "Großer General Store", "Kleiner spezialisierter Händler", "Spezialisierter Händler", "Großer spezialisierter Händler"],
    }
    settlement_modifiers = {
        "Wildnis": (3, 5),
        "Ansiedlung": (2, 4),
        "Dorf": (1, 3),
        "Großes Dorf": (0, 2),
        "Kleinstadt": (-1, 1),
        "Stadt": (-2, 0),
        "Großstadt": (-3, -1),
        "Metropole": (-4, -2),
    }
    trade_modifiers = {
        "Völlig abgelegen von jeglichem Handel": (2, 3),
        "Abgelegen von jeglichem Handel": (1, 2),
        "Handelsroute / Kleiner Handels-Hafen": (0, 1),
        "Handelsstadt / Großer Handels-Hafen": (-1, 0),
        "Handelszentrum": (-2, -1),
    }
    manufacturing_modifiers = {
        "Import, wird nicht hergestellt": (2, 3),
        "Teilweiser Import, Herstellung im Reich": (1, 2),
        "Herstellung im Reich": (0, 1),
        "Herstellung in der Region": (-1, 0),
        "Herstellung direkt vor Ort": (-2, -1),
    }
    merchant_modifiers = {
        "Kein Laden / kein Händler": (3, 5),
        "Fahrender Trödelhändler": (2, 4),
        "Fahrender spezialisierter Händler": (1, 2),
        "Kleiner General Store": (1, 3),
        "Großer General Store": (0, 2),
        "Kleiner spezialisierter Händler": (0, 1),
        "Spezialisierter Händler": (-1, 0),
        "Großer spezialisierter Händler": (-2, -1),
    }
    for category, entries in values.items():
        for entry in entries:
            if category == "Welt":
                weapon_mod = artifact_mod = None
                status = "BESTÄTIGT"
                note = "Auswahlwert bestätigt; AW/NW wählt den Grundwert."
            elif category == "Siedlungsgröße":
                weapon_mod, artifact_mod = settlement_modifiers[entry]
                status = "BESTÄTIGT"
                note = "Bestätigte additive Siedlungsmodifikatoren."
            elif category == "Handelsstufe":
                weapon_mod, artifact_mod = trade_modifiers[entry]
                status = "BESTÄTIGT"
                note = "Bestätigte additive Handelsstufenmodifikatoren."
            elif category == "Herstellungsort":
                weapon_mod, artifact_mod = manufacturing_modifiers[entry]
                status = "BESTÄTIGT"
                note = "Bestätigte additive Herstellungsortmodifikatoren."
            elif category == "Händler":
                weapon_mod, artifact_mod = merchant_modifiers[entry]
                status = "BESTÄTIGT"
                if "spezialisiert" in entry:
                    note = "Gilt nur für die per Pflicht-Dropdown gewählte Warengruppe."
                else:
                    note = "Bestätigter additiver Händlermodifikator."
            else:
                weapon_mod = artifact_mod = None
                status = "OFFEN"
                note = "Numerische additive Werte festlegen."
            ws.append([category, entry, weapon_mod, artifact_mod, status, note])
    end = ws.max_row
    style_header(ws, 4, 1, len(headers))
    add_table(ws, 4, end, len(headers), "OrtsauspraegungenTable")
    add_list_validation(ws, STATUS_VALUES, [f"E5:E{end}"])
    add_status_formatting(ws, f"E5:E{end}", "E5")
    set_widths(ws, {"A": 22, "B": 48, "C": 24, "D": 18, "E": 18, "F": 46})
    ws.freeze_panes = "A5"
    return ws


def canonical_product_groups(preisliste):
    core = {"NK-Waffen", "Fernkampfwaffen", "Feuerwaffen", "Rüstungen", "Schilde", "Artefakte"}
    existing = {str(row.get("art") or "").strip() for row in preisliste}
    return sorted(core | {value for value in existing if value}, key=str.casefold)


def create_product_groups(wb, preisliste):
    ws = wb.create_sheet("Warengruppen")
    add_title(ws, "Zentrale Warengruppen", "43 bestätigte Gruppen; das Händler-Dropdown enthält nur die 38 spezialisierbaren Warengruppen.", 6)
    headers = ["Warengruppe", "Quelle", "Status", "Händler-spezialisierbar", "Verwendung", "Prüfnotiz"]
    ws.append([])
    ws.append(headers)
    preisliste_groups = {str(row.get("art") or "").strip() for row in preisliste}
    groups = canonical_product_groups(preisliste)
    excluded = MERCHANT_SPECIALIZATION_EXCLUDED
    for group in groups:
        source = "Preisliste: Art" if group in preisliste_groups else "Ausrüstungsmodul / Nutzerbeispiel"
        merchant_specializable = "NEIN" if group in excluded else "JA"
        use = "Item-Zuordnung; nicht im Händler-Dropdown" if group in excluded else "Item-Zuordnung und Händler-Dropdown"
        ws.append([group, source, "BESTÄTIGT", merchant_specializable, use, "Vorhandene Benennung unverändert übernehmen."])
    end = ws.max_row
    eligible_groups = [group for group in groups if group not in excluded]
    ws["G4"] = "Dropdown-Hilfsliste"
    for row_index, group in enumerate(eligible_groups, 5):
        ws.cell(row_index, 7, group)
    dropdown_end = 4 + len(eligible_groups)
    ws.column_dimensions["G"].hidden = True
    style_header(ws, 4, 1, len(headers))
    add_table(ws, 4, end, len(headers), "WarengruppenTable")
    add_list_validation(ws, STATUS_VALUES, [f"C5:C{end}"])
    add_status_formatting(ws, f"C5:C{end}", "C5")
    set_widths(ws, {"A": 34, "B": 30, "C": 18, "D": 24, "E": 48, "F": 48})
    ws.freeze_panes = "A5"
    return end, dropdown_end


def create_example_locations(wb):
    ws = wb.create_sheet("Beispielorte")
    add_title(ws, "Beispielorte", "Normalisierte Erfassung der drei vorgegebenen Herkunftsorte. Händler stehen normalisiert im Blatt Ort-Händler.", 10)
    headers = ["Name", "Welt", "Region", "Siedlungsgröße", "Hauptspezies", "Etablierte Minderheiten", "Handelsstufe", "Herstellungsort", "Händler-Zusammenfassung", "Lokale Produktion"]
    ws.append([])
    ws.append(headers)
    rows = [
        ("Straitmor", "NW", "Orkisches Protektorat Straitmor", "Metropole", "Orks", "Zwerge | Goblins", "Handelszentrum", "Teilweiser Import, Herstellung im Reich", "6 spezialisierte Händler (siehe Ort-Händler)", "Feuerwaffen: Orks | Rüstungen | NK-Waffen"),
        ("Zwogón", "AW", "Großkönigliche Kernprovinz Zwogón", "Metropole", "Zwerge", "Orks | Elfen | Gnome", "Handelszentrum", "Herstellung direkt vor Ort", "38 große spezialisierte Händler (jede zulässige Warengruppe)", ""),
        ("Phoenix-Feste", "NW", "Neuweltliches Protektorat Neu-Zwogón", "Dorf", "Zwerge", "Indianer", "Handelsroute / Kleiner Handels-Hafen", "Import, wird nicht hergestellt", "1 kleiner General Store", "NK-Waffen: Zwerge | Rüstungen"),
    ]
    for row in rows:
        ws.append(row)
    style_header(ws, 4, 1, len(headers))
    add_table(ws, 4, 4 + len(rows), len(headers), "BeispielorteTable")
    set_widths(ws, {"A": 20, "B": 10, "C": 40, "D": 18, "E": 18, "F": 30, "G": 38, "H": 42, "I": 48, "J": 46})
    for row in ws.iter_rows(min_row=5, max_row=7):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A5"
    return ws


def create_location_merchants(wb, product_group_dropdown_end, eligible_product_groups):
    ws = wb.create_sheet("Ort-Händler")
    add_title(ws, "Händler der Beispielorte", "Ein Händlerdatensatz pro Zeile. Spezialisierte Händler benötigen genau eine Warengruppe.", 5)
    headers = ["Ort", "Händler-Typ", "Warengruppe", "Status", "Prüfnotiz"]
    ws.append([])
    ws.append(headers)
    rows = [
        ("Straitmor", "Spezialisierter Händler", "Sklaven", "BESTÄTIGT", "Nutzerangabe"),
        ("Straitmor", "Spezialisierter Händler", "Feuerwaffen", "BESTÄTIGT", "Nutzerangabe"),
        ("Straitmor", "Spezialisierter Händler", "Rüstungen", "BESTÄTIGT", "Nutzerangabe"),
        ("Straitmor", "Spezialisierter Händler", "NK-Waffen", "BESTÄTIGT", "Nutzerangabe"),
        ("Straitmor", "Spezialisierter Händler", "Edelsteine", "BESTÄTIGT", "Nutzerangabe; Bergbau ausdrücklich nicht angelegt"),
        ("Straitmor", "Spezialisierter Händler", "Metall", "BESTÄTIGT", "Nutzerangabe; Bergbau ausdrücklich nicht angelegt"),
    ]
    rows.extend(
        ("Zwogón", "Großer spezialisierter Händler", group, "BESTÄTIGT", "Nutzerangabe: jede händlerspezialisierbare Warengruppe")
        for group in eligible_product_groups
    )
    rows.append(("Phoenix-Feste", "Kleiner General Store", "", "BESTÄTIGT", "Nicht spezialisiert; Warengruppe bleibt leer"))
    for row in rows:
        ws.append(row)
    end = ws.max_row
    style_header(ws, 4, 1, len(headers))
    add_table(ws, 4, end, len(headers), "OrtHaendlerTable")
    merchant_types = '"Kein Laden / kein Händler,Fahrender Trödelhändler,Fahrender spezialisierter Händler,Kleiner General Store,Großer General Store,Kleiner spezialisierter Händler,Spezialisierter Händler,Großer spezialisierter Händler"'
    add_list_validation(ws, merchant_types, [f"B5:B{end}"])
    add_list_validation(ws, f'=INDIRECT("Warengruppen!$G$5:$G${product_group_dropdown_end}")', [f"C5:C{end}"])
    add_list_validation(ws, STATUS_VALUES, [f"D5:D{end}"])
    add_status_formatting(ws, f"D5:D{end}", "D5")
    ws.conditional_formatting.add(f"C5:C{end}", FormulaRule(formula=['AND(ISNUMBER(SEARCH("spezialisiert",B5)),C5="")'], fill=PatternFill("solid", fgColor=YELLOW)))
    set_widths(ws, {"A": 22, "B": 38, "C": 28, "D": 18, "E": 58})
    ws.freeze_panes = "A5"
    return ws


def create_people_model(wb):
    ws = wb.create_sheet("Völker-Modell")
    add_title(ws, "Völkerzuweisung für Ausrüstung", "Ein Datenmodell für Alle-, Einfach- und Mehrfachzuweisung.", 6)
    ws.append([])
    ws.append(["Feld", "Zulässige Werte", "Regel", "Beispiel 1", "Beispiel 2", "Status"])
    rows = [
        ("voelkerModus", "ALLE | AUSWAHL", "ALLE nutzt keine Einzelliste; AUSWAHL verlangt mindestens ein Volk", "ALLE", "AUSWAHL", "VORSCHLAG"),
        ("voelker", "Liste von Völker-IDs", "1 Eintrag = Einfachauswahl; mehrere Einträge = Mehrfachauswahl", "", "Orks | Zwerge", "VORSCHLAG"),
        ("voelkerStatus", "BESTÄTIGT | ABGELEITET | OFFEN", "Transparente Datenpflege", "BESTÄTIGT", "OFFEN", "VORSCHLAG"),
        ("voelkerNotiz", "Text optional", "Begründung/Quelle dokumentieren", "Allgemeines Handelsgut", "Quelle prüfen", "VORSCHLAG"),
    ]
    for row in rows:
        ws.append(row)
    style_header(ws, 4, 1, 6)
    add_table(ws, 4, 4 + len(rows), 6, "VoelkerModellTable")
    ws["A11"] = "Zentrale Völkerliste"
    ws["A11"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A11"].font = Font(bold=True, color=WHITE)
    peoples = ["Dalkini", "Draw", "Elfen", "Gnome", "Goblins", "Indianer", "Katzen", "Orks", "Trolle", "Zentauren", "Zwerge"]
    for index, people in enumerate(peoples, 12):
        ws.cell(index, 1, people)
    ws["C11"] = "Wirkung"
    ws["C11"].fill = PatternFill("solid", fgColor=NAVY)
    ws["C11"].font = Font(bold=True, color=WHITE)
    ws["C12"] = "Die Zuweisung beschreibt Hersteller-/Kulturherkunft, keine Benutzungsbeschränkung."
    ws.merge_cells("C12:F14")
    ws["C12"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["C16"] = "Abgleich mit Ortsbevölkerung"
    ws["C16"].fill = PatternFill("solid", fgColor=NAVY)
    ws["C16"].font = Font(bold=True, color=WHITE)
    ws.merge_cells("C16:F16")
    population_headers = ["Völkermodus", "Beste Übereinstimmung", "Modifikator", "Regel"]
    for column, value in enumerate(population_headers, 3):
        ws.cell(17, column, value)
    style_header(ws, 17, 3, 6)
    population_rows = [
        ("ALLE", "unabhängig", 0, "Immer 0"),
        ("AUSWAHL", "Hauptspezies", 0, "Mindestens eine Zuweisung passt"),
        ("AUSWAHL", "etablierte Minderheit", 1, "Nur wenn keine Hauptspezies passt"),
        ("AUSWAHL", "keine Spezies", 3, "Keine Zuweisung ist am Ort vertreten"),
    ]
    for row_index, row in enumerate(population_rows, 18):
        for column, value in enumerate(row, 3):
            ws.cell(row_index, column, value)
    set_widths(ws, {"A": 22, "B": 38, "C": 58, "D": 22, "E": 28, "F": 16})
    ws.freeze_panes = "A5"
    return ws


def create_material_reference_sheet(wb):
    ws = wb.create_sheet("Material-Referenz")
    add_title(ws, "Globale Materialreferenz", "Verbindliche AW-/NW-Werte und Herstellerzuweisungen für das gesamte Verfügbarkeitsprojekt.", 7)
    headers = ["Material", "Verfügbarkeit AW", "Verfügbarkeit NW", "Völker-Modus", "Völker", "Status", "Prüfnotiz"]
    ws.append([])
    ws.append(headers)
    for material in GLOBAL_MATERIAL_ORDER:
        aw, nw = GLOBAL_MATERIAL_AVAILABILITY[material]
        people = global_material_people(material)
        mode, people_list, _ = normalize_people(people)
        ws.append([material, aw, nw, mode, people_list, "BESTÄTIGT", "Globale Referenz; Waffen, Schilde und abgeleitete Schaftmaterialien verwenden diese Zeile."])
    end = ws.max_row
    style_header(ws, 4, 1, len(headers))
    add_table(ws, 4, end, len(headers), "MaterialReferenzTable")
    add_availability_validation(ws, [f"B5:C{end}"])
    add_list_validation(ws, PEOPLE_MODE_VALUES, [f"D5:D{end}"])
    add_list_validation(ws, STATUS_VALUES, [f"F5:F{end}"])
    add_status_formatting(ws, f"F5:F{end}", "F5")
    set_widths(ws, {"A": 26, "B": 20, "C": 20, "D": 18, "E": 58, "F": 18, "G": 72})
    ws.freeze_panes = "A5"
    return ws


def create_armor_sheet(wb, armor):
    ws = wb.create_sheet("Rüstungen Bestand")
    add_title(ws, "Vorhandene Rüstungsverfügbarkeiten", "Vollständiger Bestand: 12 Metallbasen mit eingeschränkter Herstellerliste; übrige 16 Datensätze ALLE.", 11)
    headers = ["Gruppe", "Name", "SourceRow", "Verfügbarkeit NW", "Verfügbarkeit AW", "Differenz NW-AW", "Muster", "Status Verfügbarkeit", "Völker-Modus", "Völker", "Status Völker"]
    ws.append([])
    ws.append(headers)
    labels = {"basis": "Basis", "verarbeitung": "Verarbeitung", "anpassung": "Anpassung"}
    for key in ("basis", "verarbeitung", "anpassung"):
        for row in armor[key]:
            nw = int(row["Verfuegbarkeit-NW"])
            aw = int(row["Verfuegbarkeit-AW"])
            delta = nw - aw
            pattern = "gleich" if delta == 0 else f"NW um {delta} schlechter"
            is_chain = key == "basis" and row["name"] in CHAIN_ARMOR_NAMES
            is_plate = key == "basis" and row["name"] in PLATE_ARMOR_NAMES
            restricted = is_chain or is_plate
            people_mode = "AUSWAHL" if restricted else "ALLE"
            people = CHAIN_ARMOR_MANUFACTURERS if is_chain else PLATE_ARMOR_MANUFACTURERS if is_plate else ""
            ws.append([labels[key], row["name"], row["sourceRow"], nw, aw, delta, pattern, "BESTÄTIGT", people_mode, people, "BESTÄTIGT"])
    end = ws.max_row
    style_header(ws, 4, 1, len(headers))
    add_table(ws, 4, end, len(headers), "RuestungenBestandTable")
    add_availability_validation(ws, [f"D5:E{end}"])
    add_list_validation(ws, STATUS_VALUES, [f"H5:H{end}", f"K5:K{end}"])
    add_list_validation(ws, PEOPLE_MODE_VALUES, [f"I5:I{end}"])
    add_status_formatting(ws, f"H5:H{end}", "H5")
    add_status_formatting(ws, f"K5:K{end}", "K5")
    ws.conditional_formatting.add(
        f"F5:F{end}",
        FormulaRule(formula=["F5>0"], fill=PatternFill("solid", fgColor=YELLOW)),
    )
    set_widths(ws, {"A": 18, "B": 36, "C": 12, "D": 20, "E": 20, "F": 20, "G": 26, "H": 22, "I": 18, "J": 58, "K": 18})
    ws.freeze_panes = "A5"
    return ws


def create_artifact_sheet(wb, artifacts):
    ws = wb.create_sheet("Artefakte Bestand")
    add_title(ws, "Kaufbare Artefaktvarianten", "Verfügbarkeit nach Grad; Völkerzuweisung wird von 57 Grunddatensätzen an alle 749 Varianten vererbt.", 15)
    headers = ["Referenz", "Name", "Grad", "Variante", "Preis", "Verfügbarkeit Bestand", "Verfügbarkeit AW", "NW-Modifikator", "Verfügbarkeit NW", "Status", "SourceRow", "Prüfnotiz", "Völker-Modus", "Völker", "Status Völker"]
    ws.append([])
    ws.append(headers)
    for row in artifacts["kosten"]:
        for variant, cost_key, availability_key in (
            ("einmalig", "kostenEinmalig", "verfuegbarkeitEinmalig"),
            ("permanent", "kostenPermanent", "verfuegbarkeitPermanent"),
        ):
            price = row.get(cost_key)
            if price in (None, ""):
                continue
            availability = row.get(availability_key)
            numeric = is_availability(availability)
            master = is_master_availability(availability)
            status = "BESTÄTIGT" if numeric or master else ("SONDERFALL" if availability not in (None, "") else "OFFEN")
            note = "" if numeric else ("Bestandswert Meister wird kanonisch als M geführt; nur Meister-Modul." if master else ("Nichtnumerischer Bestandswert fachlich klären." if availability not in (None, "") else "Verfügbarkeit fehlt."))
            grade = int(row.get("grad")) if str(row.get("grad", "")).isdigit() else None
            nw_modifier = 0 if grade is not None and grade <= 2 else 1 if grade is not None and grade <= 5 else 2 if grade is not None else None
            availability_aw = int(availability) if numeric else "M" if master else None
            availability_nw = min(7, availability_aw + nw_modifier) if numeric and nw_modifier is not None else "M" if master else None
            if master:
                nw_modifier = None
            ws.append([
                row.get("referenz"), row.get("name"), row.get("grad"), variant, price,
                availability, availability_aw, nw_modifier, availability_nw, status, row.get("sourceRow"), note,
                "AUSWAHL", ARTIFACT_MANUFACTURERS, "BESTÄTIGT",
            ])
    end = ws.max_row
    style_header(ws, 4, 1, len(headers))
    add_table(ws, 4, end, len(headers), "ArtefakteBestandTable")
    add_availability_validation(ws, [f"G5:G{end}", f"I5:I{end}"])
    add_list_validation(ws, STATUS_VALUES, [f"J5:J{end}", f"O5:O{end}"])
    add_list_validation(ws, PEOPLE_MODE_VALUES, [f"M5:M{end}"])
    add_status_formatting(ws, f"J5:J{end}", "J5")
    add_status_formatting(ws, f"O5:O{end}", "O5")
    set_widths(ws, {"A": 38, "B": 34, "C": 10, "D": 14, "E": 16, "F": 24, "G": 18, "H": 18, "I": 18, "J": 18, "K": 12, "L": 48, "M": 18, "N": 52, "O": 18})
    ws.freeze_panes = "A5"
    return ws


def audit_rows(armor, artifacts, preisliste, weapons, shields):
    rows = []

    def add(family, table, row, variant="", availability="", aw="", nw="", av_status="OFFEN", people_raw="", people_status_override="", note=""):
        people_mode, people, people_status = normalize_people(people_raw)
        if people_status_override:
            people_status = people_status_override
        rows.append([
            family, table, row.get("sourceRow"), row.get("name") or row.get("referenz") or "",
            variant, availability, aw, nw, av_status, people_mode, people, people_status, note,
        ])

    for table in ("basis", "verarbeitung", "anpassung"):
        for row in armor[table]:
            is_chain = table == "basis" and row.get("name") in CHAIN_ARMOR_NAMES
            is_plate = table == "basis" and row.get("name") in PLATE_ARMOR_NAMES
            restricted = is_chain or is_plate
            people_raw = CHAIN_ARMOR_MANUFACTURERS if is_chain else PLATE_ARMOR_MANUFACTURERS if is_plate else "ALLE"
            note = "Hersteller-Auswahl bestätigt; Katzen, Indianer, Zentauren, Gnome und Trolle ausgeschlossen." if is_chain else "Hersteller-Auswahl bestätigt; Katzen, Indianer, Zentauren und Gnome ausgeschlossen." if is_plate else "Völkerzuweisung ALLE bestätigt."
            add("Rüstung", table, row, aw=row.get("Verfuegbarkeit-AW"), nw=row.get("Verfuegbarkeit-NW"), av_status="BESTÄTIGT", people_raw=people_raw, people_status_override="BESTÄTIGT", note=note)

    for row in artifacts["kosten"]:
        for variant, cost_key, availability_key in (
            ("einmalig", "kostenEinmalig", "verfuegbarkeitEinmalig"),
            ("permanent", "kostenPermanent", "verfuegbarkeitPermanent"),
        ):
            if row.get(cost_key) in (None, ""):
                continue
            availability = row.get(availability_key)
            master = is_master_availability(availability)
            status = "BESTÄTIGT" if is_availability(availability) or master else ("SONDERFALL" if availability not in (None, "") else "OFFEN")
            if is_availability(availability) and str(row.get("grad", "")).isdigit():
                grade = int(row["grad"])
                nw_modifier = 0 if grade <= 2 else 1 if grade <= 5 else 2
                availability_aw = int(availability)
                availability_nw = min(7, availability_aw + nw_modifier)
                note = f"Artefakt-NW-Modifikator Grad {grade}: +{nw_modifier}; Hersteller-Auswahl vom Grunddatensatz geerbt."
            elif master:
                availability_aw = "M"
                availability_nw = "M"
                note = "Bestandswert Meister als M normalisiert; nur Meister-Modul; Hersteller-Auswahl vom Grunddatensatz geerbt."
            else:
                availability_aw = ""
                availability_nw = ""
                note = "Sonderwert fachlich klären; Hersteller-Auswahl vom Grunddatensatz geerbt."
            add("Artefakt", "Artefakt-Kosten", row, variant=f"Grad {row.get('grad')} / {variant}", availability=availability, aw=availability_aw, nw=availability_nw, av_status=status, people_raw=ARTIFACT_MANUFACTURERS, people_status_override="BESTÄTIGT", note=note)

    for row in preisliste:
        if not row.get("preisAvailable"):
            continue
        add("Preisliste", "Preisliste", row, av_status="OFFEN", note="Kaufbar, aber Verfügbarkeit und Völkerzuweisung fehlen.")

    for table in ("basis", "material", "fertigung", "anpassung", "schaftmaterial"):
        for row in weapons[table]:
            if table == "basis" and row.get("Volk"):
                people_raw = NON_PURCHASABLE_WEAPON_PEOPLE[row.get("Volk")]
                add("Waffe", table, row, av_status="NICHT KAUFBAR", people_raw=people_raw, people_status_override="BESTÄTIGT", note="Natürlicher Angriff oder Kampfstil; aus Ausrüstungskäufen und Ortsverfügbarkeit ausgeschlossen; Spezies als Anwendungsbeschränkung beibehalten.")
            elif table == "basis":
                people_raw = CULTURAL_WEAPON_PEOPLE.get(row.get("name"), "ALLE")
                note = "Grundverfügbarkeit 1/1 bestätigt; ausdrücklich kulturell benannte Waffenbasis der genannten Spezies zugewiesen." if row.get("name") in CULTURAL_WEAPON_PEOPLE else "Grundverfügbarkeit 1/1 und Völkerzuweisung ALLE bestätigt."
                add("Waffe", table, row, aw=1, nw=1, av_status="BESTÄTIGT", people_raw=people_raw, people_status_override="BESTÄTIGT", note=note)
            elif table == "material":
                source_name = row.get("name") or ""
                display_row = dict(row)
                if source_name == "Kolhartz":
                    display_row["name"] = "Kolharz"
                    people_raw = "Zentauren"
                elif source_name == "Schwarzfells":
                    display_row["name"] = "Schwarzfels"
                    people_raw = "Indianer | Katzen"
                elif source_name == "Vulkanglas":
                    people_raw = "Indianer | Katzen"
                elif source_name == "Diamantspat":
                    people_raw = "Draw"
                elif source_name in {"Mithril", "Nasium"}:
                    people_raw = "Elfen | Zwerge"
                elif source_name == "Alchemistensilb.":
                    people_raw = "Dalkini | Elfen | Goblins | Orks | Zwerge"
                elif source_name in METAL_WEAPON_MATERIALS:
                    people_raw = METAL_SHIELD_MANUFACTURERS
                else:
                    people_raw = "ALLE"
                canonical_name = canonical_material_name(source_name)
                availability_aw, availability_nw = GLOBAL_MATERIAL_AVAILABILITY[canonical_name]
                add("Waffe", table, display_row, aw=availability_aw, nw=availability_nw, av_status="BESTÄTIGT", people_raw=people_raw, people_status_override="BESTÄTIGT", note="Verfügbarkeit und Herstellerzuweisung aus der globalen Materialreferenz übernommen.")
            elif table == "fertigung":
                availability_aw, availability_nw = SHIELD_CRAFT_AVAILABILITY[row.get("name")]
                people_raw = "Goblins" if row.get("name") == "Goblin Massenfab." else "ALLE"
                add("Waffe", table, row, aw=availability_aw, nw=availability_nw, av_status="BESTÄTIGT", people_raw=people_raw, people_status_override="BESTÄTIGT", note="Verfügbarkeit und Herstellerzuweisung identisch zur bestätigten Schildfertigung.")
            elif table == "anpassung":
                availability_aw, availability_nw = WEAPON_ADJUSTMENT_AVAILABILITY[row.get("name")]
                add("Waffe", table, row, aw=availability_aw, nw=availability_nw, av_status="BESTÄTIGT", people_raw="ALLE", people_status_override="BESTÄTIGT", note="Verfügbarkeit analog zur bestätigten Rüstungsanpassung; Völkerzuweisung ALLE.")
            elif table == "schaftmaterial":
                source_name = row.get("name") or ""
                display_row = dict(row)
                if source_name == "Adamandit Verst.":
                    display_row["name"] = "Adamandit Verstärkt"
                elif source_name == "Adamantit Voll":
                    display_row["name"] = "Adamandit Voll"
                base_material = WEAPON_SHAFT_BASE_MATERIAL.get(source_name)
                availability_aw, availability_nw = (1, 1) if base_material is None else GLOBAL_MATERIAL_AVAILABILITY[base_material]
                people_raw = "ALLE" if base_material is None else global_material_people(base_material)
                add("Waffe", table, display_row, aw=availability_aw, nw=availability_nw, av_status="BESTÄTIGT", people_raw=people_raw, people_status_override="BESTÄTIGT", note="Verfügbarkeit und Herstellerzuweisung vom bestätigten Grundmaterial übernommen; Verstärkt/Voll ohne Zusatzmodifikator.")
            else:
                add("Waffe", table, row, av_status="OFFEN", people_raw=row.get("Volk"), note="Verfügbarkeit fehlt; vorhandene Volk-Angabe nur als abgeleiteter Prüfwert übernommen." if row.get("Volk") else "Verfügbarkeit und Völkerzuweisung fehlen.")

    for table in ("material", "fertigung", "bespannung"):
        for row in shields[table]:
            source_name = row.get("name") or ""
            display_row = dict(row)
            people_status = "BESTÄTIGT"
            if table == "material" and source_name == "Kolhartz":
                display_row["name"] = "Kolharz"
                people_raw = "Zentauren"
                note = "Verfügbarkeit fehlt; Schreibweise Kolhartz zu Kolharz korrigiert; vorhandener Quellenverweis: nur Zentauren."
            elif table == "material" and source_name == "Diamantspart":
                display_row["name"] = "Diamantspat"
                people_raw = "Draw"
                note = "Verfügbarkeit fehlt; Schreibweise Diamantspart zu Diamantspat korrigiert; Hersteller Draw bestätigt."
            elif table == "material" and source_name == "Schwarzfells":
                display_row["name"] = "Schwarzfels"
                people_raw = "Indianer | Katzen"
                note = "Verfügbarkeit fehlt; Schreibweise Schwarzfells zu Schwarzfels korrigiert; Hersteller Indianer und Katzen bestätigt."
            elif table == "material" and source_name == "Vulkanglas":
                people_raw = "Indianer | Katzen"
                note = "Verfügbarkeit fehlt; Hersteller Indianer und Katzen bestätigt."
            elif table == "material" and source_name in {"Mithril", "Nasium"}:
                people_raw = "Elfen | Zwerge"
                note = "Verfügbarkeit und Hersteller Elfen/Zwerge bestätigt."
            elif table == "material" and source_name == "Spinnenwebe":
                people_raw = "Draw"
                note = "Verfügbarkeit und Hersteller Draw bestätigt."
            elif table == "bespannung" and source_name == "Spinnenwebe":
                people_raw = "Draw"
                note = "Verfügbarkeit und Hersteller Draw bestätigt."
            elif table == "material" and source_name == "Alchemistensilb.":
                people_raw = "Dalkini | Elfen | Goblins | Orks | Zwerge"
                note = "Verfügbarkeit und explizite Herstellerliste bestätigt."
            elif table == "bespannung" and source_name == "Kohlharz":
                people_raw = "Zentauren"
                note = "Verfügbarkeit fehlt; vorhandener Quellenverweis: nur Zentauren."
            elif table == "fertigung" and source_name == "Goblin Massenfab.":
                people_raw = "Goblins"
                note = "Verfügbarkeit fehlt; Spezies aus der vorhandenen Benennung übernommen."
            elif table == "material" and source_name in METAL_SHIELD_MATERIALS:
                people_raw = METAL_SHIELD_MANUFACTURERS
                note = "Verfügbarkeit fehlt; Metallschilde werden nicht von Katzen, Indianern, Zentauren oder Gnomen hergestellt."
            else:
                people_raw = "ALLE"
                note = "Verfügbarkeit fehlt; Völkerzuweisung ALLE bestätigt."
            if table == "fertigung" and source_name in SHIELD_CRAFT_AVAILABILITY:
                availability_aw, availability_nw = SHIELD_CRAFT_AVAILABILITY[source_name]
                availability_status = "BESTÄTIGT"
                note = note.replace("Verfügbarkeit fehlt; ", "Verfügbarkeit analog zur bestätigten Rüstungsfertigung; ")
            elif table == "material" and canonical_material_name(source_name) in GLOBAL_MATERIAL_AVAILABILITY:
                availability_aw, availability_nw = GLOBAL_MATERIAL_AVAILABILITY[canonical_material_name(source_name)]
                availability_status = "BESTÄTIGT"
                note = note.replace("Verfügbarkeit fehlt; ", "Verfügbarkeit bestätigt; ")
            elif table == "bespannung" and source_name in SHIELD_COVER_AVAILABILITY:
                availability_aw, availability_nw = SHIELD_COVER_AVAILABILITY[source_name]
                availability_status = "BESTÄTIGT"
                note = note.replace("Verfügbarkeit fehlt; ", "Verfügbarkeit bestätigt; ")
            else:
                availability_aw = availability_nw = ""
                availability_status = "OFFEN"
            add("Schild", table, display_row, aw=availability_aw, nw=availability_nw, av_status=availability_status, people_raw=people_raw, people_status_override=people_status, note=note)

    return rows


def create_audit_sheet(wb, rows):
    ws = wb.create_sheet("Audit Ausrüstung")
    add_title(
        ws,
        "Ausrüstungs-Audit",
        "Filterbare Arbeitsliste. Leere Verfügbarkeiten und Status OFFEN sind bewusste Nacharbeitsmarkierungen; kein Wert 0.",
        13,
    )
    headers = [
        "Familie", "Tabelle", "SourceRow", "Name/Referenz", "Variante", "Verfügbarkeit Bestand",
        "Verfügbarkeit AW", "Verfügbarkeit NW", "Status Verfügbarkeit", "Völker-Modus",
        "Völker", "Status Völker", "Prüfnotiz",
    ]
    ws.append([])
    ws.append(headers)
    for row in rows:
        ws.append(row)
    end = ws.max_row
    style_header(ws, 4, 1, len(headers))
    add_table(ws, 4, end, len(headers), "AusruestungAuditTable")
    add_availability_validation(ws, [f"F5:H{end}"])
    add_list_validation(ws, STATUS_VALUES, [f"I5:I{end}", f"L5:L{end}"])
    add_list_validation(ws, PEOPLE_MODE_VALUES, [f"J5:J{end}"])
    add_status_formatting(ws, f"I5:I{end}", "I5")
    add_status_formatting(ws, f"L5:L{end}", "L5")
    set_widths(ws, {"A": 14, "B": 20, "C": 11, "D": 42, "E": 22, "F": 24, "G": 19, "H": 19, "I": 23, "J": 17, "K": 30, "L": 18, "M": 58})
    ws.freeze_panes = "D5"
    return ws


def create_open_questions(wb):
    ws = wb.create_sheet("Offene Entscheidungen")
    add_title(ws, "Offene fachliche Entscheidungen", "Vor einer Umsetzung in Werte beziehungsweise App punktweise abnehmen.", 6)
    headers = ["Nr.", "Thema", "Entscheidung benötigt", "Empfehlung", "Status", "Antwort"]
    ws.append([])
    ws.append(headers)
    questions = [
        (1, "Händlermodifikatoren", "Additive Werte, Spezialisierung und mehrere Händler", "Spezialisierte Händler an genau eine Warengruppe binden; besten anwendbaren Händler nutzen", "GEKLÄRT", "Mehrere Händler je Ort; niedrigster anwendbarer Modifikator; ohne Treffer 'Kein Laden'"),
        (2, "Lokale Produktion", "Override-Regel je Warengruppe und optionalem Volk", "Allgemeinen Herstellungsort ersetzen; keinen Bonus addieren", "GEKLÄRT", "Bei Treffer gilt 'Herstellung direkt vor Ort'; ALLE erfüllt keinen volksgebundenen Eintrag"),
        (3, "Ortsbevölkerung", "Abgleich mit Völkerzuweisung des Gegenstands", "Beste Übereinstimmung bei Mehrfachzuweisung verwenden", "GEKLÄRT", "ALLE/Hauptspezies 0; Minderheit +1; keine Spezies +3"),
        (4, "Kaufsperre", "Welche Werte sperren Spielercharaktere?", "Zentrale Prüfung nach Modifikatoren und Begrenzung", "GEKLÄRT", "1–4 kaufbar; 5–7 gesperrt; M nur Meister-Modul; OFFEN gesperrt mit Pflegehinweis"),
        (5, "Artefakt Meister", "Bedeutung des permanenten Werts Meister bei 49 Varianten", "Kanonisch als M führen", "GEKLÄRT", "M; ausschließlich im Meister-Modul verfügbar"),
        (6, "Komposition", "Wie werden Basis und Komponenten zusammengeführt?", "Jede gewählte Pflichtkomponente vollständig separat berechnen", "GEKLÄRT", "Numerisch Maximum; M vor numerisch; OFFEN vor allem; nicht addieren/mitteln; Ursachen anzeigen"),
        (7, "AW/NW Artefakte", "Wie wird der vorhandene Artefaktwert für NW verändert?", "Vorhandener Wert als AW/Grundwert verwenden", "GEKLÄRT", "Grad 1–2 +0; Grad 3–5 +1; Grad 6–7 +2; maximal 7"),
        (8, "Master-Änderungen", "Globaler Ort oder Abenteuer-Snapshot?", "Persistenten Grundzustand plus optionalen Abenteuerzustand führen", "GEKLÄRT", "Aktiver Zustand gilt für Käufe; danach verwerfen, weiterführen oder bewusst als Grundzustand übernehmen; keine Rückwirkung"),
        (9, "Warengruppen-Normalisierung", "Müssen Preislistenarten zusammengefasst oder umbenannt werden?", "Vorhandene Namen erhalten; Händler-Eignung separat markieren", "GEKLÄRT", "43 unverändert; 38 spezialisierbar; Miete/Post/Reisekosten/Tavernen-Preise/Zoll ausgeschlossen; Dropdown durchsuchbar"),
        (10, "Meisterfreigabe", "Darf eine gesperrte Verfügbarkeit konkret freigegeben werden?", "Transaktionsbezogen protokollieren; Stammdaten unverändert lassen", "GEKLÄRT", "5–7/M freigebbar; Charakter, Gegenstand, Ort, Wert, Zeitpunkt; Begründung optional; OFFEN nicht freigebbar"),
        (11, "Rüstungen Völker", "Völkerzuweisung für 20 Basen, 4 Verarbeitungen, 4 Anpassungen", "Herstellerregeln nach Bauart anwenden", "GEKLÄRT", "6 Kettenbasen ohne Katzen/Indianer/Zentauren/Gnome/Trolle; 6 Metallplattenbasen ohne Katzen/Indianer/Zentauren/Gnome; übrige 16 ALLE"),
        (12, "Artefakte Völker", "Völkerzuweisung für 57 Grunddatensätze und 749 Varianten", "Am Grunddatensatz pflegen und an Varianten vererben", "GEKLÄRT", "AUSWAHL Draw/Elfen/Gnome/Goblins/Orks/Zwerge; alle Varianten BESTÄTIGT"),
        (13, "Schilde Völker", "Völkerzuweisung und Schreibweisen für 22 Materialien, 7 Fertigungen und 5 Bespannungen", "Querverweise und Bauartregel einheitlich anwenden", "GEKLÄRT", "Kolharz/Kohlharz Zentauren; Diamantspat und Spinnenwebe Draw; Schwarzfels/Vulkanglas Indianer und Katzen; Mithril/Nasium Elfen und Zwerge; Alchemistensilber Dalkini/Elfen/Goblins/Orks/Zwerge; Goblin Massenfab. Goblins; übrige Metalle ohne Katzen/Indianer/Zentauren/Gnome"),
        (14, "Schildfertigung Verfügbarkeit", "AW-/NW-Werte der sieben Fertigungsstufen", "Bestätigte Rüstungsfertigung übernehmen; drei niedrigere Massen-/Ausschussstufen 1/1", "GEKLÄRT", "Ausschuss, Goblin Massenfab., Massenfabrikation, Gesellenarbeit 1/1; Meisterarbeit 2/3; Großmeisterarbeit 3/5; Einzelstück 7/7"),
        (15, "Globale Materialreferenz", "AW-/NW-Werte und Völkerzuweisung der 22 Materialien", "Eine einzige Referenz für Waffen, Schilde und abgeleitete Komponenten", "GEKLÄRT", "Korrekturen: Schwarzfels 5/1; Chitin 5/2; Qualitätsstahl 2/4; Vulkanglas 7/3; Kolharz 5/7; Alchemistensilber 5/6. Alle übrigen Werte unverändert"),
        (16, "Schildbespannung", "Verfügbarkeit und Völkerzuweisung der fünf Bespannungen", "Gemeinsam als vollständigen Bespannungsblock abnehmen", "GEKLÄRT", "Stoff/Leder 1/1 ALLE; Spinnenwebe 4/4 Draw; Kohlharz 4/4 Zentauren; Drachenschuppe M/M ALLE"),
        (17, "Waffenmaterial Völker", "Völkerzuweisung der 19 Waffenmaterialien", "Bestätigte Schildmaterial-Regeln auf gleichnamige Waffenmaterialien übertragen", "GEKLÄRT", "Mithril/Nasium Elfen und Zwerge; Alchemistensilber Dalkini/Elfen/Goblins/Orks/Zwerge; Diamantspat Draw; Schwarzfels/Vulkanglas Indianer und Katzen; Kolharz Zentauren; übrige Metalle allgemeine Metallliste; Nichtmetalle ALLE"),
        (18, "Waffenfertigung", "Verfügbarkeit und Völkerzuweisung der sieben Fertigungen", "Bestätigte Schildfertigung vollständig übernehmen", "GEKLÄRT", "Ausschuss/Goblin Massenfab./Massenfabrikation/Gesellenarbeit 1/1; Meisterarbeit 2/3; Großmeisterarbeit 3/5; Einzelstück 7/7; nur Goblin Massenfab. Goblins, sonst ALLE"),
        (19, "Waffenanpassung", "Verfügbarkeit und Völkerzuweisung der drei Anpassungen", "Bestätigte Rüstungsanpassung übernehmen", "GEKLÄRT", "Von der Stange 1/1; angepasst 2/3; Perfekt angepasst 3/5; alle ALLE"),
        (20, "Waffenschaftmaterial", "Verfügbarkeit und Völkerzuweisung der 15 Schaftvarianten", "Globale Grundmaterialwerte übernehmen; Verstärkt/Voll ohne Zusatz", "GEKLÄRT", "Standard 1/1 ALLE; Eisen/Bronze 1/1; Stahl 1/2; Qualitätsstahl 2/4; Faltstahl 3/5; Mithril 7/7 Elfen/Zwerge; Adamandit M/M; übrige Metalle allgemeine Metallliste; Adamandit-Schreibweise normalisiert"),
        (21, "Nicht kaufbare Waffenbasis", "Natürliche Angriffe und Kampfstile im Waffenbestand", "Nicht als Ausrüstung behandeln; Speziesbeschränkung behalten", "GEKLÄRT", "17 Zeilen NICHT KAUFBAR; aus Kauf/Ortsverfügbarkeit ausgeschlossen; Singular-, Drow- und andere-Voelker-Angaben kanonisch normalisiert"),
        (22, "Waffenbasis Völker", "Völkerzuweisung der 226 kaufbaren Waffenbasis-Zeilen", "Nur ausdrückliche kulturelle Benennungen einschränken; Rest ALLE", "GEKLÄRT", "18 kulturell benannte Modelle auf 20 Zeilen der jeweiligen Spezies; übrige 206 Zeilen ALLE; Trolltöter Widerhaken ausdrücklich ALLE"),
        (23, "Waffenbasis Verfügbarkeit", "Grundverfügbarkeit der 226 kaufbaren Waffenbasis-Zeilen", "Basisform 1/1; Seltenheit über Komponenten, Spezies und Ort", "GEKLÄRT", "Alle 226 kaufbaren Waffenbasis-Zeilen AW/NW 1/1; 17 natürliche Angriffe/Kampfstile bleiben NICHT KAUFBAR"),
    ]
    for row in questions:
        ws.append(row)
    end = ws.max_row
    style_header(ws, 4, 1, len(headers))
    add_table(ws, 4, end, len(headers), "OffeneEntscheidungenTable")
    add_list_validation(ws, '"OFFEN,GEKLÄRT,ZURÜCKGESTELLT"', [f"E5:E{end}"])
    ws.conditional_formatting.add(f"E5:E{end}", FormulaRule(formula=['E5="OFFEN"'], fill=PatternFill("solid", fgColor=YELLOW)))
    ws.conditional_formatting.add(f"E5:E{end}", FormulaRule(formula=['E5="GEKLÄRT"'], fill=PatternFill("solid", fgColor=GREEN)))
    set_widths(ws, {"A": 8, "B": 24, "C": 60, "D": 58, "E": 18, "F": 58})
    for row in ws.iter_rows(min_row=5, max_row=end):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A5"
    return ws


def create_claude_sheet(wb):
    ws = wb.create_sheet("Claude-Auftrag")
    add_title(ws, "Claude-Implementierungsauftrag", "Erst nach fachlicher Abnahme aller offenen Regelwerte ausführen.", 4)
    headers = ["Schritt", "Arbeitsauftrag", "Abnahmekriterium", "Status"]
    ws.append([])
    ws.append(headers)
    rows = [
        (1, "Ortsmodell und kontrollierte Auswahllisten implementieren", "Typen, Validierung und drei Beispielorte vorhanden", "BLOCKIERT"),
        (2, "Heimat durch Herkunft ersetzen und bestehende Charakterdaten migrieren", "Headerdruck exakt Ort, Region, AW/NW", "BLOCKIERT"),
        (3, "Herkunftsdropdown und Formular in Charaktererschaffung", "Vorhandener oder neuer Ort wählbar; Name Pflicht", "BLOCKIERT"),
        (4, "Lokale persistente Speicherung für Orte, Abenteuer-Ortszustände und Charaktere", "Reload erhält Grundzustand, aktive/abgeschlossene Abenteuerzustände und Charaktere; Storage hinter Schnittstelle", "BLOCKIERT"),
        (5, "Meister-Modul zur Ortsverwaltung", "Grundzustand direkt editierbar; Abenteuerzustand erzeugen, verwerfen, weiterführen oder bewusst übernehmen", "BLOCKIERT"),
        (6, "Verfügbarkeit und Völkerzuweisung aller kaufbaren Daten", "Keine stillen Null-/0-Fallbacks; offene Daten sichtbar", "BLOCKIERT"),
        (7, "Zentrale Verfügbarkeitsberechnung für alle Kaufpfade", "Waffen, Rüstung, Schilde, Preisliste und Artefakte nutzen dieselbe Prüfung", "BLOCKIERT"),
        (8, "Artefakt-Verfügbarkeit sperren", "Einmalig/permanent und Grad werden korrekt geprüft", "BLOCKIERT"),
        (9, "Tests und Migration", "Berechnung, Kaufpfade, Persistenz und Bestandsmigration getestet", "BLOCKIERT"),
    ]
    for row in rows:
        ws.append(row)
    end = ws.max_row
    style_header(ws, 4, 1, len(headers))
    add_table(ws, 4, end, len(headers), "ClaudeAuftragTable")
    add_list_validation(ws, '"BLOCKIERT,BEREIT,IN ARBEIT,FERTIG"', [f"D5:D{end}"])
    set_widths(ws, {"A": 10, "B": 68, "C": 64, "D": 16})
    for row in ws.iter_rows(min_row=5, max_row=end):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A5"
    return ws


def finalize_styles(wb):
    for ws in wb.worksheets:
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.page_margins.left = 0.3
        ws.page_margins.right = 0.3
        ws.page_margins.top = 0.5
        ws.page_margins.bottom = 0.5
        for row in ws.iter_rows():
            for cell in row:
                if cell.row > 2:
                    cell.font = cell.font.copy(name="Aptos", size=10)
                    if cell.alignment == Alignment():
                        cell.alignment = Alignment(vertical="top")
        ws.auto_filter.ref = None


def main():
    armor = read_json("armor.json")
    artifacts = read_json("artefakte.json")
    preisliste = read_json("preisliste.json")
    weapons = read_json("weapons.json")
    shields = read_json("shields.json")

    wb = Workbook()
    wb.remove(wb.active)
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"

    create_overview(wb)
    create_location_model(wb)
    create_location_values(wb)
    _, product_group_dropdown_end = create_product_groups(wb, preisliste)
    create_example_locations(wb)
    eligible_product_groups = [
        group for group in canonical_product_groups(preisliste)
        if group not in MERCHANT_SPECIALIZATION_EXCLUDED
    ]
    create_location_merchants(wb, product_group_dropdown_end, eligible_product_groups)
    create_people_model(wb)
    create_material_reference_sheet(wb)
    create_armor_sheet(wb, armor)
    create_artifact_sheet(wb, artifacts)
    rows = audit_rows(armor, artifacts, preisliste, weapons, shields)
    create_audit_sheet(wb, rows)
    create_open_questions(wb)
    create_claude_sheet(wb)
    finalize_styles(wb)
    wb.save(OUTPUT)

    check = load_workbook(OUTPUT, read_only=True, data_only=False)
    print(f"output={OUTPUT}")
    print(f"sheets={len(check.sheetnames)}")
    print(f"audit_rows={len(rows)}")
    print(f"armor_rows={sum(len(armor[k]) for k in ('basis','verarbeitung','anpassung'))}")
    print(f"artifact_buyable_variants={sum(1 for row in artifacts['kosten'] for key in ('kostenEinmalig','kostenPermanent') if row.get(key) not in (None,''))}")


if __name__ == "__main__":
    main()
