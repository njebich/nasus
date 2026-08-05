"""One-off: turn the single "Gute Wunder" Talent (added by add_geweihte_stufen.py) into a
2-stufe pair (talente_geweihte_gute_wunder_stufe_1/2), mirroring talente_spruchgute_stufe_1/2 -
Stufe 1 = Karma, Stufe 2 = Karma + Aura, both capped at Normale:2. Naming with "_stufe_<N>"
gets the generic Vorstufenketten-Hardblock from engine/talenteStufenKette.ts for free (Stufe 2
requires Stufe 1). See Nutzer-Ask 2026-08-06 (project_geweihte_tab memory).

Run once, then recalc via Excel COM before re-running generate_data_ts.py.

Usage: python scripts/add_gute_wunder_stufe2.py "werte 0.8-claude.xlsx"
"""
import sys
import openpyxl

WERTE_HEADERS = [
    'Referenz', 'Kategorie', 'Beschreibung', 'Abkürzung', 'Info', 'Parent', 'Art',
    'Formel', 'Pool', 'Flag', 'Grad', 'Kosten', 'Verfuegbarkeit', None,
    'Mindest-TaW', 'Eig-Bonus', 'Wirkung',
]


def col_index(header_name):
    return WERTE_HEADERS.index(header_name) + 1


def main():
    path = sys.argv[1] if len(sys.argv) > 1 else 'werte 0.8-claude.xlsx'
    wb = openpyxl.load_workbook(path, read_only=False, data_only=False)
    ws = wb['Werte']

    renamed_row = None
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        if row[col_index('Referenz') - 1].value == 'talente_geweihte_gute_wunder':
            row[col_index('Referenz') - 1].value = 'talente_geweihte_gute_wunder_stufe_1'
            row[col_index('Beschreibung') - 1].value = 'Gute Wunder Stufe 1'
            row[col_index('Wirkung') - 1].value = (
                'Ersetzt die gute Wunder-Probe durch Karma. Die höchstmögliche Gute ist Normale : 2.'
            )
            renamed_row = row[0].row
            break
    if renamed_row is None:
        raise SystemExit('talente_geweihte_gute_wunder nicht gefunden - schon migriert?')

    last_referenz_row = 1
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        if row[0].value is not None:
            last_referenz_row = row[0].row
    next_row = last_referenz_row + 1

    stufe2 = {
        'Referenz': 'talente_geweihte_gute_wunder_stufe_2',
        'Kategorie': 'Talente',
        'Beschreibung': 'Gute Wunder Stufe 2',
        'Parent': 'Geweihte',
        'Art': 'Auswahl',
        'Kosten': 20,
        'Wirkung': 'Ersetzt die gute Wunder-Probe durch Karma + Aura. Die höchstmögliche Gute ist Normale : 2.',
    }
    for header, value in stufe2.items():
        ws.cell(row=next_row, column=col_index(header), value=value)

    wb.save(path)
    print(f'Renamed row {renamed_row}: talente_geweihte_gute_wunder -> talente_geweihte_gute_wunder_stufe_1')
    print(f'Added row {next_row}: talente_geweihte_gute_wunder_stufe_2')


if __name__ == '__main__':
    main()
